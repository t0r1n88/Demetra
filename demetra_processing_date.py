"""
Функция для подсчета текущего возраста и разбиения по возрастным категориям
"""
from demetra_support_functions import write_group_df_to_excel
import pandas as pd
import numpy as np
import datetime
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time
import platform
import warnings
import re

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging

logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)


def extract_number_month(cell):
    """
    Функция для извлечения номера месяца
    """
    return cell.month


def extract_name_month(cell):
    """
    Функция для извлечения названия месяца
    Взято отсюда https://ru.stackoverflow.com/questions/1045154/Вывод-русских-символов-из-pd-timestamp-month-name
    """
    if name_os == 'Windows':
        return cell.month_name(locale='Russian')
    else:
        return cell.month_name()


def extract_year(cell):
    """
    Функция для извлечения года рождения
    """
    return cell.year


def calculate_age(born, raw_selected_date):
    """
    Функция для расчета текущего возраста взято с https://stackoverflow.com/questions/2217488/age-from-birthdate-in-python/9754466#9754466
    :param born: дата рождения
    :return: возраст
    """

    try:
        selected_date = pd.to_datetime(raw_selected_date, dayfirst=True)
        return selected_date.year - born.year - ((selected_date.month, selected_date.day) < (born.month, born.day))

    except ValueError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Введена некорректная дата относительно которой нужно провести обработку\nПример корректной даты 01.09.2022')
        logging.exception('AN ERROR HAS OCCURRED')
        quit()


def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        if cell is np.nan:
            return 'Не заполнено'
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return f'Неправильное значение - {cell}'
    except TypeError:
        return f'Неправильное значение - {cell}'


def calculation_maturity(value):
    """
    Присвоение признака совершеннолетия
    :param value: значение
    :return: совершеннолетний если >=18, несовершеннолетний если от нуля до 18, отрицательный возраст если Текущий_возраст
     меньше нуля
    """
    try:
        int_value = int(value)
        if int_value >= 18:
            return 'совершеннолетний'
        elif 0 <= int_value < 18:
            return 'несовершеннолетний'
        else:
            return 'отрицательный возраст'
    except:
        return None

def comparison_date(value, pattern):
    """
    Функция для проверки соответсвия формата даты
    :param value:значение
    :param pattern: объект re.compile
    :return:
    """
    value = str(value)
    if re.fullmatch(pattern,value):
        return value
    else:
        return f'Ошибка: Дата должна иметь формат ДД.ММ.ГГГГ, например 21.10.2024. В ячейке указано - {value}'


def proccessing_date(raw_selected_date, name_column, df: pd.DataFrame, path_to_end_folder_date: str):
    """
   Функция для разбиения по категориям 1-ПК 1-ПО СПО-1, подсчета текущего возраста и выделения месяца,года
    :param raw_selected_date: дата на момент которой нужно подсчитать Текущий_возраст в формате DD.MM.YYYY
    :param name_column: название колонки с датами рождения
    :param name_file_data_date: датафрейм с данными
    :param path_to_end_folder_date: папка куда будет сохранен итоговый файл
    :return: файл Excel  содержащий исходный файл с добавленными колонками категорий и т.п.
    """

    try:
        global name_os  # делаем глобальной, чтобы проверять месяц
        name_os = platform.system()

        lst_create_date_columns =['Текущий_возраст','Совершеннолетие','Порядковый_номер_месяца_рождения','Название_месяца_рождения',
                                  'Год_рождения','Один_ПК','Один_ПО','СПО_Один','Росстат']
        date_pattern = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')  # создаем паттерн
        # создаем временную колонку которой в конце заменим исходную колонку
        df['temp'] = pd.to_datetime(df[name_column], dayfirst=True, errors='ignore')
        df['temp'] = df['temp'].fillna('Пустая ячейка')
        df['temp'] = df['temp'].apply(lambda x:comparison_date(x,date_pattern))


        # В случае ошибок заменяем значение NaN
        df[name_column] = pd.to_datetime(df[name_column], dayfirst=True, errors='coerce')

        # Создаем файл excel
        wb = openpyxl.Workbook()
        # Создаем листы
        wb.create_sheet(title='Свод сов_несов', index=1)
        wb.create_sheet(title='Свод по возрастам', index=2)
        wb.create_sheet(title='Свод по месяцам', index=3)
        wb.create_sheet(title='Свод по годам', index=4)
        wb.create_sheet(title='Свод по 1-ПК', index=5)
        wb.create_sheet(title='Свод по 1-ПО', index=6)
        wb.create_sheet(title='Свод по СПО-1', index=7)
        wb.create_sheet(title='Свод по категориям Росстата', index=8)

        # Подсчитываем Текущий_возраст
        df['Текущий_возраст'] = df[name_column].apply(lambda x: calculate_age(x, raw_selected_date))


        # Добавлем признак совершеннолетия
        df['Совершеннолетие'] = df['Текущий_возраст'].apply(calculation_maturity)

        # Получаем номер месяца
        df['Порядковый_номер_месяца_рождения'] = df[name_column].apply(extract_number_month)

        # Получаем название месяца
        df['Название_месяца_рождения'] = df[name_column].apply(extract_name_month)
        dct_month = {'January': 'Январь', 'February': 'Февраль', 'March': 'Март', 'April': 'Апрель', 'May': 'Май',
                     'June': 'Июнь', 'July': 'Июль',
                     'August': 'Август', 'September': 'Сентябрь', 'October': 'Октябрь',
                     'November': 'Ноябрь', 'December': 'Декабрь'}
        df['Название_месяца_рождения'] = df['Название_месяца_рождения'].replace(dct_month)

        # Получаем год рождения
        df['Год_рождения'] = df[name_column].apply(extract_year)

        # Присваиваем категорию по 1-ПК
        df['Один_ПК'] = pd.cut(df['Текущий_возраст'], [0, 24, 29, 34, 39, 44, 49, 54, 59, 64, 101, 10000],
                                      labels=['моложе 25 лет', '25-29', '30-34', '35-39',
                                              '40-44', '45-49', '50-54', '55-59', '60-64',
                                              '65 и более',
                                              'Возраст  больше 101'])
        # Приводим к строковому виду, иначе не запишется на лист
        df['Один_ПК'] = df['Один_ПК'].astype(str)
        df['Один_ПК'] = df['Один_ПК'].replace('nan', 'Ошибочное значение!!!')

        # Присваиваем категорию по 1-ПО
        df['Один_ПО'] = pd.cut(df['Текущий_возраст'],
                                      [0, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25,
                                       26, 27, 28,
                                       29, 34, 39, 44, 49, 54, 59, 64, 101],
                                      labels=['моложе 14 лет', '14 лет', '15 лет',
                                              '16 лет',
                                              '17 лет', '18 лет', '19 лет', '20 лет',
                                              '21 год', '22 года',
                                              '23 года', '24 года', '25 лет',
                                              '26 лет', '27 лет', '28 лет', '29 лет',
                                              '30-34 лет',
                                              '35-39 лет', '40-44 лет', '45-49 лет',
                                              '50-54 лет',
                                              '55-59 лет',
                                              '60-64 лет',
                                              '65 лет и старше'])
        # Приводим к строковому виду, иначе не запишется на лист
        df['Один_ПО'] = df['Один_ПО'].astype(str)
        df['Один_ПО'] = df['Один_ПО'].replace('nan', 'Ошибочное значение!!!')

        # Присваиваем категорию по 1-СПО
        df['СПО_Один'] = pd.cut(df['Текущий_возраст'],
                                       [0, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 34,
                                        39,
                                        101],
                                       labels=['моложе 13 лет', '13 лет', '14 лет', '15 лет', '16 лет', '17 лет',
                                               '18 лет',
                                               '19 лет', '20 лет'
                                           , '21 год', '22 года', '23 года', '24 года', '25 лет', '26 лет', '27 лет',
                                               '28 лет',
                                               '29 лет',
                                               '30-34 лет', '35-39 лет', '40 лет и старше'])
        ## Приводим к строковому виду, иначе не запишется на лист
        df['СПО_Один'] = df['СПО_Один'].astype(str)
        df['СПО_Один'] = df['СПО_Один'].replace('nan', 'Ошибочное значение!!!')

        # Присваиваем категорию по Росстату
        df['Росстат'] = pd.cut(df['Текущий_возраст'],
                                         [0, 4, 9, 14, 19, 24, 29, 34, 39, 44, 49, 54, 59, 64, 69, 200],
                                         labels=['0-4', '5-9', '10-14', '15-19', '20-24', '25-29', '30-34',
                                                 '35-39', '40-44', '45-49', '50-54', '55-59', '60-64', '65-69',
                                                 '70 лет и старше'])
        ## Приводим к строковому виду, иначе не запишется на лист
        df['Росстат'] = df['Росстат'].astype(str)
        df['Росстат'] = df['Росстат'].replace('nan', 'Ошибочное значение!!!')

        # Заполняем пустые строки
        df[lst_create_date_columns]=df[lst_create_date_columns].fillna('Ошибочное значение!!!')


        # заполняем сводные таблицы
        # Количество совершенолетних
        df_svod_by_matur = df.groupby(['Совершеннолетие']).agg({'ФИО': 'count'})
        df_svod_by_matur.columns = ['Количество']
        for r in dataframe_to_rows(df_svod_by_matur, index=True, header=True):
            wb['Свод сов_несов'].append(r)

        # Сводная по возрастам

        df_svod_by_age = df.groupby(['Текущий_возраст']).agg({'ФИО': 'count'})
        df_svod_by_age.columns = ['Количество']
        for r in dataframe_to_rows(df_svod_by_age, index=True, header=True):
            wb['Свод по возрастам'].append(r)

        # Сводная по месяцам
        df_svod_by_month = df.groupby(['Название_месяца_рождения']).agg({'ФИО': 'count'})

        # Сортируем индекс чтобы месяцы шли в хоронологическом порядке
        # Взял отсюда https://stackoverflow.com/questions/40816144/pandas-series-sort-by-month-index
        df_svod_by_month.index = pd.CategoricalIndex(df_svod_by_month.index,
                                                     categories=['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                                                                 'Июль',
                                                                 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь',
                                                                 'Ошибочное значение!!!'],
                                                     ordered=True)
        df_svod_by_month.columns = ['Количество']
        df_svod_by_month.sort_index(inplace=True)
        for r in dataframe_to_rows(df_svod_by_month, index=True, header=True):
            wb['Свод по месяцам'].append(r)

        # Сводная по годам
        df_svod_by_year = df.groupby(['Год_рождения']).agg({'ФИО': 'count'})
        df_svod_by_year.columns = ['Количество']
        for r in dataframe_to_rows(df_svod_by_year, index=True, header=True):
            wb['Свод по годам'].append(r)

        # Сводная по 1-ПК
        df_svod_by_1PK = df.groupby(['Один_ПК']).agg({'ФИО': 'count'})
        df_svod_by_1PK.columns = ['Количество']
        for r in dataframe_to_rows(df_svod_by_1PK, index=True, header=True):
            wb['Свод по 1-ПК'].append(r)

        # Сводная по 1-ПО
        df_svod_by_1PO = df.groupby(['Один_ПО']).agg({'ФИО': 'count'})
        df_svod_by_1PO.columns = ['Количество']
        for r in dataframe_to_rows(df_svod_by_1PO, index=True, header=True):
            wb['Свод по 1-ПО'].append(r)

        # Сводная по СПО-1
        df_svod_by_SPO1 = df.groupby(['СПО_Один']).agg({'ФИО': 'count'})
        df_svod_by_SPO1.columns = ['Количество']
        for r in dataframe_to_rows(df_svod_by_SPO1, index=True, header=True):
            wb['Свод по СПО-1'].append(r)

        # Сводная по Росстату
        df_svod_by_Ros = df.groupby(['Росстат']).agg({'ФИО': 'count'})

        # Сортируем индекс
        df_svod_by_Ros.index = pd.CategoricalIndex(df_svod_by_Ros.index,
                                                   categories=['0-4', '5-9', '10-14', '15-19', '20-24', '25-29',
                                                               '30-34',
                                                               '35-39', '40-44', '45-49', '50-54', '55-59', '60-64',
                                                               '65-69',
                                                               '70 лет и старше', 'Ошибочное значение!!!'],
                                                   ordered=True)
        df_svod_by_Ros.sort_index(inplace=True)
        df_svod_by_Ros.columns = ['Количество']
        for r in dataframe_to_rows(df_svod_by_Ros, index=True, header=True):
            wb['Свод по категориям Росстата'].append(r)

        df[name_column] = df['temp']  # заменяем временной колонкой
        df.drop(columns=['temp'], inplace=True)

        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # удаляем пустой лист
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(f'{path_to_end_folder_date}/Свод по возрастам от {current_time}.xlsx')

        return df
    except NameError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Выберите файл с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')
    except KeyError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'В таблице нет такой колонки!\nПроверьте написание названия колонки')
        logging.exception('AN ERROR HAS OCCURRED')
    except FileNotFoundError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')

    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log')




if __name__ == '__main__':
    raw_selected_date_main = '01.10.2023'
    name_column_main = 'Дата_рождения'
    name_file_data_date_main = pd.read_excel('data/data.xlsx',dtype=str)
    path_to_end_folder_date_main = 'data'

    proccessing_date(raw_selected_date_main, name_column_main, name_file_data_date_main, path_to_end_folder_date_main)
    print('Lindy Booth')
