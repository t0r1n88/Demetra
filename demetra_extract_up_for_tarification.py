"""
Скрипт для извлечения данных о дисциплинах находящихся в учебном плане скачанного с сетевого города для использования при подсчете тарификации
"""
from demetra_support_functions import write_df_up_to_excel,del_sheet
import numpy as np
import pandas as pd
pd.options.display.width= None
pd.options.display.max_columns= None
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time
import re
import copy
import os
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None

from tkinter import messagebox


class NotCorrectParams(Exception):
    """
    Исключение для случаев когда нет ни одного корректного параметра
    """
    pass

class NotFile(Exception):
    """
    Обработка случаев когда нет файлов в папке
    """
    pass


def save_convert_to_int(value):
    if isinstance(value,str):
        if value.isdigit():
            return int(value)
        else:
            return value
    else:
        return value


def convert_params(quantity_header,number_main_column,quantity_cols):
    """
    Функция для конвертации строк в целые числа
    :param quantity_header: строка с количеством строк заголовка
    :param number_main_column: строка с порядковым номером колонки с извлекаемыми данными
    :param quantity_cols: количество колонок с данными которые нужно извлечь
    :return: словарь с данными
    """
    dct_params =dict()
    try:
        dct_params['quantity_header'] = int(quantity_header)
        dct_params['number_main_column'] = int(number_main_column)
        dct_params['quantity_cols'] = int(quantity_cols)
        return dct_params
    except:
        raise NotCorrectParams


def processing_data_up_for_tarification(data_folder:str,result_folder:str,name_sheet:str, quantity_header:str,number_main_column:str,quantity_cols:str):
    """
    Функция для извлечения данных для тарификации из учебных планов
    :param data_folder:папка с данными
    :param result_folder:конечная папка
    :param name_sheet:имя листа, где находятся данные
    :param quantity_header:количество строк заголовка
    :param number_main_column:порядковый номер колонки с данными которые нужно извлечь
    :param quantity_cols: количество извлекаемых колонок
    :return:
    """

    try:
        error_df = pd.DataFrame(
            columns=['Название файла','Описание ошибки'])  # датафрейм для ошибок

        dct_params = convert_params(quantity_header,number_main_column,quantity_cols)
        quantity_header = dct_params['quantity_header'] # количество строк заголовка
        number_main_column = dct_params['number_main_column'] - 1 # порядковый номер колонки с наименованиями
        quantity_cols = dct_params['quantity_cols'] # количество колонок с данными которые нужно собрать

        finish_df = pd.DataFrame(columns=list(range(quantity_cols+1))) # создаем итоговый датафрейм куда будут добавляться все данные

        for dirpath, dirnames, filenames in os.walk(data_folder):
            for file in filenames:
                if file.endswith('.xls') or file.endswith('.ods'):
                    temp_error_df = pd.DataFrame(
                        data=[[f'{file}',
                               f'Программа обрабатывает файлы с разрешением xlsx. XLS и ODS файлы не обрабатываются !'
                               ]],
                        columns=['Название файла',
                                 'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0,
                                         ignore_index=True)
                    continue
                if not file.startswith('~$') and file.endswith('.xlsx'):
                    name_file = file.split('.xlsx')[0]
                    print(name_file)  # обрабатываемый файл
                    try:
                        temp_df = pd.read_excel(f'{dirpath}/{file}',sheet_name=name_sheet,skiprows=quantity_header,header=None)  # открываем файл
                        temp_df = temp_df.iloc[:,:quantity_cols]
                    except ValueError:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{file}',
                                   f'Не найден лист с указанным названием {name_sheet}'
                                   ]],
                            columns=['Название файла',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0,
                                             ignore_index=True)
                        continue
                    except:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{file}',
                                   f'Не удалось обработать файл. Возможно файл поврежден'
                                   ]],
                            columns=['Название файла',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0,
                                             ignore_index=True)
                        continue


                    temp_df.dropna(subset=[number_main_column],inplace=True) # убираем пустые строки в колонке с наименованиями
                    temp_df.insert(0,'Имя файла',name_file) # добавляем название файла

                    # добавляем нехватающие колонки если количество колонок отличается от указанного
                    if len(temp_df.columns) < quantity_cols+1:
                        count_add_cols = quantity_cols+1 - len(temp_df.columns)
                        for i in range(count_add_cols):
                            col_name = f'Col_{i}'
                            temp_df[col_name] = None

                    temp_df.columns = list(range(quantity_cols+1)) # переименовываем колонки
                    finish_df = pd.concat([finish_df,temp_df])



        t = time.localtime()  # получаем текущее время
        current_time = time.strftime('%H_%M_%S', t)

        # Сохраняем ошибки
        wb = openpyxl.Workbook()
        for r in dataframe_to_rows(error_df, index=False, header=True):
            wb['Sheet'].append(r)

        wb['Sheet'].column_dimensions['A'].width = 30
        wb['Sheet'].column_dimensions['B'].width = 40
        wb['Sheet'].column_dimensions['C'].width = 50

        wb.save(f'{result_folder}/ОШИБКИ от {current_time}.xlsx')

        name_finish_column = finish_df.columns[number_main_column+1]
        finish_df[name_finish_column] = finish_df[name_finish_column].astype(str)
        finish_df.sort_values(by=name_finish_column,inplace=True)
        lst_finish_cols = ['Название файла']
        lst_finish_cols.extend(list(range(1,quantity_cols+1)))
        finish_df.columns = lst_finish_cols # присваиваем более понятный названия колонок

        finish_df = finish_df.applymap(save_convert_to_int)

        temp_wb = write_df_up_to_excel(
            {'Общий свод': finish_df},
            write_index=False)
        temp_wb = del_sheet(temp_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        temp_wb.save(f'{result_folder}/Общий результат {current_time}.xlsx')

        # Сохраняем по отдельным файлам
        name_column = finish_df.columns[number_main_column+1]  # получаем название колонки
        lst_value_column = finish_df.iloc[:,number_main_column+1].unique()
        lst_value_column = [value for value in lst_value_column if pd.notna(value)]
        lst_value_column = list(map(str, lst_value_column))

        used_name_file = set()  # множество для уже использованных имен файлов
        for idx, value in enumerate(lst_value_column):
            wb = openpyxl.Workbook()  # создаем файл
            temp_df = finish_df[finish_df[name_column] == value]  # отфильтровываем по значению
            # short_name = value[:40]  # получаем обрезанное значение
            value = re.sub(r'\s',' ',value)

            short_name = re.sub(r'[\r\b\n\t\'+()<>:"?*|\\/]', '_', value)
            if short_name.lower() in used_name_file:
                short_name = f'{short_name}_{idx}'  # добавляем окончание
            for row in dataframe_to_rows(temp_df, index=False, header=True):
                wb['Sheet'].append(row)

            # Устанавливаем автоширину для каждой колонки
            for column in wb['Sheet'].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                if adjusted_width >= 80:
                    wb['Sheet'].column_dimensions[column_name].width = 80
                    for cell in wb['Sheet'][column_name]:
                        cell.alignment = Alignment(horizontal='left',wrap_text=True)
                else:
                    wb['Sheet'].column_dimensions[column_name].width = adjusted_width+3

            finish_path = f'{result_folder}/По отдельности'
            if not os.path.exists(finish_path):
                os.makedirs(finish_path)

            wb.save(f'{finish_path}/{short_name}.xlsx')
            used_name_file.add(short_name.lower())
            wb.close()

        if error_df.shape[0] != 0:
            messagebox.showwarning('Деметра Отчеты социальный паспорт студента',
                                   f'Обнаружены ошибки в файлах с данными.\n'
                                   f'Проверьте файл ОШИБКИ')

    except FileNotFoundError as e:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Ошибка {e}\n\nПеренесите файлы, конечную папку с которой вы работаете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    except PermissionError as e:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Ошибка {e}\n\nЗакройте все файлы созданные Деметрой.')
    except NotCorrectParams as e:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Неправильные параметры! Количество строк заголовка, порядковый номер колонки с извлекаемыми данными, количество колонок с данными которые нужно извлечь\n'
                             f'должны быть записаны ЦЕЛЫМИ числами например 5.')
    else:
        messagebox.showinfo('Деметра Отчеты социальный паспорт студента', 'Данные успешно обработаны')













if __name__ == '__main__':
    main_data_folder = 'data/Учебные планы'
    main_result_folder = 'data/Результат'
    main_name_sheet = 'Учебный план'
    main_quantity_header = '6' # количество строк заголовка
    main_number_main_column = '2' # порядковый номер колонки с наименованиями
    main_quantity_cols = '21' # количество колонок с данными которые нужно собрать
    processing_data_up_for_tarification(main_data_folder,main_result_folder,main_name_sheet,main_quantity_header,main_number_main_column,main_quantity_cols)
    print('Lindy Booth')