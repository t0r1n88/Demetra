"""
Скрипт для создания  отчета по социальному паспорту студента БРИТ
"""
from demetra_support_functions import (write_df_to_excel, write_df_to_excel_report_brit, del_sheet,
                                       declension_fio_by_case,
                                       extract_parameters_egisso, write_df_big_dct_to_excel, check_error_in_pers_data)
from demetra_processing_date import proccessing_date
from demetra_egisso import create_part_egisso_data, create_full_egisso_data
from tkinter import messagebox

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import time
import datetime
from collections import Counter
import re
import os
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import sys
import locale


class NotColumn(Exception):
    """
    Исключение для обработки случая когда отсутствуют нужные колонки
    """
    pass


class NotGoodSheet(Exception):
    """
    Исключение для случая когда ни один лист не подхожит под эталон
    """
    pass


class ExceedingQuantity(Exception):
    """
    Исключение для случаев когда числа уникальных значений больше 255
    """
    pass


class BadEtalonFile(Exception):
    """
    Исключение для обработки случая когда эталонный файл поврежден и его нельзя открыть
    """
    pass


def set_rus_locale():
    """
    Функция чтобы можно было извлечь русские названия месяцев
    """
    locale.setlocale(
        locale.LC_ALL,
        'rus_rus' if sys.platform == 'win32' else 'ru_RU.UTF-8')


def create_value_str(df: pd.DataFrame, name_column: str, target_name_column: str, dct_str: dict,checkbox_expelled:int) -> pd.DataFrame:
    """
    Функция для формирования строки нужного формата с использованием переменных
    :param df:датафрейм
    :param name_column:название колонки для значений которой нужно произвести подсчет
    :param target_name_column: название колонки по которой будет производится подсчет
    :param dct_str:словарь с параметрами
    :param checkbox_expelled: как будет отображаться количество людей в академ и отчисленных
    :return:датафрейм
    """
    temp_counts = df[name_column].value_counts()  # делаем подсчет
    new_value_df = temp_counts.to_frame().reset_index()  # создаем датафрейм с данными
    new_value_df.columns = ['Показатель', 'Значение']  # делаем одинаковыми названия колонок
    new_value_df['Показатель'] = new_value_df['Показатель'].astype(str)
    new_value_df.sort_values(by='Показатель', inplace=True)
    for idx, row in enumerate(new_value_df.iterrows()):
        name_op = row[1].values[0]  # получаем название ОП
        temp_df = df[df[name_column] == name_op]  # отфильтровываем по названию ОП
        quantity_study_student = temp_df[temp_df[target_name_column] == dct_str['Обучается']].shape[
            0]  # со статусом Обучается
        quantity_academ_student = temp_df[temp_df[target_name_column].str.contains(dct_str['Академ'])].shape[
            0]
        quantity_not_status_student = \
        temp_df[temp_df[target_name_column].str.contains(dct_str['Не указан статус'])].shape[
            0]

        if checkbox_expelled == 0:
            out_str = f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student},' \
                      f' Не указан статус - {quantity_not_status_student}, Всего {len(temp_df)}'
        elif checkbox_expelled == 1:
            out_str = f'Обучается - {quantity_study_student},' \
                      f' Не указан статус - {quantity_not_status_student}, Всего {len(temp_df)}'
        else:
            quantity_except = temp_df[temp_df[target_name_column].str.contains('Отчислен')].shape[
                0]
            out_str = f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student}, Отчислено - {quantity_except}, Не указан статус - {quantity_not_status_student}, Всего {len(temp_df)}'




        new_value_df.iloc[idx, 1] = out_str  # присваиваем значение

    return new_value_df


def count_value(group: pd.Series, target_value: str):
    """
    Функция для группировки по конкретному значению
    Возвращает количество значений target_value в группе
    """
    # считаем сколько значений подходят по условие
    count_group = group.str.contains(target_value)
    return sum(count_group)


def convert_number(value):
    """
    Функция для конвертации в float значений колонок содержащих в названии Подсчет_
    :param value: значение
    :return: число в формате float
    """
    try:
        return float(value)
    except:
        return 0


def convert_to_date(value):
    """
    Функция для конвертации строки в текст
    :param value: значение для конвертации
    :return:
    """
    try:
        if value == 'Нет статуса':
            return 'Не заполнено'
        else:
            date_value = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
            return date_value
    except ValueError:
        result = re.search(r'^\d{2}\.\d{2}\.\d{4}$', value)
        if result:
            try:
                return datetime.datetime.strptime(result.group(0), '%d.%m.%Y')
            except ValueError:
                # для случаев вида 45.09.2007
                return f'Некорректный формат даты - {value}, проверьте лишние пробелы,наличие точек'
        else:
            return f'Некорректный формат даты - {value}, проверьте лишние пробелы,наличие точек'
    except:
        return f'Некорректный формат даты - {value}, проверьте лишние пробелы,наличие точек'


def extract_part_status_op(value,part_extract:str):
    """
    Функция для извлечения кода или наименования специальности профессии
    :param value: значение
    :param part_extract: что именно нужно извлекать либо Код_ОП либо Наименование_ОП
    :return:
    """
    lst_part = str(value).split(' ',maxsplit=1) # делим по первому пробелу
    if len(lst_part) == 2:
        if part_extract == 'Код_ОП':
            return lst_part[0]
        else:
            return lst_part[1]
    else:
        return f''






def create_report_brit(df: pd.DataFrame, dct_slice: dict, path_end_folder: str) -> None:
    """
    Функция для создания отчета по стандарту БРИТ
    :param df: копия общего датафрейма
    :param dct_slice:  словарь с названиями колонок по которым нужно сделать срез
    :param path_end_folder: куда сохранять результаты
    """
    dct_name_sheet = dict()  # словарь, где ключ это названия листа, а значение датафрейм на основе которого был произведен подсчет
    dct_report_sheet = dict()  # создаем словарь, где ключ это название листа, а содержимое сводный датафрейм
    df.fillna('Нет статуса', inplace=True)  # заполняем Наны

    for slice_column, prefix in dct_slice.items():
        group_main_df = pd.DataFrame(index=list(df[slice_column].unique()))

        # Считаем общее количество студентов
        study_df = df[df['Статус_Учёба'] == 'Обучается']
        study_df_group_df = study_df.groupby(by=[slice_column]).agg({'ФИО': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(study_df_group_df)  # добавляем в свод
        group_main_df.rename(columns={'ФИО': 'Обучается'}, inplace=True)
        study_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Обучается'] = study_df  # добавляем в словарь


        # Считаем количество студентов в академе
        akadem_df = df[df['Статус_Учёба'].isin(['Академический отпуск(декрет)','Академический отпуск(служба в РА)','Академический отпуск(по болезни)','Академический отпуск(ученич. договор)'])]
        dct_name_sheet['Академ'] = akadem_df  # добавляем в словарь
        akadem_df_group_df = akadem_df.groupby(by=[slice_column]).agg({'ФИО': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(akadem_df_group_df)  # добавляем в свод
        group_main_df.rename(columns={'ФИО': 'Академ. отпуск'}, inplace=True)
        akadem_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Академ'] = akadem_df  # добавляем в словарь



        # Совершеннолетние
        maturity_df = df[df['Совершеннолетие'] == 'совершеннолетний']
        maturity_df_group_df = maturity_df.groupby(by=[slice_column]).agg(
            {'Совершеннолетие': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(maturity_df_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Совершеннолетие': 'Совершеннолетние'}, inplace=True)
        maturity_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Совершеннолетние'] = maturity_df  # добавляем в словарь

        # Несовершеннолетние
        not_maturity_df = df[df['Совершеннолетие'] == 'несовершеннолетний']
        not_maturity_df_group_df = not_maturity_df.groupby(by=[slice_column]).agg(
            {'Совершеннолетие': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(not_maturity_df_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Совершеннолетие': 'Несовершеннолетние'}, inplace=True)
        not_maturity_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Несовершеннолетние'] = not_maturity_df  # добавляем в словарь

        # Создаем датафрейм с сиротами
        orphans_df = df[df['Статус_Сиротство'].isin(['гособеспечение + постинтернатное сопровождение',
                                                     'дети-сироты, находящиеся на полном государственном обеспечении',
                                                     'дети-сироты, находящиеся под опекой'])]

        orphans_group_df = orphans_df.groupby(by=[slice_column]).agg({'Статус_Сиротство': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(orphans_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Сиротство': 'Дети-сироты'}, inplace=True)
        orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты'] = orphans_df  # добавляем в словарь



        # Создаем датафрейм с инвалидами
        invalid_df = df[df['Статус_Уровень_здоровья'].isin(['Инвалид детства', 'Инвалид 1,2,3, группы'])]

        invalid_group_df = invalid_df.groupby(by=[slice_column]).agg(
            {'Статус_Уровень_здоровья': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(invalid_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Уровень_здоровья': 'Инвалиды'}, inplace=True)
        all_invalid_df = invalid_df.copy()
        all_invalid_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Инвалиды'] = all_invalid_df  # добавляем в словарь


        # Создаем датафрейм с совершеннолетними инвалидами
        maturity_invalid_df = invalid_df[invalid_df['Совершеннолетие'] == 'совершеннолетний']
        maturity_invalid_group_df = maturity_invalid_df.groupby(by=[slice_column]).agg(
            {'Статус_Уровень_здоровья': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(maturity_invalid_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Уровень_здоровья': 'Инвалиды совер-ние'}, inplace=True)
        maturity_invalid_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Инвалиды_сов'] = maturity_invalid_df  # добавляем в словарь


        # Создаем датафрейм с несовершеннолетними инвалидами
        not_maturity_invalid_df = invalid_df[invalid_df['Совершеннолетие'] == 'несовершеннолетний']
        not_maturity_invalid_group_df = not_maturity_invalid_df.groupby(by=[slice_column]).agg(
            {'Статус_Уровень_здоровья': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(not_maturity_invalid_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Уровень_здоровья': 'Инвалиды несов-ние'}, inplace=True)
        not_maturity_invalid_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Инвалиды_несов'] = not_maturity_invalid_df  # добавляем в словарь


        # Создаем датафрейм с получателями социальной стипендии
        soc_benefit_df = df[df['Статус_Соц_стипендия'].isin(['да'])]
        all_soc_benefit_df = soc_benefit_df.copy()
        all_soc_benefit_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Соц. стипендия Все'] = all_soc_benefit_df  # добавляем в словарь

        # получаем малоимущих
        poor_soc_benefit_df = soc_benefit_df[soc_benefit_df['Статус_Соц_положение_семьи'].isin(['Малоимущая'])]
        poor_soc_benefit_group_df = poor_soc_benefit_df.groupby(by=[slice_column]).agg(
            {'Статус_Соц_положение_семьи': 'count'})
        group_main_df = group_main_df.join(poor_soc_benefit_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Соц_положение_семьи': 'Соц. стипендия малоимущие'}, inplace=True)
        poor_soc_benefit_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Соц. стипендия малоим.'] = poor_soc_benefit_df


        # получаем детей сирот
        orphans_soc_benefit_df = soc_benefit_df[soc_benefit_df['Статус_Сиротство'].isin(
            ['гособеспечение + постинтернатное сопровождение',
             'дети-сироты, находящиеся на полном государственном обеспечении',
             'дети-сироты, находящиеся под опекой'])]
        orphans_soc_benefit_group_df = orphans_soc_benefit_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})
        group_main_df = group_main_df.join(orphans_soc_benefit_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Сиротство': 'Соц. стипендия дети-сироты'}, inplace=True)
        orphans_soc_benefit_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Соц. стипендия сироты'] = orphans_soc_benefit_df



        # получаем инвалидов
        invalid_soc_benefit_df = soc_benefit_df[
            soc_benefit_df['Статус_Уровень_здоровья'].isin(['Инвалид детства', 'Инвалид 1,2,3, группы'])]
        invalid_soc_benefit_group_df = invalid_soc_benefit_df.groupby(by=[slice_column]).agg(
            {'Статус_Уровень_здоровья': 'count'})
        group_main_df = group_main_df.join(invalid_soc_benefit_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Уровень_здоровья': 'Соц. стипендия инвалиды'}, inplace=True)
        invalid_soc_benefit_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Соц. стипендия инвалиды'] = invalid_soc_benefit_df


        # Создаем датафрейм с получателями бесплатного питания
        eating_df = df[df['Статус_Питание'].isin(
            ['получает компенсацию за питание', 'питается в ПОО', 'получает компенсацию за питание + питается в ПОО'])]
        all_eating_df = eating_df.copy()
        all_eating_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Питание все'] = all_eating_df  # добавляем в словарь

        # получаем малоимущих
        poor_eating_df = eating_df[eating_df['Статус_Соц_положение_семьи'].isin(['Малоимущая'])]
        poor_eating_group_df = poor_eating_df.groupby(by=[slice_column]).agg({'Статус_Соц_положение_семьи': 'count'})
        group_main_df = group_main_df.join(poor_eating_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Соц_положение_семьи': 'Питание малоимущие'}, inplace=True)
        poor_eating_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Питание малоим.'] = poor_eating_df


        # получаем сирот
        orphans_eating_df = eating_df[
            eating_df['Статус_Сиротство'].isin(['гособеспечение + постинтернатное сопровождение',
                                                'дети-сироты, находящиеся на полном государственном обеспечении',
                                                'дети-сироты, находящиеся под опекой'])]
        orphans_eating_group_df = orphans_eating_df.groupby(by=[slice_column]).agg({'Статус_Сиротство': 'count'})
        group_main_df = group_main_df.join(orphans_eating_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Сиротство': 'Питание сироты'}, inplace=True)
        orphans_eating_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Питание сироты'] = orphans_eating_df


        # получаем инвалидов
        invalid_eating_df = eating_df[
            eating_df['Статус_Уровень_здоровья'].isin(['Инвалид детства', 'Инвалид 1,2,3, группы'])]
        invalid_eating_group_df = invalid_eating_df.groupby(by=[slice_column]).agg({'Статус_Уровень_здоровья': 'count'})
        group_main_df = group_main_df.join(invalid_eating_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Уровень_здоровья': 'Питание инвалиды'}, inplace=True)
        invalid_eating_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Питание инвалиды'] = invalid_eating_df


        # получаем СВО
        svo_eating_df = eating_df[eating_df['Статус_Родитель_СВО'].isin(['да'])]
        svo_eating_group_df = svo_eating_df.groupby(by=[slice_column]).agg({'Статус_Родитель_СВО': 'count'})
        group_main_df = group_main_df.join(svo_eating_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Родитель_СВО': 'Питание СВО'}, inplace=True)
        svo_eating_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Питание СВО'] = svo_eating_df


        # Создаем датафрейм с проживающими в общежитии
        dormitory_df = df[df['Статус_Общежитие'].isin(['да'])]
        dormitory_group_df = dormitory_df.groupby(by=[slice_column]).agg({'Статус_Общежитие': 'count'})
        group_main_df = group_main_df.join(dormitory_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Общежитие': 'Общежитие Всего'}, inplace=True)
        all_dormitory_df = dormitory_df.copy()
        all_dormitory_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Общежитие все'] = all_dormitory_df  # добавляем в словарь


        # получаем не сирот и сохраняем для списка
        not_orphans_dormitory_df = dormitory_df[~dormitory_df['Статус_Сиротство'].isin(
            ['гособеспечение + постинтернатное сопровождение',
             'дети-сироты, находящиеся на полном государственном обеспечении',
             'дети-сироты, находящиеся под опекой'])]
        not_orphans_dormitory_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Общежитие кроме сирот'] = not_orphans_dormitory_df

        # получаем сирот
        orphans_dormitory_df = dormitory_df[dormitory_df['Статус_Сиротство'].isin(
            ['гособеспечение + постинтернатное сопровождение',
             'дети-сироты, находящиеся на полном государственном обеспечении',
             'дети-сироты, находящиеся под опекой'])]
        orphans_dormitory_group_df = orphans_dormitory_df.groupby(by=[slice_column]).agg({'Статус_Сиротство': 'count'})
        group_main_df = group_main_df.join(orphans_dormitory_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Сиротство': 'Общежитие сироты'}, inplace=True)
        orphans_dormitory_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Общежитие сироты'] = orphans_dormitory_df


        # Считаем выпуск текущего год
        release_df = df[df['Статус_Выпуск'].isin(['да'])]
        all_release_df = release_df.copy()
        all_release_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Выпуск текущий год'] = all_release_df  # добавляем в словарь

        # Считаем сирот
        release_orphans_df = release_df[release_df['Статус_Сиротство'].isin(
            ['гособеспечение + постинтернатное сопровождение',
             'дети-сироты, находящиеся на полном государственном обеспечении',
             'дети-сироты, находящиеся под опекой'])]

        release_orphans_group_df = release_orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(release_orphans_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Сиротство': 'Выпуск сироты'}, inplace=True)
        release_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Выпуск сироты'] = release_orphans_df  # добавляем в словарь


        # Считаем инвалидов
        release_invalid_df = release_df[
            release_df['Статус_Уровень_здоровья'].isin(['Инвалид детства', 'Инвалид 1,2,3, группы'])]

        release_invalid_group_df = release_invalid_df.groupby(by=[slice_column]).agg(
            {'Статус_Уровень_здоровья': 'count'})  # создаем базовый
        group_main_df = group_main_df.join(release_invalid_group_df)  # добавляем в свод
        group_main_df.rename(columns={'Статус_Уровень_здоровья': 'Выпуск инвалиды'}, inplace=True)
        release_invalid_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Выпуск инвалиды'] = release_invalid_df  # добавляем в словарь


        group_main_df.fillna(0, inplace=True)  # заполняем наны
        group_main_df = group_main_df.astype(int)  # приводим к инту
        sum_row = group_main_df.sum(axis=0)  # суммируем колонки

        if slice_column == 'Текущий_возраст' or slice_column == 'Год_рождения':
            # для проведения сортировки
            group_main_df.rename(index={'Ошибочное значение!!!': 100000000}, inplace=True)
            group_main_df.sort_index(inplace=True)
            group_main_df.rename(index={100000000: 'Ошибочное значение!!!'}, inplace=True)

        group_main_df.loc['Итого'] = sum_row  # добавляем суммирующую колонку

        group_orphans_main_df = pd.DataFrame(index=list(df[slice_column].unique()))  # Базовый датафрейм для сирот

        # Считаем общее количество
        all_orphans_group_df = orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})  # создаем базовый
        group_orphans_main_df = group_orphans_main_df.join(all_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Всего'}, inplace=True)

        # Считаем сирот совершеннолетних
        maturity_orphans_df = orphans_df[orphans_df['Совершеннолетие'] == 'совершеннолетний']
        maturity_orphans_group_df = maturity_orphans_df.groupby(by=[slice_column]).agg({'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(maturity_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Сироты совершеннолетние'}, inplace=True)
        maturity_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты совер-ние'] = maturity_orphans_df


        # Считаем сирот несовершеннолетних
        not_maturity_orphans_df = orphans_df[orphans_df['Совершеннолетие'] == 'несовершеннолетний']
        not_maturity_orphans_group_df = not_maturity_orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(not_maturity_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Сироты несовершеннолетние'}, inplace=True)
        not_maturity_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты несовер-ние'] = not_maturity_orphans_df


        # считаем академ
        akadem_orphans_df = orphans_df[orphans_df['Статус_Учёба'].isin(['Академический отпуск(декрет)','Академический отпуск(служба в РА)','Академический отпуск(по болезни)','Академический отпуск(ученич. договор)'])]
        akadem_orphans_group_df = akadem_orphans_df.groupby(by=[slice_column]).agg({'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(akadem_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Из них в академическом отпуске'}, inplace=True)
        akadem_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты академ'] = akadem_orphans_df



        # считаем постинтернатное сопровождение
        postinternat_orphans_df = orphans_df[
            orphans_df['Статус_Сиротство'].isin(['гособеспечение + постинтернатное сопровождение'])]
        postinternat_orphans_group_df = postinternat_orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(postinternat_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Из них постинтернат'}, inplace=True)
        postinternat_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты постинтернат'] = postinternat_orphans_df


        # считаем сирот на гособеспечении
        gos_orphans_df = orphans_df[
            orphans_df['Статус_Сиротство'].isin(['гособеспечение + постинтернатное сопровождение',
                                                 'дети-сироты, находящиеся на полном государственном обеспечении'
                                                 ])]
        gos_orphans_group_df = gos_orphans_df.groupby(by=[slice_column]).agg({'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(gos_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Из них на гособеспечении'}, inplace=True)
        gos_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты гособеспечение'] = gos_orphans_df


        # считаем сирот с опекой
        custody_orphans_df = orphans_df[orphans_df['Статус_Сиротство'].isin(['дети-сироты, находящиеся под опекой'])]
        custody_orphans_group_df = custody_orphans_df.groupby(by=[slice_column]).agg({'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(custody_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Из них под опекой'}, inplace=True)
        custody_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты опека'] = custody_orphans_df


        # считаем сирот получающих питание
        all_eating_orphans_df = orphans_df[
            orphans_df['Статус_Питание'].isin(['питается в ПОО', 'получает компенсацию за питание',
                                               'получает компенсацию за питание + питается в ПОО'])]
        all_eating_orphans_group_df = all_eating_orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(all_eating_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Питается всего'}, inplace=True)
        all_eating_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты питание все'] = all_eating_orphans_df


        # считаем сирот получающих питание в брит
        brit_eating_orphans_df = orphans_df[orphans_df['Статус_Питание'].isin(['питается в ПОО'])]
        brit_eating_orphans_group_df = brit_eating_orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(brit_eating_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Питается в ПОО'}, inplace=True)
        brit_eating_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты питание ПОО'] = brit_eating_orphans_df



        # считаем сирот получающих компенсацию
        compens_eating_orphans_df = orphans_df[orphans_df['Статус_Питание'].isin(['получает компенсацию за питание'])]
        compens_eating_orphans_group_df = compens_eating_orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(compens_eating_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'Получает компенсацию'}, inplace=True)
        compens_eating_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты питание компенсация'] = compens_eating_orphans_df


        # считаем сирот получающих компенсацию + питание в брит
        brit_compens_eating_orphans_df = orphans_df[
            orphans_df['Статус_Питание'].isin(['получает компенсацию за питание + питается в ПОО'])]
        brit_compens_eating_orphans_group_df = brit_compens_eating_orphans_df.groupby(by=[slice_column]).agg(
            {'Статус_Сиротство': 'count'})
        group_orphans_main_df = group_orphans_main_df.join(brit_compens_eating_orphans_group_df)  # добавляем в свод
        group_orphans_main_df.rename(columns={'Статус_Сиротство': 'ПОО+компенсация'}, inplace=True)
        brit_compens_eating_orphans_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Сироты питание ПОО+компенсация'] = brit_compens_eating_orphans_df


        group_orphans_main_df.fillna(0, inplace=True)  # заполняем наны
        group_orphans_main_df = group_orphans_main_df.astype(int)  # приводим к инту
        sum_row = group_orphans_main_df.sum(axis=0)  # суммируем колонки
        if slice_column == 'Текущий_возраст' or slice_column == 'Год_рождения':
            # для проведения сортировки
            group_orphans_main_df.rename(index={'Ошибочное значение!!!': 100000000}, inplace=True)
            group_orphans_main_df.sort_index(inplace=True)
            group_orphans_main_df.rename(index={100000000: 'Ошибочное значение!!!'}, inplace=True)
        group_orphans_main_df.loc['Итого'] = sum_row  # добавляем суммирующую колонку

        # отчет по учетам

        group_accounting_main_df = pd.DataFrame(index=list(df[slice_column].unique()))  # Базовый датафрейм для учетов

        # Считаем КДН
        kdn_accounting_df = df[df['Статус_КДН'].isin(['состоит'])]
        kdn_group_accounting_df = kdn_accounting_df.groupby(by=[slice_column]).agg({'Статус_КДН': 'count'})
        group_accounting_main_df = group_accounting_main_df.join(kdn_group_accounting_df)  # добавляем в свод
        kdn_accounting_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['КДН'] = kdn_accounting_df



        # Считаем ПДН
        pdn_accounting_df = df[df['Статус_ПДН'].isin(['состоит'])]
        pdn_group_accounting_df = pdn_accounting_df.groupby(by=[slice_column]).agg({'Статус_ПДН': 'count'})
        group_accounting_main_df = group_accounting_main_df.join(pdn_group_accounting_df)  # добавляем в свод
        pdn_accounting_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['ПДН'] = pdn_accounting_df



        # Слаживаем колонки и удаляем лишнее
        group_accounting_main_df.fillna(0, inplace=True)
        group_accounting_main_df['Учет КДН, ПДН'] = group_accounting_main_df['Статус_КДН'] + group_accounting_main_df[
            'Статус_ПДН']
        group_accounting_main_df.drop(columns=['Статус_КДН', 'Статус_ПДН'], inplace=True)

        # Считаем внутренний учет
        inside_accounting_df = df[df['Статус_Внутр_учет'].isin(['состоит'])]
        inside_group_accounting_df = inside_accounting_df.groupby(by=[slice_column]).agg({'Статус_Внутр_учет': 'count'})
        group_accounting_main_df = group_accounting_main_df.join(inside_group_accounting_df)  # добавляем в свод
        group_accounting_main_df.rename(columns={'Статус_Внутр_учет': 'Внутренний учет'}, inplace=True)
        inside_accounting_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Внутр_учет'] = inside_accounting_df


        # Считаем самовольный уход
        awol_accounting_df = df[df['Статус_Самовольный_уход'].isin(['да'])]
        awol_group_accounting_df = awol_accounting_df.groupby(by=[slice_column]).agg(
            {'Статус_Самовольный_уход': 'count'})
        group_accounting_main_df = group_accounting_main_df.join(awol_group_accounting_df)  # добавляем в свод
        group_accounting_main_df.rename(columns={'Статус_Самовольный_уход': 'Самовольный уход'}, inplace=True)
        awol_accounting_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['Самовольный уход'] = awol_accounting_df



        # Считаем соп
        sop_accounting_df = df[df['Статус_Соц_положение_семьи'].isin(['СОП'])]
        sop_group_accounting_df = sop_accounting_df.groupby(by=[slice_column]).agg(
            {'Статус_Соц_положение_семьи': 'count'})
        group_accounting_main_df = group_accounting_main_df.join(sop_group_accounting_df)  # добавляем в свод
        group_accounting_main_df.rename(columns={'Статус_Соц_положение_семьи': 'СОП'}, inplace=True)
        sop_accounting_df.replace('Нет статуса', '', inplace=True)
        dct_name_sheet['СОП'] = sop_accounting_df



        group_accounting_main_df.fillna(0, inplace=True)  # заполняем наны
        group_accounting_main_df = group_accounting_main_df.astype(int)  # приводим к инту
        sum_row = group_accounting_main_df.sum(axis=0)  # суммируем колонки
        if slice_column == 'Текущий_возраст' or slice_column == 'Год_рождения':
            # для проведения сортировки
            group_accounting_main_df.rename(index={'Ошибочное значение!!!': 100000000}, inplace=True)
            group_accounting_main_df.sort_index(inplace=True)
            group_accounting_main_df.rename(index={100000000: 'Ошибочное значение!!!'}, inplace=True)

        group_accounting_main_df.loc['Итого'] = sum_row  # добавляем суммирующую колонку

        # Добавляем в словарь
        dct_report_sheet[f'Соцпаспорт_{prefix}'] = group_main_df
        dct_report_sheet[f'Сироты_{prefix}'] = group_orphans_main_df
        dct_report_sheet[f'Учет_{prefix}'] = group_accounting_main_df

    # получаем текущее время
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    # Создаем файл для самого отчета
    report_wb = write_df_to_excel_report_brit(dct_report_sheet, write_index=True)
    report_wb = del_sheet(report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
    report_wb.save(f'{path_end_folder}/Отчет по стандарту БРИТ от {current_time}.xlsx')

    # Сохраняем списки
    lst_report_wb = write_df_to_excel(dct_name_sheet, write_index=False)
    lst_report_wb = del_sheet(lst_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
    lst_report_wb.save(f'{path_end_folder}/Списки для отчета по стандарту БРИТ от {current_time}.xlsx')


def create_svod_counting(df: pd.DataFrame, name_column: str, postfix_file: str, lst_counting_name_columns: list,
                         path: str, current_time):
    """
    Функция для создания сводов по колонкам подсчета
    :param df: основной датафрейм
    :param name_column: название колонки
    :param postfix_file: дополнение к имени файла
    :param lst_counting_name_columns: список колонок типа Подсчет
    :param path: путь куда сохранять файлы
    :param current_time: время под которым будет сохраняться файл
    """
    dct_counting_df = {}  # словарь для хранения датафреймов
    for name_counting_column in lst_counting_name_columns:
        temp_svod_df = (pd.pivot_table(df, index=[name_column],
                                       values=[name_counting_column],
                                       aggfunc=[np.mean, np.sum, np.median, np.min, np.max, len]))
        temp_svod_df = temp_svod_df.reset_index()  # убираем мультииндекс
        temp_svod_df = temp_svod_df.droplevel(axis=1, level=0)  # убираем мультиколонки
        temp_svod_df.columns = [name_column, 'Среднее', 'Сумма', 'Медиана', 'Минимум', 'Максимум',
                                'Количество']
        if name_column == 'Текущий_возраст' or name_column == 'Год_рождения':
            # для проведения сортировки
            temp_svod_df.rename(index={'Ошибочное значение!!!': 100000000}, inplace=True)
            temp_svod_df.sort_index(inplace=True)
            temp_svod_df.rename(index={100000000: 'Ошибочное значение!!!'}, inplace=True)

        dct_counting_df[name_counting_column] = temp_svod_df  # сохраняем в словарь

    # Сохраняем контролируя количество листов в файле
    if len(dct_counting_df) >= 252:
        lst_dct = list(dct_counting_df.items())  # превращаем в список кортежей
        start_threshold = 0
        end_threshold = 252
        count = 1
        while end_threshold <= len(lst_dct) + 1:
            big_dct = dict(lst_dct[start_threshold:end_threshold])
            counting_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
            counting_report_wb = del_sheet(counting_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            counting_report_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
            count += 1
            start_threshold += 252
            end_threshold += 252
            # контролируем диапазон
            if end_threshold > len(lst_dct):
                big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                counting_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                counting_report_wb = del_sheet(counting_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                counting_report_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
    else:
        counting_report_wb = write_df_big_dct_to_excel(dct_counting_df, write_index=False)
        counting_report_wb = del_sheet(counting_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        counting_report_wb.save(f'{path}/Свод по {postfix_file} от {current_time}.xlsx')

def create_svod_list(df: pd.DataFrame, name_column: str, postfix_file: str, lst_list_name_columns: list,
                         path: str, current_time, dct_values_in_list_columns:dict):
    """
    Функция для создания сводов по колонкам списков
    :param df: основной датафрейм
    :param name_column: название колонки
    :param postfix_file: дополнение к имени файла
    :param lst_counting_name_columns: список колонок типа Список
    :param path: путь куда сохранять файлы
    :param current_time: время под которым будет сохраняться файл
    """
    dct_svod_list_df = {}  # словарь в котором будут храниться датафреймы по названию колонок
    for name_lst_column in lst_list_name_columns:
        file_cols = dct_values_in_list_columns[name_lst_column]
        file_cols.insert(0, name_column)
        dct_svod_list_df[name_lst_column] = pd.DataFrame(columns=file_cols)

    lst_file = df[name_column].unique()  # список файлов
    for name in lst_file:
        temp_df = df[df[name_column] == name]
        for name_lst_column in lst_list_name_columns:

            temp_col_value_lst = temp_df[name_lst_column].tolist()  # создаем список
            temp_col_value_lst = [value for value in temp_col_value_lst if value]  # отбрасываем пустые значения
            temp_unwrap_lst = []
            temp_col_value_lst = list(map(str, temp_col_value_lst))  # делаем строковым каждый элемент
            for value in temp_col_value_lst:
                temp_unwrap_lst.extend(value.split(','))
            temp_unwrap_lst = list(map(str.strip, temp_unwrap_lst))  # получаем список
            # убираем повторения и сортируем
            dct_values_in_list_columns[name_lst_column] = sorted(list(set(temp_unwrap_lst)))

            dct_value_list = dict(Counter(temp_unwrap_lst))  # Превращаем в словарь
            sorted_dct_value_lst = dict(sorted(dct_value_list.items()))  # сортируем словарь
            # создаем датафрейм
            temp_svod_df = pd.DataFrame(list(sorted_dct_value_lst.items()),
                                        columns=['Показатель', 'Значение']).transpose()
            new_temp_cols = temp_svod_df.iloc[0]  # получаем первую строку для названий
            temp_svod_df = temp_svod_df[1:]  # удаляем первую строку
            temp_svod_df.columns = new_temp_cols
            temp_svod_df.insert(0, name_column, name)
            # добавляем значение в датафрейм
            dct_svod_list_df[name_lst_column] = pd.concat([dct_svod_list_df[name_lst_column], temp_svod_df])

    for key, value_df in dct_svod_list_df.items():
        dct_svod_list_df[key].fillna(0, inplace=True)  # заполняем наны
        dct_svod_list_df[key] = dct_svod_list_df[key].astype(int, errors='ignore')
        if name_column == 'Текущий_возраст' or name_column == 'Год_рождения':
            dct_svod_list_df[key] = dct_svod_list_df[key].astype(str) # делаем строковой
            # заменяем ошибочное значение числом чтобы отсортировать
            dct_svod_list_df[key] = dct_svod_list_df[key].apply(lambda x:x.replace('Ошибочное значение!!!',100000000))
            dct_svod_list_df[key] = dct_svod_list_df[key].astype(float) # делаем числом
            dct_svod_list_df[key] = dct_svod_list_df[key].astype(int) # делаем числом
            if name_column == 'Текущий_возраст':
                dct_svod_list_df[key].sort_values(by='Текущий_возраст',inplace=True)
            else:
                dct_svod_list_df[key].sort_values(by='Год_рождения', inplace=True)


        sum_row = dct_svod_list_df[key].sum(axis=0)  # суммируем колонки
        dct_svod_list_df[key].loc['Итого'] = sum_row  # добавляем суммирующую колонку
        dct_svod_list_df[key].iloc[-1, 0] = 'Итого'
        if name_column == 'Текущий_возраст':
            dct_svod_list_df[key]['Текущий_возраст'] = dct_svod_list_df[key]['Текущий_возраст'] .astype(str)  # делаем строковой
            dct_svod_list_df[key]['Текущий_возраст']  = dct_svod_list_df[key]['Текущий_возраст'] .apply(lambda x: x.replace('100000000','Ошибочное значение!!!'))
        if name_column == 'Год_рождения':
            dct_svod_list_df[key]['Год_рождения'] = dct_svod_list_df[key]['Год_рождения'] .astype(str)  # делаем строковой
            dct_svod_list_df[key]['Год_рождения']  = dct_svod_list_df[key]['Год_рождения'] .apply(lambda x: x.replace('100000000','Ошибочное значение!!!'))
        # Сохраняем

    if len(dct_svod_list_df) >= 252:
        lst_dct = list(dct_svod_list_df.items())  # превращаем в список кортежей
        start_threshold = 0
        end_threshold = 252
        count = 1
        while end_threshold <= len(lst_dct) + 1:
            big_dct = dict(lst_dct[start_threshold:end_threshold])
            list_columns_svod_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
            list_columns_svod_wb = del_sheet(list_columns_svod_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            list_columns_svod_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
            count += 1
            start_threshold += 252
            end_threshold += 252
            # контролируем диапазон
            if end_threshold > len(lst_dct):
                big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                list_columns_svod_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                list_columns_svod_wb = del_sheet(list_columns_svod_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                list_columns_svod_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
    else:
        list_columns_svod_wb = write_df_big_dct_to_excel(dct_svod_list_df, write_index=False)
        list_columns_svod_wb = del_sheet(list_columns_svod_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        list_columns_svod_wb.save(f'{path}/Свод по {postfix_file} {current_time}.xlsx')


def create_svod_status(df: pd.DataFrame, name_column: str, postfix_file: str, lst_status: list,
                         path: str, current_time):
    """
    Функция для создания сводов по колонкам подсчета
    :param df: основной датафрейм
    :param name_column: название колонки
    :param postfix_file: дополнение к имени файла
    :param lst_counting_name_columns: список колонок типа Список
    :param path: путь куда сохранять файлы
    :param current_time: время под которым будет сохраняться файл
    """
    dct_status = {}  # словарь для хранения сводных датафреймов
    lst_status_columns = lst_status.copy()

    if name_column == 'Статус_Пол':
        lst_status_columns.remove('Статус_Пол')
    if name_column == 'Статус_ОП':
        lst_status_columns.remove('Статус_ОП')

    for name_status_column in lst_status_columns:
        svod_df = pd.pivot_table(df, index=name_column, columns=name_status_column, values='ФИО',
                                 aggfunc='count', fill_value=0, margins=True,
                                 margins_name='Итого').reset_index()
        name_sheet = name_status_column.replace('Статус_', '')
        dct_status[name_sheet] = svod_df  # сохраняем в словарь сводную таблицу

    # Сохраняем
    if len(dct_status) >= 252:
        lst_dct = list(dct_status.items())  # превращаем в список кортежей
        start_threshold = 0
        end_threshold = 252
        count = 1
        while end_threshold <= len(lst_dct) + 1:
            big_dct = dict(lst_dct[start_threshold:end_threshold])
            svod_status_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
            svod_status_wb = del_sheet(svod_status_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            svod_status_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
            count += 1
            start_threshold += 252
            end_threshold += 252
            # контролируем диапазон
            if end_threshold > len(lst_dct):
                big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                svod_status_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                svod_status_wb = del_sheet(svod_status_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                svod_status_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
    else:
        svod_status_wb = write_df_big_dct_to_excel(dct_status, write_index=False)
        svod_status_wb = del_sheet(svod_status_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        svod_status_wb.save(f'{path}/Свод по {postfix_file} {current_time}.xlsx')



def create_social_report(etalon_file: str, data_folder: str, path_egisso_params: str, path_end_folder: str,
                         checkbox_expelled: int, raw_date) -> None:
    """
    Функция для генерации отчета по социальному статусу студентов БРИТ
    """
    try:
        set_rus_locale()  # устанавливаем русскую локаль что категоризация по месяцам работала
        # обязательные колонки
        name_columns_set = {'ФИО', 'Дата_рождения', 'Статус_ОП', 'Статус_Бюджет', 'Статус_Общежитие', 'Статус_Учёба',
                            'Статус_Всеобуч',
                            'Статус_Соц_стипендия', 'Статус_Соц_положение_семьи',
                            'СНИЛС','ИНН', 'Пол', 'Серия_паспорта', 'Номер_паспорта', 'Дата_выдачи_паспорта', 'Кем_выдан',
                            'Адрес_регистрации','Фактический_адрес',
                            'Статус_Питание',
                            'Статус_Состав_семьи', 'Статус_Уровень_здоровья', 'Статус_Сиротство',
                            'Статус_Место_регистрации', 'Статус_Студенческая_семья',
                             'Статус_Родитель_СВО',
                            'Статус_ПДН', 'Статус_КДН', 'Статус_Внутр_учет', 'Статус_Самовольный_уход', 'Статус_Выпуск'}
        error_df = pd.DataFrame(
            columns=['Название файла', 'Название листа', 'Значение ошибки', 'Описание ошибки'])  # датафрейм для ошибок
        try:
            wb = openpyxl.load_workbook(etalon_file)  # загружаем эталонный файл
            quantity_sheets = 0  # считаем количество групп
            main_sheet = wb.sheetnames[0]  # получаем название первого листа с которым и будем сравнивать новые файлы
            main_df = pd.read_excel(etalon_file, sheet_name=main_sheet,
                                    nrows=0)  # загружаем датафрейм чтобы получить эталонные колонки
        except:
            raise BadEtalonFile

        # Проверяем на обязательные колонки
        etallon_always_cols = name_columns_set.difference(set(main_df.columns))
        if len(etallon_always_cols) != 0:
            raise NotColumn
        etalon_cols = set(main_df.columns)  # эталонные колонки

        for idx, file in enumerate(os.listdir(data_folder)):
            if not file.startswith('~$') and not file.endswith('.xlsx'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '', '',
                                                    'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(f'Файл: {name_file}')
                try:
                    temp_wb = openpyxl.load_workbook(f'{data_folder}/{file}')  # открываем
                except:
                    temp_error_df = pd.DataFrame(
                        data=[[f'{name_file}', f'', f'',
                               'Не удалось обработать файл.']],
                        columns=['Название файла', 'Название листа', 'Значение ошибки',
                                 'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue  # не обрабатываем лист, где найдены ошибки
                lst_sheets_temp_wb = temp_wb.sheetnames  # получаем список листов в файле
                for name_sheet in lst_sheets_temp_wb:
                    if name_sheet != 'Данные для выпадающих списков':  # отбрасываем лист с даннными выпадающих списков
                        try:
                            temp_df = pd.read_excel(f'{data_folder}/{file}', sheet_name=name_sheet,
                                                    dtype=str)  # получаем колонки которые есть на листе
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'', f'',
                                       'Не удалось обработать файл.']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки
                        # Проверяем на обязательные колонки
                        always_cols = name_columns_set.difference(set(temp_df.columns))
                        if len(always_cols) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(always_cols)}',
                                       'В файле на указанном листе не найдены указанные обязательные колонки. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки

                        # Проверяем порядок колонок
                        order_main_columns = list(main_df.columns)  # порядок колонок эталонного файла
                        order_temp_df_columns = list(temp_df.columns)  # порядок колонок проверяемого файла
                        error_order_lst = []  # список для несовпадающих пар
                        # Сравниваем попарно колонки
                        for main, temp in zip(order_main_columns, order_temp_df_columns):
                            if main != temp:
                                error_order_lst.append(f'На месте колонки {main} находится колонка {temp}')
                        if len(error_order_lst) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(error_order_lst)}',
                                       'Неправильный порядок колонок по сравнению с эталоном. Приведите порядок колонок в соответствии с порядком в эталоне. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки


                        # проверяем на соответствие эталонному файлу
                        diff_cols = (set(temp_df.columns).symmetric_difference(etalon_cols))
                        if len(diff_cols) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(diff_cols)}',
                                       'В файле на указанном листе найдены лишние или отличающиеся колонки по сравнению с эталоном. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки

                        # Проверяем на наличие колонок без названия Unnamed
                        unnamed_lst = [f'{idx} колонка не имеет названия' for idx, name_column in
                                       enumerate(temp_df.columns, 1) if 'Unnamed' in name_column]
                        if len(unnamed_lst) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(unnamed_lst)}',
                                       'В файле на указанном листе найдена(ы) колонка(и) у которых нет названия. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки


                        # указываем где есть ФИО в которых пробельные символы
                        lst_fio = temp_df['ФИО'].tolist()  # делаем список
                        lst_spaces = []  # список в котором будут храниться индексы
                        for idx, fio in enumerate(lst_fio):
                            lst_spaces.append(fio)
                            if str(fio).strip() == '':
                                lst_spaces[idx] = f'Строка {idx + 2} состоит из только из пробельных символов'

                        temp_df['ФИО'] = lst_spaces  #
                        temp_df.dropna(how='all', inplace=True)  # удаляем пустые строки
                        temp_df = temp_df[temp_df['ФИО'].notna()] # отсекаем строки где не заполнено ФИО


                        # проверяем наличие колонок Файл и Группа
                        if 'Файл' not in temp_df.columns:
                            temp_df.insert(0, 'Файл', name_file)

                        if 'Группа' not in temp_df.columns:
                            temp_df.insert(0, 'Группа', name_sheet)  # вставляем колонку с именем листа

                        if checkbox_expelled == 0:
                            temp_df = temp_df[
                                temp_df['Статус_Учёба'] != 'Отчислен']  # отбрасываем отчисленных если поставлен чекбокс
                        elif checkbox_expelled == 1:
                            temp_df = temp_df[
                                temp_df['Статус_Учёба'] != 'Отчислен']  # отбрасываем отчисленных и академистов
                            temp_df = temp_df[~temp_df['Статус_Учёба'].isin(['Академический отпуск(декрет)','Академический отпуск(служба в РА)','Академический отпуск(по болезни)','Академический отпуск(ученич. договор)'])]
                        else:
                            temp_df = temp_df


                        main_df = pd.concat([main_df, temp_df], axis=0, ignore_index=True)  # добавляем в общий файл
                        if len(temp_df) == 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'',
                                       'Пустой файл']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        else:
                            quantity_sheets += 1
        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        # Проверяем есть ли данные в общем файле, если нет то вызываем исключение
        if len(main_df) == 0:
            error_df.to_excel(f'{path_end_folder}/Ошибки от {current_time}.xlsx', index=False)
            raise NotGoodSheet
        main_df.rename(columns={'Группа': 'Для переноса', 'Файл': 'файл для переноса'},
                       inplace=True)  # переименовываем группу чтобы перенести ее в начало таблицы
        main_df.insert(0, 'Файл', main_df['файл для переноса'])
        main_df.insert(1, 'Группа', main_df['Для переноса'])
        main_df.drop(columns=['Для переноса', 'файл для переноса'], inplace=True)

        main_df.fillna('Нет статуса', inplace=True)  # заполняем пустые ячейки

        # Сохраняем лист со всеми данными
        lst_date_columns = []  # список для колонок с датами
        for column in main_df.columns:
            if 'дата' in column.lower():
                lst_date_columns.append(column)
        main_df[lst_date_columns] = main_df[lst_date_columns].applymap(convert_to_date)  # Приводим к типу
        main_df[lst_date_columns] = main_df[lst_date_columns].applymap(
            lambda x: x.strftime('%d.%m.%Y') if isinstance(x, (pd.Timestamp, datetime.datetime)) and pd.notna(x) else x)

        main_df.replace('Нет статуса', '', inplace=True)

        # Добавляем раздельные колонки для кода ОП и названия ОП
        main_df['Код_ОП'] = main_df['Статус_ОП'].apply(lambda x:extract_part_status_op(x,'Код_ОП'))
        main_df['Наименование_ОП'] = main_df['Статус_ОП'].apply(lambda x:extract_part_status_op(x,'Наименование_ОП'))

        # Добавляем разбиение по датам
        main_df = proccessing_date(raw_date, 'Дата_рождения', main_df, path_end_folder)

        # Добавляем колонки со склоненными ФИО
        main_df = declension_fio_by_case(main_df)

        main_df.columns = list(map(str, list(main_df.columns)))

        # Обрабатываем колонки типа Подсчет если они есть
        lst_counting_name_columns = [name_column for name_column in main_df.columns if 'Подсчет_' in name_column]
        if len(lst_counting_name_columns) != 0:
            # Создаем файл в котором будут сводные данные по колонкам с Подсчетом
            dct_counting_save_name = {'Файл': 'по группам', 'Текущий_возраст': 'по возрастам','Год_рождения': 'по годам рождения', 'Статус_ОП': 'по ОП',
                                      'Пол': 'по полам'}  # словарь для названий колонок по которым будут создаваться файлы
            # Создаем папку для хранения сводов по колонкам подсчета
            path_counting_file = f'{path_end_folder}/Своды по колонкам Подсчетов'  #
            if not os.path.exists(path_counting_file):
                os.makedirs(path_counting_file)
            # приводим к числовому формату
            for name_counting_column in lst_counting_name_columns:
                main_df[name_counting_column] = main_df[name_counting_column].apply(convert_number)
            # создаем своды
            for name_column, name_file in dct_counting_save_name.items():
                create_svod_counting(main_df.copy(), name_column, name_file, lst_counting_name_columns,
                                     path_counting_file, current_time)

        # генерируем отчет по стандарту БРИТ
        create_report_brit(main_df.copy(), {'Файл': 'по группам', 'Текущий_возраст': 'по возрастам','Год_рождения': 'по годам рождения',
                                            'Пол': 'по полам', 'Статус_ОП': 'по ОП', }, path_end_folder)

        # Генерируем файлы егиссо
        # генерируем полный вариант
        df_params_egisso, temp_params_egisso_error_df = extract_parameters_egisso(path_egisso_params,
                                                                                  list(main_df.columns))
        path_egisso_file = f'{path_end_folder}/ЕГИССО'  # создаем папку для хранения файлов егиссо
        if not os.path.exists(path_egisso_file):
            os.makedirs(path_egisso_file)

        if len(df_params_egisso) != 0:
            egisso_full_wb, egisso_not_find_wb, egisso_error_wb = create_full_egisso_data(main_df, df_params_egisso,
                                                                                          path_egisso_file)  # создаем полный набор данных
            egisso_full_wb.save(f'{path_egisso_file}/ЕГИССО полные данные от {current_time}.xlsx')
            egisso_not_find_wb.save(f'{path_egisso_file}/ЕГИССО Не найденные льготы {current_time}.xlsx')
            egisso_error_wb.save(f'{path_egisso_file}/ЕГИССО перс данные ОШИБКИ от {current_time}.xlsx')


        else:
            # генерируем вариант только с персональными данными
            egisso_clean_wb, egisso_error_wb = create_part_egisso_data(main_df)
            egisso_clean_wb.save(f'{path_egisso_file}/ЕГИССО перс данные от {current_time}.xlsx')
            egisso_error_wb.save(f'{path_egisso_file}/ЕГИССО перс данные ОШИБКИ от {current_time}.xlsx')

        # Сохраняем лист с ошибками

        error_df = pd.concat([error_df, temp_params_egisso_error_df], axis=0, ignore_index=True)
        error_wb = write_df_to_excel({'Ошибки': error_df}, write_index=False)
        error_wb.save(f'{path_end_folder}/Ошибки в файлах от {current_time}.xlsx')


        # Обрабатываем колонки типа Список
        lst_list_name_columns = [name_column for name_column in main_df.columns if 'Список_' in name_column]
        if len(lst_list_name_columns) != 0:
            dct_list_columns = {}  # словарь в котором будут храниться датафреймы созданные для каждой колонки со списокм
            dct_values_in_list_columns = {}  # словарь в котором будут храниться названия колонок и все значения которые там встречались
            dct_df_list_in_columns = {}  # словарь где будут храниться значения в колонках и датафреймы где в указанных колонках есть соответствующее значение

            dct_list_save_name = {'Файл': 'по группам', 'Текущий_возраст': 'по возрастам', 'Год_рождения': 'по годам рождения', 'Статус_ОП': 'по ОП',
                                      'Пол': 'по полам'}  # словарь для названий колонок по которым будут создаваться срезы
            # Создаем папку для хранения сводов по колонкам списков
            path_list_file = f'{path_end_folder}/Своды по колонкам Списков'  #
            if not os.path.exists(path_list_file):
                os.makedirs(path_list_file)


            main_df[lst_list_name_columns] = main_df[lst_list_name_columns].astype(str)
            for name_lst_column in lst_list_name_columns:
                temp_col_value_lst = main_df[name_lst_column].tolist()  # создаем список
                temp_col_value_lst = [value for value in temp_col_value_lst if value]  # отбрасываем пустые значения
                unwrap_lst = []
                temp_col_value_lst = list(map(str, temp_col_value_lst))  # делаем строковым каждый элемент
                for value in temp_col_value_lst:
                    unwrap_lst.extend(value.split(','))
                unwrap_lst = list(map(str.strip, unwrap_lst))  # получаем список
                # убираем повторения и сортируем
                dct_values_in_list_columns[name_lst_column] = sorted(list(set(unwrap_lst)))

                dct_value_list = dict(Counter(unwrap_lst))  # Превращаем в словарь
                sorted_dct_value_lst = dict(sorted(dct_value_list.items()))  # сортируем словарь
                # создаем датафрейм
                temp_df = pd.DataFrame(list(sorted_dct_value_lst.items()), columns=['Показатель', 'Значение'])
                dct_list_columns[name_lst_column] = temp_df

            # Создаем датафреймы для подтверждения цифр
            for key, lst in dct_values_in_list_columns.items():
                for value in lst:
                    temp_list_df = main_df[main_df[key].str.contains(value)]
                    name_sheet = key.replace('Список_', '')

                    dct_df_list_in_columns[f'{name_sheet}_{value}'] = temp_list_df

                # Сохраняем
            if len(dct_df_list_in_columns) >= 252:
                lst_dct = list(dct_df_list_in_columns.items()) # превращаем в список кортежей
                start_threshold = 0
                end_threshold = 252
                count = 1
                while end_threshold <= len(lst_dct) + 1:
                    big_dct = dict(lst_dct[start_threshold:end_threshold])
                    list_columns_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                    list_columns_report_wb = del_sheet(list_columns_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                    list_columns_report_wb.save(f'{path_list_file}/Данные по сводам Списков {count}.xlsx')
                    count += 1
                    start_threshold += 252
                    end_threshold += 252
                    # контролируем диапазон
                    if end_threshold > len(lst_dct):
                        big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                        list_columns_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                        list_columns_report_wb = del_sheet(list_columns_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                        list_columns_report_wb.save(f'{path_list_file}/Данные по сводам Списков {count}.xlsx')

            else:
                list_columns_report_wb = write_df_big_dct_to_excel(dct_df_list_in_columns, write_index=False)
                list_columns_report_wb = del_sheet(list_columns_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                list_columns_report_wb.save(f'{path_list_file}/Данные по сводам Списков {current_time}.xlsx')

            # создаем срезы
            for name_column, prefix_file in dct_list_save_name.items():
                create_svod_list(main_df.copy(), name_column, prefix_file, lst_list_name_columns,
                                     path_list_file, current_time,dct_values_in_list_columns)

        # Сохраняем общий файл
        main_wb = write_df_to_excel({'Общий список': main_df}, write_index=False)
        main_wb = del_sheet(main_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        main_wb.save(f'{path_end_folder}/Общий файл от {current_time}.xlsx')

        # Ищем ошибки в персональных данных
        main_df.replace('', 'Нет статуса', inplace=True)
        check_error_in_pers_data(main_df.copy(),path_end_folder,current_time)



        # Создаем Свод по статусам
        # Заменяем название колонки Пол на Статус_Пол чтобы обработка проходила нормально
        main_df.rename(columns={'Пол': 'Статус_Пол'}, inplace=True)

        # Создаем папку для хранения сводов по статусам
        path_svod_file = f'{path_end_folder}/Своды по колонкам Статусов'  #
        if not os.path.exists(path_svod_file):
            os.makedirs(path_svod_file)


        # Создаем раскладку по колонкам статусов
        lst_status_columns = [column for column in main_df.columns if 'Статус_' in column]
        dct_status_save_name = {'Файл': 'по группам', 'Текущий_возраст': 'по возрастам', 'Год_рождения': 'по годам рождения', 'Статус_ОП': 'по ОП',
                              'Статус_Пол': 'по полам'}  # словарь для названий колонок по которым будут создаваться срезы

        for name_column, prefix_file in dct_status_save_name.items():
            create_svod_status(main_df.copy(), name_column, prefix_file, lst_status_columns,
                             path_svod_file, current_time)


        # Собираем колонки содержащие слово Статус_ и Подсчет_
        lst_status = [name_column for name_column in main_df.columns if
                      'Статус_' in name_column or 'Подсчет_' in name_column or 'Список_' in name_column]
        # Создаем датафрейм с данными по статусам
        soc_df = pd.DataFrame(columns=['Показатель', 'Значение'])  # датафрейм для сбора данных отчета
        soc_df.loc[len(soc_df)] = ['Количество учебных групп', quantity_sheets]  # добавляем количество учебных групп
        # считаем количество студентов
        quantity_study_student = main_df[main_df['Статус_Учёба'] == 'Обучается'].shape[0]  # со статусом Обучается
        quantity_academ_student = main_df[main_df['Статус_Учёба'].str.contains('Академический отпуск')].shape[
            0]
        quantity_not_status_student = main_df[main_df['Статус_Учёба'].str.contains('Нет статуса')].shape[
            0]
        quantity_except_deducted = main_df[~main_df['Статус_Учёба'].str.contains('Отчислен')].shape[
            0]  # все студенты кроме отчисленных
        if checkbox_expelled == 0:
            soc_df.loc[len(soc_df)] = ['Количество студентов (контингент)',
                                       f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student}, Не указан статус - {quantity_not_status_student}, Всего - {quantity_except_deducted} ']  # добавляем количество студентов
        elif checkbox_expelled == 1:
            # Без отчисленных и академа
            soc_df.loc[len(soc_df)] = ['Количество студентов (контингент)',
                                       f'Обучается - {quantity_study_student}, Не указан статус - {quantity_not_status_student}, Всего {quantity_except_deducted} (без студентов в академическом отпуске и отчисленных)']  # добавляем количество студентов
        else:
            # Включая отчисленных и в академе
            quantity_except = main_df[main_df['Статус_Учёба'].str.contains('Отчислен')].shape[0] # количество отчисленных
            soc_df.loc[len(soc_df)] = ['Количество студентов (контингент)',
                                       f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student}, Не указан статус - {quantity_not_status_student},Отчислено - {quantity_except}, Всего -  {len(main_df)}']  # добавляем количество студентов

        # считаем количество совершенолетних студентов
        quantity_maturity_students = len(main_df[main_df['Совершеннолетие'] == 'совершеннолетний'])
        quantity_not_maturity_students = len(main_df[main_df['Совершеннолетие'] == 'несовершеннолетний'])
        quantity_error_maturity_students = len(
            main_df[main_df['Совершеннолетие'].isin(['отрицательный возраст', 'Ошибочное значение!!!'])])
        soc_df.loc[len(soc_df)] = ['Возраст',
                                   f'Совершеннолетних - {quantity_maturity_students}, Несовершеннолетних - {quantity_not_maturity_students}, Неправильная дата рождения - {quantity_error_maturity_students}, Всего {len(main_df)} ']

        # Распределение по СПО-1
        header_spo = pd.DataFrame(columns=['Показатель', 'Значение'],
                                  data=[['Статус_СПО-1', None]])  # создаем строку с заголовком
        df_svod_by_SPO1 = main_df.groupby(['СПО_Один']).agg({'ФИО': 'count'})
        df_svod_by_SPO1 = df_svod_by_SPO1.reset_index()
        df_svod_by_SPO1.columns = ['Показатель', 'Значение']
        soc_df = pd.concat([soc_df, header_spo], axis=0)
        soc_df = pd.concat([soc_df, df_svod_by_SPO1], axis=0)
        for name_column in lst_status:
            if name_column == 'Статус_ОП':
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                # создаем строки с описанием
                new_value_df = create_value_str(main_df, name_column, 'Статус_Учёба',
                                                {'Обучается': 'Обучается', 'Академ': 'Академический отпуск',
                                                 'Не указан статус': 'Нет статуса'},checkbox_expelled)
            elif 'Статус_' in name_column:
                temp_counts = main_df[name_column].value_counts()  # делаем подсчет
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                new_value_df = temp_counts.to_frame().reset_index()  # создаем датафрейм с данными
                new_value_df.columns = ['Показатель', 'Значение']  # делаем одинаковыми названия колонок
                new_value_df['Показатель'] = new_value_df['Показатель'].astype(str)
                new_value_df.sort_values(by='Показатель', inplace=True)
            elif 'Подсчет_' in name_column:
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                main_df[name_column] = main_df[name_column].apply(convert_number)
                temp_desccribe = main_df[name_column].describe()
                sum_column = main_df[name_column].sum()
                _dct_describe = temp_desccribe.to_dict()
                dct_describe = {'Среднее': round(_dct_describe['mean'], 2), 'Сумма': round(sum_column, 2),
                                'Медиана': _dct_describe['50%'],
                                'Минимум': _dct_describe['min'], 'Максимум': _dct_describe['max'],
                                'Количество': _dct_describe['count'], }
                new_value_df = pd.DataFrame(list(dct_describe.items()), columns=['Показатель', 'Значение'])
            elif 'Список_' in name_column:
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                new_value_df = dct_list_columns[name_column]

            new_part_df = pd.concat([new_part_df, new_value_df], axis=0)  # соединяем
            soc_df = pd.concat([soc_df, new_part_df], axis=0)
        # Создаем раскладку по группам
        new_group_header_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[['Статус_студентов по группам', None]])  # создаем строку с заголовком
        new_group_df = create_value_str(main_df, 'Группа', 'Статус_Учёба',
                                        {'Обучается': 'Обучается', 'Академ': 'Академический отпуск',
                                         'Не указан статус': 'Нет статуса'},checkbox_expelled)
        new_group_header_df = pd.concat([new_group_header_df, new_group_df], axis=0)

        soc_df = pd.concat([soc_df, new_group_header_df], axis=0)

        soc_wb = write_df_to_excel({'Свод по статусам': soc_df}, write_index=False)
        soc_wb = del_sheet(soc_wb, ['Sheet', 'Sheet1', 'Для подсчета'])

        column_number = 0  # номер колонки в которой ищем слово Статус_
        # Создаем  стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in soc_wb['Свод по статусам'].iter_rows(min_row=1, max_row=soc_wb['Свод по статусам'].max_row,
                                                        min_col=column_number,
                                                        max_col=column_number):  # Перебираем строки
            if 'Статус_' in str(row[column_number].value) or 'Подсчет_' in str(
                    row[column_number].value) or 'Список_' in str(
                    row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill
        soc_wb.save(f'{path_end_folder}/Сводка от {current_time}.xlsx')

        # Создаем файл excel в котороым будет находится отчет
        wb = openpyxl.Workbook()

        # Проверяем наличие возможных дубликатов ,котороые могут получиться если обрезать по 30 символов
        lst_length_column = [column[:30] for column in main_df.columns]
        check_dupl_length = [k for k, v in Counter(lst_length_column).items() if v > 1]

        # проверяем наличие объединенных ячеек
        check_merge = [column for column in main_df.columns if 'Unnamed' in column]
        # если есть хоть один Unnamed то просто заменяем названия колонок на Колонка №цифра
        if check_merge or check_dupl_length:
            main_df.columns = [f'Колонка №{i}' for i in range(1, main_df.shape[1] + 1)]
        # очищаем названия колонок от символов */\ []''
        # Создаем регулярное выражение
        pattern_symbols = re.compile(r"[/*'\[\]/\\]")
        clean_main_df_columns = [re.sub(pattern_symbols, '', column) for column in main_df.columns]
        main_df.columns = clean_main_df_columns

        # Добавляем столбец для облегчения подсчета по категориям
        main_df['Для подсчета'] = 1

        # Создаем листы
        for idx, name_column in enumerate(main_df.columns):
            # Делаем короткое название не более 30 символов
            wb.create_sheet(title=name_column[:30], index=idx)

        for idx, name_column in enumerate(main_df.columns):
            group_main_df = main_df.astype({name_column: str}).groupby([name_column]).agg({'Для подсчета': 'sum'})
            group_main_df.columns = ['Количество']
            # Сортируем по убыванию
            group_main_df.sort_values(by=['Количество'], inplace=True, ascending=False)

            for r in dataframe_to_rows(group_main_df, index=True, header=True):
                if len(r) != 1:
                    wb[name_column[:30]].append(r)
            wb[name_column[:30]].column_dimensions['A'].width = 50

        # Удаляем листы
        wb = del_sheet(wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        # Сохраняем итоговый файл
        wb.save(f'{path_end_folder}/Свод по каждой колонке таблицы от {current_time}.xlsx')

        # проверяем на наличие ошибок
        if error_df.shape[0] != 0:
            messagebox.showwarning('Деметра Отчеты социальный паспорт студента',
                                   f'Обнаружены ошибки в файлах с данными.\n'
                                   f'Проверьте файл Ошибки в файле')
    except FileNotFoundError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    except NotColumn:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Проверьте названия колонок в первом листе эталонного файла, для работы программы\n'
                             f' требуюется наличие колонок: {";".join(etallon_always_cols)}'
                             )
    except NotGoodSheet:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Заголовки ни одного листа не соответствуют эталонному файлу,\n'
                             f'Откройте файл с ошибками и устраните проблему'
                             )
    except ExceedingQuantity:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Количество групп или вариантов в колонках начинающихся с Список_ превышает 253 !\n'
                             f'Программа не может создать больше 253 листов в файле xlsx'
                             f'Сократите количество обрабатываемых значений')
    except BadEtalonFile:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Эталонный файл поврежден, скачайте заново или пересохраните его с помощью Excel'
                             )
    else:
        messagebox.showinfo('Деметра Отчеты социальный паспорт студента', 'Данные успешно обработаны')


if __name__ == '__main__':
    main_etalon_file = 'data/Эталон.xlsx'
    # main_etalon_file = 'data/Эталон подсчет.xlsx'
    main_data_folder = 'data/Данные'
    main_egisso_params = 'data/Параметры ЕГИССО.xlsx'
    main_end_folder = 'data/Результат'
    main_checkbox_expelled = 0
    main_raw_date = '05.09.2024'

    create_social_report(main_etalon_file, main_data_folder, main_egisso_params, main_end_folder,
                         main_checkbox_expelled, main_raw_date)

    print('Lindy Booth !!!')

