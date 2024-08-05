"""
Вспомогательные функции
"""
import pandas as pd
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


def write_df_to_excel(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

    return wb


def write_df_to_excel_report_brit(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel отчета по стандарту БРИТ
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1
        column_number = 0  # номер колонки в которой ищем слово Статус_
        # Создаем  стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                            min_col=column_number, max_col=df.shape[1] + 1):  # Перебираем строки
            if 'Итого' in str(row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill

    return wb


def write_df_to_excel_expired_docs(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        # Красим в зависимости от цвета
        last_column_number = wb.active.max_column  # номер колонки в которой ищем слово Осталось дней
        # Создаем  стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        # Месяц
        fill_week = PatternFill(fill_type='solid', fgColor='ff0000')  # Красный цвет
        fill_two_week = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        fill_month = PatternFill(fill_type='solid', fgColor='ffff00')  # Желтый цвет
        for row in wb[name_sheet].iter_rows(min_row=2, max_row=wb[name_sheet].max_row,
                                            min_col=0,
                                            max_col=last_column_number):  # Перебираем строки
            if int(row[last_column_number-1].value) <= 7:
                for cell in row:
                    cell.font = font
                    cell.fill = fill_week
            elif int(row[last_column_number-1].value) <= 14:
                for cell in row:
                    cell.font = font
                    cell.fill = fill_two_week
            else:
                for cell in row:
                    cell.font = font
                    cell.fill = fill_month


        count_index += 1

    return wb


def del_sheet(wb: openpyxl.Workbook, lst_name_sheet: list) -> openpyxl.Workbook:
    """
    Функция для удаления лишних листов из файла
    :param wb: объект таблицы
    :param lst_name_sheet: список удаляемых листов
    :return: объект таблицы без удаленных листов
    """
    for del_sheet in lst_name_sheet:
        if del_sheet in wb.sheetnames:
            del wb[del_sheet]

    return wb
