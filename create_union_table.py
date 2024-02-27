"""
Скрипт для создания общего файла в котором каждый список был бы отдельным листом
"""
from support_functions import *
import pandas as pd
import openpyxl
from copy import copy
import time
from collections import Counter
import re
import os
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None


def merge_table(etalon_file:str, folder_update_file:str, result_folder:str)->None:
    """
    Функция для оздания новой таблицы на основании эталона из разных файлов
    :param etalon_file: файл по образу которого нужно общий файл
    :param folder_update_file: папка в которой лежат файлы для обновления
    :param result_folder: папка в которой будет находится итоговый файл
    """
    wb = openpyxl.load_workbook(etalon_file) # загружаем эталонный файл
    print(wb.sheetnames)
    main_sheet = wb.sheetnames[0] # получаем название первого листа с которым и будем сравнивать новые файлы
    main_df = pd.read_excel(etalon_file,sheet_name=main_sheet) # загружаем датафрейм чтобы получить эталонные колонки
    etalon_cols = set(main_df.columns) # эталонные колонки
    error_df = pd.DataFrame(columns=['Название файла','Название листа','Значение ошибки','Описание ошибки'])  # датафрейм для ошибок
    for file in os.listdir(folder_update_file):
        if not file.startswith('~$') and not file.endswith('.xlsx'):
            name_file = file.split('.xls')[0]
            temp_error_df = pd.DataFrame(data=[[f'{name_file}','', '',
                                                'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                         columns=['Название файла', 'Строка или колонка с ошибкой',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
            continue
        if not file.startswith('~$') and file.endswith('.xlsx'):
            name_file = file.split('.xlsx')[0]
            print(name_file)
            temp_wb = openpyxl.load_workbook(f'{folder_update_file}/{file}') # открываем
            lst_sheets_temp_wb = temp_wb.sheetnames # получаем список листов в файле
            for name_sheet in lst_sheets_temp_wb:
                if name_sheet != 'Данные для выпадающих списков': # отбрасываем лист с даннными выпадающих списков
                    temp_df = pd.read_excel(f'{folder_update_file}/{file}',sheet_name=name_sheet) # получаем колонки которые есть на листе
                    diff_cols = set(temp_df.columns).difference(etalon_cols)
                    if len(diff_cols) != 0:
                        temp_error_df = pd.DataFrame(data=[[f'{name_file}', f'{name_sheet}', f'{";".join(diff_cols)}',
                                                            'В файле на указанном листе найдены лишние или отличающиеся колонки по сравнению с эталоном. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                                     columns=['Название файла','Название листа', 'Значение ошибки',
                                                              'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue # не обрабатываем лист где найдены ошибки
                    target_sheet = wb.create_sheet(name_sheet) # Создаем лист в итоговом файле
                    for row in temp_wb[name_sheet].iter_rows(): # копируем данные
                        for cell in row:
                            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.alignment = copy(cell.alignment)
                                new_cell.protection = copy(cell.protection)




    del wb[main_sheet] # Удаляем эталонный лист
    t = time.localtime()  # получаем текущее время
    current_time = time.strftime('%H_%M_%S', t)
    current_date = time.strftime('%d_%m_%Y', t)


    wb.save(f'{result_folder}/Cвод за месяц от {current_date}.xlsx')


    # Создаем документ
    error_wb = openpyxl.Workbook()
    for r in dataframe_to_rows(error_df, index=False, header=True):
        error_wb['Sheet'].append(r)

    error_wb['Sheet'].column_dimensions['A'].width = 50
    error_wb['Sheet'].column_dimensions['B'].width = 40
    error_wb['Sheet'].column_dimensions['C'].width = 50
    error_wb.save(f'{result_folder}/ОШИБКИ Сбор данных групп от {current_time}.xlsx')
    error_wb.close()



if __name__=='__main__':
    main_etalon_file = 'data/Таблица для заполнения социального паспорта студентов.xlsx'
    main_folder_update = 'data/27.02'
    main_folder_result = 'data/Результат'
    merge_table(main_etalon_file,main_folder_update,main_folder_result)

    print('Lindy Booth')




