"""
Скрипт для добавления новых колонок в файлы с данными
"""
from tkinter import messagebox

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import time
import datetime
from collections import Counter
import re
import os
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import sys
import locale

class BadEtalonFile(Exception):
    """
    Исключение для обработки случая когда эталонный файл поврежден и его нельзя открыть
    """
    pass

class NotColumn(Exception):
    """
    Исключение для обработки случая когда отсутствуют нужные колонки
    """
    pass

def set_rus_locale():
    """
    Функция чтобы можно было извлечь русские названия месяцев
    """
    locale.setlocale(
        locale.LC_ALL,
        'rus_rus' if sys.platform == 'win32' else 'ru_RU.UTF-8')




def add_new_columns(etalon_file:str,data_folder:str,path_end_folder:str):
    """
    Функция для приведения в соответствие эталону путем переименования или добавления новых колонок
    :param etalon_file: эталонный файл
    :param data_folder: папка с файлами с данными
    :param path_end_folder: конечная папка
    :return:
    """
    try:
        set_rus_locale()  # устанавливаем русскую локаль что категоризация по месяцам работала
        # обязательные колонки
        name_columns_set = {'ФИО', 'Дата_рождения', 'Статус_ОП', 'Статус_Бюджет', 'Статус_Общежитие', 'Статус_Учёба',
                            'Статус_Всеобуч',
                            'Статус_Соц_стипендия', 'Статус_Соц_положение_семьи',
                            'СНИЛС', 'Пол', 'Серия_паспорта', 'Номер_паспорта', 'Дата_выдачи_паспорта', 'Кем_выдан',
                            'Статус_Питание',
                            'Статус_Состав_семьи', 'Статус_Уровень_здоровья', 'Статус_Сиротство',
                            'Статус_Место_регистрации', 'Статус_Студенческая_семья',
                            'Статус_Воинский_учет', 'Статус_Родитель_СВО', 'Статус_Участник_СВО',
                            'Статус_ПДН', 'Статус_КДН', 'Статус_Внутр_учет', 'Статус_Самовольный_уход', 'Статус_Выпуск'}
        error_df = pd.DataFrame(
            columns=['Название файла', 'Название листа', 'Значение ошибки', 'Описание ошибки'])  # датафрейм для ошибок
        try:
            wb = openpyxl.load_workbook(etalon_file)  # загружаем эталонный файл
            quantity_sheets = 0  # считаем количество групп
            main_sheet = wb.sheetnames[0]  # получаем название первого листа с которым и будем сравнивать новые файлы
            main_df = pd.read_excel(etalon_file, sheet_name=main_sheet,
                                    nrows=0)  # загружаем датафрейм чтобы получить эталонные колонки
        except:
            raise BadEtalonFile

        # Проверяем на обязательные колонки
        etallon_always_cols = name_columns_set.difference(set(main_df.columns))
        if len(etallon_always_cols) != 0:
            raise NotColumn
        etalon_cols = set(main_df.columns)  # эталонные колонки

        for idx, file in enumerate(os.listdir(data_folder)):
            if not file.startswith('~$') and not file.endswith('.xlsx'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '', '',
                                                    'Расширение файла НЕ XLSX! Программа обрабатывает только XLSX ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(f'Файл: {name_file}')
                try:
                    temp_wb = openpyxl.load_workbook(f'{data_folder}/{file}')  # открываем
                except:
                    temp_error_df = pd.DataFrame(
                        data=[[f'{name_file}', f'', f'',
                               'Не удалось обработать файл.']],
                        columns=['Название файла', 'Название листа', 'Значение ошибки',
                                 'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue  # не обрабатываем лист, где найдены ошибки
                lst_sheets_temp_wb = temp_wb.sheetnames  # получаем список листов в файле
                for name_sheet in lst_sheets_temp_wb:
                    if name_sheet != 'Данные для выпадающих списков':  # отбрасываем лист с даннными выпадающих списков
                        try:
                            temp_df = pd.read_excel(f'{data_folder}/{file}', sheet_name=name_sheet,
                                                    dtype=str)  # получаем колонки которые есть на листе
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'', f'',
                                       'Не удалось обработать файл.']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки
                        # Проверяем на обязательные колонки
                        always_cols = name_columns_set.difference(set(temp_df.columns))
                        if len(always_cols) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(always_cols)}',
                                       'В файле на указанном листе не найдены указанные обязательные колонки. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки
                        # Проверяем на наличие колонок без названия Unnamed
                        unnamed_lst = [f'{idx} колонка не имеет названия' for idx, name_column in
                                       enumerate(temp_df.columns, 1) if 'Unnamed' in name_column]
                        if len(unnamed_lst) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(unnamed_lst)}',
                                       'В файле на указанном листе найдена(ы) колонка(и) у которых нет названия. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки

                        # Проверяем порядок колонок
                        order_main_columns = list(main_df.columns)  # порядок колонок эталонного файла
                        order_temp_df_columns = list(temp_df.columns)  # порядок колонок проверяемого файла
                        error_order_lst = []  # список для несовпадающих пар
                        # Сравниваем попарно колонки
                        for main, temp in zip(order_main_columns, order_temp_df_columns):
                            if main != temp:
                                error_order_lst.append(f'На месте колонки {main} находится колонка {temp}')
                        if len(error_order_lst) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(error_order_lst)}',
                                       'Неправильный порядок колонок по сравнению с эталоном. Приведите порядок колонок в соответствии с порядком в эталоне. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки




        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        error_df.to_excel(f'{path_end_folder}/Ошибки от {current_time}.xlsx', index=False)
    except UnicodeError:
        print('fdsf')







if __name__ == '__main__':
    main_etalon_file = 'data/Эталон.xlsx'
    # main_etalon_file = 'data/Эталон подсчет.xlsx'
    main_data_folder = 'data/Данные'
    main_end_folder = 'data/Результат'
    # main_checkbox_expelled = 1

    add_new_columns(main_etalon_file, main_data_folder, main_end_folder,
                         )

    print('Lindy Booth !!!')





