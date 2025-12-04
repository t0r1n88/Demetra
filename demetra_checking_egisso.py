"""
Скрипт для обработки и нахождения ошибок в файлах ЕГИССО
"""
import numpy as np

from demetra_support_functions import write_df_to_excel_cheking_egisso,del_sheet,convert_to_date_egisso_cheking,create_doc_convert_date_egisso_cheking,convert_to_date_start_finish_egisso_cheking,write_df_error_egisso_to_excel # вспомогательные функции
import os
import pandas as pd
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
import time
from datetime import datetime
import re
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',)



class NotFile(Exception):
    """
    Обработка случаев когда нет файлов в папке
    """
    pass

class BadOrderCols(Exception):
    """
    Исключение для обработки случая когда колонки не совпадают
    """
    pass


class NotRecColsLMSZ(Exception):
    """
    Обработка случаев когда нет обязательных колонок в файле с мерами ЛСМЗ
    """
    pass


def strip_if_string(value):
    """Убирает пробелы вокруг строк"""
    if isinstance(value, str):
        return value.strip()
    return value

def delete_semicolon(value):
    """Убирает точку с запятой"""
    if isinstance(value, str):
        return value.replace(';',' ')
    return value


def drop_space_symbols(value:str):
    """
    Функция для замены пробельных символов в строке
    :param value:
    """
    if 'Ошибка' in value:
        return value

    result = re.sub(r'\s','',value)
    return result


def processing_rectype(value:str):
    """
    Функция для проверки колонки RecType
    :param value:
    """
    if 'Ошибка' in value:
        return value

    if value in ('Fact','Reason','Criteria','AlterationBasedOnRecalculation',
                 'AlterationBasedOnDeathOrMovingRecalculation','AlterationBasedOnPreviousErrors',
                 'TerminationBasedOnRecalculation','TerminationBasedOnIneligibility',
                 'TerminationBasedOnEndOfRight','TerminationBasedOnEndOfValidityPeriod',
                 'DuplicatesInvalidation','ErrorsInvalidation'):
        return value
    else:
        return f'Ошибка: значение {value} отличается от допустимого'




def check_symbols_uuid(value:str):
    """
    Функция для проверки символов которые есть в строке
    :param value:
    """
    if 'Ошибка' in value:
        return value

    pattern = r'^[a-z0-9-]{36}$'
    result = re.fullmatch(pattern,value)
    if result:
        return value
    else:
        return f'Ошибка: в значении {value} найдены символы отличающиеся от маленьких латинских букв, цифр и тире или в значении не 36 символов.'


def processing_assignment_fact_uuid(row:pd.Series):
    """
    Функция для обработки assignmentFactUuid
    :param row: значение RecType и assignmentFactUuid
    """

    rec_type, fact_uid = row.tolist() # распаковываем
    if 'Ошибка' in rec_type:
        return fact_uid

    if rec_type in ('AlterationBasedOnRecalculation',
                 'AlterationBasedOnDeathOrMovingRecalculation','AlterationBasedOnPreviousErrors',
                 'TerminationBasedOnRecalculation','TerminationBasedOnIneligibility',
                 'TerminationBasedOnEndOfRight','TerminationBasedOnEndOfValidityPeriod',
                 'DuplicatesInvalidation','ErrorsInvalidation'):
        if pd.isna(fact_uid):
            return f'Ошибка: assignmentFactUuid не заполнен несмотря на то что указан RecType отличный от Fact, Reason, Criteria'

        value = check_symbols_uuid(fact_uid)
        return value
    else:
        return fact_uid




    series_doc = series_doc.replace(' ','') # очищаем от пробелов
    # если свидетельство о рождении
    if type_doc == '05':
        series_doc = series_doc.upper() # делаем заглавными
        pattern = r'^[IVXLCDM]{1,5}[-|s][А-ЯЁ]{2}$'
        result = re.fullmatch(pattern,series_doc)
        if result:
            return series_doc
        else:
            return f'Ошибка: {series_doc} серия свидетельства должна соответствовать следующим требованиям 2 римские цифры, дефис и две заглавные русские буквы'






def check_symbols_onmsz(value:str):
    """
    Функция для проверки символов которые есть в строке
    :param value:
    """
    if 'Ошибка' in value:
        return value

    pattern = r'^[0-9]{4}\.[0-9]{6}$'
    result = re.fullmatch(pattern, value)
    if result:
        return value
    else:
        return f'Ошибка: значение {value} отличается от формата 0000.000000'


def processing_snils(value):
    """
    Функция для обработки СНИЛС
    :param value:
    """
    if 'Ошибка' in value:
        return value

    result = re.findall(r'\d', value)
    if len(result) == 11:
        # проверяем на лидирующий ноль
        out_str = ''.join(result)
        return out_str

    else:
        return f'Ошибка: В СНИЛС должно быть 11 цифр а в ячейке {len(result)} цифр(ы). В ячейке указано - {value}'


def processing_fio(value):
    """
    Функция для обработки Фамилии или имени
    :param value:
    """
    if 'Ошибка' in value:
        return value

    value = re.sub(r'[^\s\w-]', '', value) # очищаем от всего кроме русских букв, пробела и тире
    value = re.sub(r'\s+', ' ', value)  # заменяем пробельные символы на один пробел

    # pattern = r'^[А-ЯЁа-яё]{0,30}(( |-)([А-ЯЁ][а-яё]{0,30})){0,2}$'
    pattern = r'^[А-ЯЁа-яё]{0,30}( |-)?([А-ЯЁа-яё]{0,30})?$'
    result = re.fullmatch(pattern, value)
    if result:
        if len(value) >100:
            return f'Ошибка: в значении {value} больше 100 символов'
        return f'{value[0].upper()}{value[1:]}'
    else:
        return f'Ошибка: в значение {value}. Допустимы только буквы русского алфавита,дефис, пробел. Возможно лишний пробел рядом с дефисом или вместо русской буквы случайно записана английская. Например c-с или o-о'


def processing_patronymic(value):
    """
    Для обработки отчества
    :param value:
    """
    if isinstance(value, str):
        value = re.sub(r'[^\s\w-]', '', value)  # очищаем от всего кроме русских букв, пробела и тире
        value = re.sub(r'\s+', ' ', value)  # заменяем пробельные символы на один пробел


        # pattern = r'^[А-ЯЁа-яё]{0,30}(( |-)([А-ЯЁ][а-яё]{0,30})){0,2}$'
        pattern = r'^[А-ЯЁа-яё]{0,30}( |-)?([А-ЯЁа-яё]{0,30})?$'

        result = re.fullmatch(pattern, value)
        if result:
            if len(value) > 100:
                return f'Ошибка: в значении {value} больше 100 символов'
            return f'{value[0].upper()}{value[1:]}'
        else:
            return f'Ошибка: в значение {value}. Допустимы только буквы русского алфавита,дефис, пробел. Возможно лишний пробел рядом с дефисом или вместо русской буквы случайно записана английская. Например c-с или o-о'

    else:
        return value

def processing_gender(value:str):
    """
    Функция для обработки колонки с полом
    :param value:
    """
    if 'Ошибка' in value:
        return value

    if value[0].upper() == 'М':
        return 'М'
    elif value[0].upper() == 'Ж':
        return 'Ж'
    else:
        return f'Ошибка: {value} неправильное значение'


def processing_doc_type(value:str):
    """
    Функция для обработки колонки doctype_recip
    :param value:
    """
    if 'Ошибка' in value:
        return value
    value = value.replace(' ','')
    if len(value) == 1 and value in ('1','2','3','4','5','6','7','8'):
        return f'0{value}'
    elif len(value) == 2 and value in ('01','02','03','04','05','06','07','08'):
        return value
    else:
        return f'Ошибка: {value} не входит в список допустимых 01,02,03,04,05,06,07,08'




def processing_doc_series(row:pd.Series):
    """
    Функция для обработки серии документа
    :param row: значение doctype_recip и doc_series
    """
    type_doc, series_doc = row.tolist() # распаковываем
    if 'Ошибка' in type_doc:
        return series_doc

    if type_doc in ('01','02','04','06','07','08') and pd.isna(series_doc):
        return series_doc

    elif type_doc in ('03','05') and pd.isna(series_doc):
        return f'Ошибка: не заполнена серия свидетельства о рождении или паспорта гражданина РФ'
    else:
        if type_doc in ('01','02','04','06','07','08'):
            return series_doc
        series_doc = series_doc.replace(' ','') # очищаем от пробелов
        # если свидетельство о рождении
        if type_doc == '05':
            series_doc = series_doc.upper() # делаем заглавными
            pattern = r'^[IVXLCDM]{1,5}[-|s][А-ЯЁ]{2}$'
            result = re.fullmatch(pattern,series_doc)
            if result:
                return series_doc
            else:
                return f'Ошибка: {series_doc} серия свидетельства должна соответствовать следующим требованиям 2 римские цифры, дефис и две заглавные русские буквы'
        elif type_doc == '03':
            result = re.findall(r'\d',series_doc)
            if result:
                if len(result) == 4:
                    out_str = ''.join(result)
                    return out_str
                else:
                    return f'Ошибка: {series_doc} серия паспорта должна состоять из 4 цифр'
            else:
                return f'Ошибка: {series_doc} серия паспорта должна состоять из 4 цифр'




def processing_doc_number(row:pd.Series):
    """
    Функция для обработки колонки doc_Number_recip
    :param value:
    """
    type_doc, number_doc = row.tolist() # распаковываем
    if 'Ошибка' in type_doc:
        return number_doc

    elif type_doc in ('03','05') and pd.isna(number_doc):
        return f'Ошибка: не заполнен номер свидетельства о рождении или паспорта гражданина РФ'

    else:
        if 'Ошибка' in number_doc:
            return number_doc
        if type_doc in ('01','02','04','06','07','08'):
            return number_doc


        value = number_doc.replace(' ','')
        result = re.findall(r'\d', value)
        if result:
            if len(result) == 6:
                out_str = ''.join(result)
                return out_str
            else:
                return f'Ошибка: {value} номер свидетельства или паспорта РФ должен состоять из 6 цифр'
        else:
            return f'Ошибка: {value} номер свидетельства или паспорта РФ должен состоять из 6 цифр'



def processing_doc_issuer(value:str):
    """
    Функция для обработки колонки doc_Issuer_recip
    :param value:
    """
    if 'Ошибка' in value:
        return value
    value = re.sub(r'\s+', ' ', value).strip()
    return value


def processing_usingsign(value):
    """
    Функция для обработки колонки usingSign
    :param value:
    """
    if 'Ошибка' in value:
        return value

    if value.lower() in ('да','нет'):
        return value.capitalize()
    else:
        f'Ошибка: {value} допустимые значения это Да, Нет, да, нет'



    return value




def processing_criteria(value):
    """
    Функция для обработки колонки criteria
    :param value:
    """
    if isinstance(value, str):
        if 'Ошибка' in value:
            return value
        if len(value) > 200:
            return f'Ошибка: Длина значения больше 200 символов {value}'
        return value








def processing_criteria_code(value):
    if isinstance(value, str):
        if 'Ошибка' in value:
            return value
        if len(value) > 10:
            return f'Ошибка: Длина значения больше 10 символов {value}'
        return value


def processing_form_code(value:str):
    """
    Функция для обработки колонки form_code
    :param value:
    """
    if 'Ошибка' in value:
        return value
    value = value.replace(' ','')
    if len(value) == 1 and value in ('1','2','3','4'):
        return f'0{value}'
    elif len(value) == 2 and value in ('01','02','03','04'):
        return value
    else:
        return f'Ошибка: {value} не входит в список допустимых 01,02,03,04'


def processing_amount(value):
    """
    Функция для обработки колонки amount
    :param value:
    """
    if 'Ошибка' in value:
        return value

    try:
        out_value = value.replace(' ', '')
        out_value = out_value.replace(',', '.')
        return float(out_value)
    except:
        return f'Ошибка: {value} не является числом или в числе есть пробелы'



def processing_measury_code(row:pd.Series):
    """
    Функция для обработки формы представления
    :param row: значение FormCode и measuryCode
    """

    form_code, measury_code = row.tolist() # распаковываем
    if form_code in ('02','03','04'):
        if isinstance(measury_code, str):
            measury_code = measury_code.replace(' ', '')
            if len(measury_code) == 1 and measury_code in ('1', '3', '4', '5','6','7'):
                return f'0{measury_code}'

            elif len(measury_code) == 2 and measury_code in ('01', '03', '04', '05','06','07'):
                return measury_code

            elif measury_code in ('383','796','778','744','9910','166'):
                return measury_code
            else:
                return f'Ошибка: {measury_code} допустимые значения в колонке measury_Code это 01,03,04,05,06,07,383,796,778,744,9910,166'
        else:
            return f'Ошибка: ячейка в колонке measury_Сode не заполнена хотя в колонке FormCode указано одно из значений 02,03,04'
    elif form_code == '01':
        return measury_code

    else:
        return f'Ошибка: допустимые значения для колонки FormCode это 01,02,03,04'




def processing_monetization(value):
    """
    Функция для обработки монетизации
    :param row: значение FormCode и monetization
    """

    if value in ('01','02','04'):
        return 'Нет'
    elif value == '03':
        return 'Да'
    else:
        return f'Ошибка: допустимые значения для колонки FormCode это 01,02,03,04'



def processing_equivalent_amount(row:pd.Series):
    """
    Функция для обработки формы представления
    :param row: значение FormCode и equivalent_amount
    """

    form_code, equivalent_amount = row.tolist() # распаковываем
    if form_code in ('02','03','04'):
        if isinstance(equivalent_amount, str):
            try:
                out_value = equivalent_amount.replace(' ', '')
                out_value = out_value.replace(',', '.')
                return float(out_value)
            except:
                return f'Ошибка: {equivalent_amount} не является числом или в числе есть пробелы'

    elif form_code == '01':
        return equivalent_amount

    else:
        return f'Ошибка: допустимые значения для колонки FormCode это 01,02,03,04'


def processing_kinship_type_code(row:pd.Series):
    """
    Функция для обработки серии документа
    :param row: значение Gender_reason и kinship_type_code
    """
    gender_reason, kinship_type_code = row.tolist() # распаковываем
    if pd.isna(kinship_type_code):
        return kinship_type_code
    if isinstance(kinship_type_code,str):
        if 'Ошибка' in gender_reason:
            return f'Ошибка: допустимые значения для колонки Gender_reason это М или Ж'
        kinship_type_code = kinship_type_code.replace(' ','')
        if len(kinship_type_code) == 7:
            if gender_reason == 'Ж':
                if kinship_type_code in ('1621010','3221010','0421011','0221011','0821011','0822011',
                                         '0420011','0824011','0828011','0120011','0802011','0402011','0000010','0000030',
                                         '0800031','0000000'):
                    return kinship_type_code
                else:
                    return f'Ошибка: значение {kinship_type_code} не входит в список допустимых значений для колонки kinshipTypeCode при значении в колонке Gender_reason равном Ж'
            else:
                if kinship_type_code in ('1611010','3211010','0411011','0811011','0211011','0812011',
                                         '0412011','0410011','0814011','0818011','0110011','0802011','0402011','0000010',
                                         '0000030','0800031','0000000'):
                    return kinship_type_code
                else:
                    return f'Ошибка: значение {kinship_type_code} не входит в список допустимых значений для колонки kinshipTypeCode при значении в колонке Gender_reason равном М'
        else:
            return f'Ошибка: количество цифр в {kinship_type_code} не равно 7'



def preparing_lsmz(data_lsmz:str):
    """
    Функция для создания словаря мер ЛСМЗ формата
    {Идентификатор ЛМСЗ:{Код:Значение, Наименование ЛМСЗ:Значение, КБК:Значение,Список категорий:[]
    Категория:{Идентификатор категории:{Код категории получателей:Значение, Наименование категории получателей:Значение, Дата категории:Значение}},
    Список ЛМСЗ:[],
    Список кодов ЛМСЗ:[],
    Список КБК:[],
    Список категорий получателей:[],
    Список кодов категорий: [],
    Список наименований:[],
    Список Дат:[]}
    :param data_lsmz:файл с данными
    :return: словарь
    """
    df = pd.read_excel(data_lsmz,dtype=str)
    lst_check_cols = ['Идентификатор ЛМСЗ','код ЛМСЗ','Наименование ЛМСЗ','КБК',
                      'Идентификатор категории получателей','Код категории получателей','Наименование категории получателей','Дата']

    df.columns = list(map(str.strip,df.columns))
    # Проверяем на обязательные колонки
    always_cols = set(lst_check_cols).difference(set(df.columns))
    if len(always_cols) != 0:
        raise NotRecColsLMSZ

    lst_lmzs = df['Идентификатор ЛМСЗ'].unique()
    dct_lmsz = {lmsz:{} for lmsz in lst_lmzs} # начинаем с мер, а потом добавим списки
    # Заполняем меры
    for lmsz in dct_lmsz.keys():
        lmsz_df = df[df['Идентификатор ЛМСЗ'] == lmsz]
        dct_lmsz[lmsz]['Код ЛМСЗ'] = lmsz_df['код ЛМСЗ'].tolist()[0]
        dct_lmsz[lmsz]['Наименование ЛМСЗ'] = lmsz_df['Наименование ЛМСЗ'].tolist()[0]
        dct_lmsz[lmsz]['КБК'] = lmsz_df['КБК'].tolist()[0]
        dct_lmsz[lmsz]['Дата ЛМСЗ'] = lmsz_df['Дата'].tolist()[0]
        dct_lmsz[lmsz]['Список идентификаторов категорий получателей'] = lmsz_df['Идентификатор категории получателей'].tolist()
        dct_lmsz[lmsz]['Список кодов категорий получателей'] = lmsz_df['Код категории получателей'].tolist()
        dct_lmsz[lmsz]['Список наименований категорий получателей'] = lmsz_df['Наименование категории получателей'].tolist()
        dct_lmsz[lmsz]['Словарь категорий'] = dict(zip(lmsz_df['Идентификатор категории получателей'],lmsz_df['Наименование категории получателей']))


    dct_lst_lmsz = dict() # словарь для списков

    dct_lst_lmsz['Список ЛМСЗ'] = list(df['Идентификатор ЛМСЗ'].unique())
    dct_lst_lmsz['Список кодов ЛМСЗ'] = list(df['код ЛМСЗ'].unique())
    dct_lst_lmsz['Список наименований ЛМСЗ'] = list(df['Наименование ЛМСЗ'].unique())
    dct_lst_lmsz['Список КБК'] = list(df['КБК'].unique())
    dct_lst_lmsz['Список идентификаторов категорий получателей'] = list(df['Идентификатор категории получателей'].unique())
    dct_lst_lmsz['Список кодов категорий получателей'] = list(df['Код категории получателей'].unique())
    dct_lst_lmsz['Список наименований категорий получателей'] = list(df['Наименование категории получателей'].unique())
    dct_lst_lmsz['Список дат'] = list(df['Дата'].unique())

    dct_lmsz.update(dct_lst_lmsz)

    return dct_lmsz


def check_exists_lmsz(value,dct_lsmz:dict):
    """
    Функция для проверки существования идентификатора ЛМСЗ
    :param value: проверяемое значение
    :param dct_lsmz: словарь с данными ЛМСЗ
    """
    if 'Ошибка' in value:
        return value
    if value in dct_lsmz['Список ЛМСЗ']:
        return value
    else:
        return f'Ошибка: указанный идентификатор ЛМСЗ(LMSZID) -{value} отсутствует в файле с реестром ЛМСЗ '

def check_exists_cat_lmsz(value,dct_lsmz:dict):
    """
    Функция для проверки существования идентификатора категорий получателей
    :param value: проверяемое значение
    :param dct_lsmz: словарь с данными ЛМСЗ
    """
    if 'Ошибка' in value:
        return value
    if value in dct_lsmz['Список идентификаторов категорий получателей']:
        return value
    else:
        return f'Ошибка: указанный идентификатор категории получателей(categoryID) -{value} отсутствует в файле с реестром ЛМСЗ'



def check_correct_cat_lmsz(row:pd.Series,dct_lsmz:dict):
    """
    Функция для проверки соответствия идентификатора категории идентификатору ЛМСЗ
    :param value: проверяемое значение
    :param dct_lsmz: словарь с данными ЛМСЗ
    """
    lmsz, cat_lmsz = row
    if 'Ошибка' in lmsz or 'Ошибка' in cat_lmsz:
        return f'Ошибка: не удается проверить соответствие идентификатора категории пользователя идентификатору ЛМСЗ. Из за наличия ошибки в LMSZID или указан идентификатор категории которого нет в реестре мер'
    if cat_lmsz in dct_lsmz[lmsz]['Список идентификаторов категорий получателей']:
        return cat_lmsz
    else:
        return f'Ошибка: указанный идентификатор категории получателей(categoryID) -{cat_lmsz} не относится к идентификатору ЛМСЗ(LMSZID) -{lmsz}.'


def create_name_lmsz(value,dct_lsmz:dict):
    """
    Функция для записи наименования идентификатора ЛМСЗ
    :param value: проверяемое значение
    :param dct_lsmz: словарь с данными ЛМСЗ
    """
    if 'Ошибка' in value:
        return f'Ошибка: обнаружена ошибка в колонке LMSZID'
    return dct_lsmz[value]['Наименование ЛМСЗ']

def create_name_cat_lmsz(row:pd.Series,dct_lsmz:dict):
    """
    Функция для записи наименования идентификатора категории ЛМСЗ
    :param row: строка из идентификатора ЛМСЗ и категории
    :param dct_lsmz: словарь с данными ЛМСЗ
    """
    lmsz, cat_lmsz = row
    if 'Ошибка' in lmsz or 'Ошибка' in cat_lmsz:
        return f'Ошибка: не удается проверить соответствие идентификатора категории пользователя идентификатору ЛМСЗ. Из за наличия ошибки в LMSZID или categoryID'

    return dct_lsmz[lmsz]['Словарь категорий'][cat_lmsz]


def check_mixing(value:str):
    """
    Функция для проверки слова на смешение алфавитов
    """
    # ищем буквы русского и английского алфавита
    russian_letters = re.findall(r'[а-яА-ЯёЁ]',value)
    english_letters = re.findall(r'[a-zA-Z]',value)
    # если найдены и те и те
    if russian_letters and english_letters:
        # если русских букв больше то указываем что в русском слове встречаются английские буквы
        if len(russian_letters) > len(english_letters):
            return (f'Ошибка: в слове {value} найдены английские буквы: {",".join(english_letters)}')
        elif len(russian_letters) < len(english_letters):
            # если английских букв больше то указываем что в английском слове встречаются русские буквы
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)}')
        else:
            # если букв поровну то просто выводим их список
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)} и английские буквы: {";".join(english_letters)}')
    else:
        # если слово состоит из букв одного алфавита
        return False


def find_mixing_alphabets(cell):
    """
    Функция для нахождения случаев смешения когда английские буквы используются в русском слове и наоборот
    """
    if isinstance(cell,str):
        lst_word = re.split(r'\W',cell) # делим по не буквенным символам
        lst_result = list(map(check_mixing,lst_word)) # ищем смешения
        lst_result = [value for value in lst_result if value] # отбираем найденые смешения если они есть
        if lst_result:
            return f'Ошибка: в тексте {cell} найдено смешение русского и английского: {"; ".join(lst_result)}'
        else:
            return cell
    else:
        return cell



def fix_files_egisso(data_folder:str, end_folder:str,data_lsmz:str):
    """
    Функция для проверки и исправления файлов ЕГИССО
    :param data_folder: папка с файлами которые нужно проверить
    :param end_folder: конечная папка
    :param data_lsmz: файл с перечислением мер соц поддержки
    """
    try:
        count_errors = 0
        error_df = pd.DataFrame(
            columns=['Название файла', 'Описание ошибки'])  # датафрейм для ошибок

        # Функция для проверки и создания словаря по мерам ЛСМЗ
        dct_lsmz = preparing_lsmz(data_lsmz)

        lst_files = []  # список для файлов
        for dirpath, dirnames, filenames in os.walk(data_folder):
            lst_files.extend(filenames)
        # отбираем файлы
        lst_xlsx = [file for file in lst_files if not file.startswith('~$') and (file.endswith('.xlsx') or file.endswith('.xlsm'))]
        quantity_files = len(lst_xlsx)  # считаем сколько xlsx файлов в папке

        # Обрабатываем в зависимости от количества файлов в папке
        if quantity_files == 0:
            raise NotFile
        else:
            lst_check_cols = ['RecType','assignmentFactUuid','LMSZID',
                              'categoryID','ONMSZCode','LMSZProviderCode',
                              'providerCode','SNILS_recip','FamilyName_recip',
                              'Name_recip','Patronymic_recip','Gender_recip',
                              'BirthDate_recip','doctype_recip','doc_Series_recip',
                              'doc_Number_recip','doc_IssueDate_recip','doc_Issuer_recip',
                              'SNILS_reason','FamilyName_reason','Name_reason',
                              'Patronymic_reason','Gender_reason','BirthDate_reason',
                              'kinshipTypeCode','doctype_reason','doc_Series_reason',
                              'doc_Number_reason','doc_IssueDate_reason','doc_Issuer_reason',
                              'decision_date','dateStart','dateFinish',
                              'usingSign','criteria','criteriaCode',
                              'FormCode','amount','measuryCode',
                              'monetization','content','comment',
                              'equivalentAmount'
                              ]
            # список колонок которые обязательно должны быть заполнены
            lst_required_filling = ['RecType','LMSZID','categoryID',
                                    'ONMSZCode','SNILS_recip','FamilyName_recip',
                                    'Name_recip','Gender_recip','BirthDate_recip',
                                    'doctype_recip','doc_Number_recip',
                                    'doc_IssueDate_recip','doc_Issuer_recip','SNILS_reason',
                                    'FamilyName_reason','Name_reason','Gender_reason',
                                    'BirthDate_reason','doctype_reason',
                                    'doc_Number_reason','doc_IssueDate_reason','doc_Issuer_reason',
                                    'decision_date','dateStart',
                                    'usingSign','FormCode','amount','monetization',
                                    ]
            lst_not_required_filling = ['assignmentFactUuid','LMSZProviderCode','providerCode','Patronymic_recip','kinshipTypeCode',
                                        'Patronymic_reason','criteria','criteriaCode','measuryCode','equivalentAmount',
                                        'content','comment'
                                        ]
            # Создаем общий файл
            main_df = pd.DataFrame(columns=lst_check_cols)
            main_df.insert(0,'Название файла','')

            t = time.localtime()
            current_time = time.strftime('%H_%M_%S', t)

            for dirpath, dirnames, filenames in os.walk(data_folder):
                for file in filenames:
                    if not file.startswith('~$') and (file.endswith('.xlsx') or file.endswith('.xlsm')):
                        try:
                            if file.endswith('.xlsx'):
                                name_file = file.split('.xlsx')[0].strip()
                            else:
                                name_file = file.split('.xlsm')[0].strip()
                            print(name_file)  # обрабатываемый файл
                            df = pd.read_excel(f'{dirpath}/{file}',dtype=str) # открываем файл
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}',
                                       f'Не удалось обработать файл. Возможно файл поврежден'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            count_errors += 1
                            continue

                        # Проверяем на обязательные колонки
                        always_cols = set(lst_check_cols).difference(set(df.columns))
                        if len(always_cols) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{";".join(always_cols)}',
                                       'В файле на листе с данными не найдены указанные обязательные колонки. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки

                        df = df[lst_check_cols] # отбираем только обязательные колонки

                        if len(df) == 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}',
                                       f'Файл пустой. Лист с данными должен быть первым по порядку'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            count_errors += 1
                            continue

                        # для строковых значений очищаем от пробельных символов в начале и конце
                        df = df.applymap(strip_if_string)
                        # очищаем от символа точка с запятой
                        df = df.applymap(delete_semicolon)

                        # Находим пропущенные значения в обязательных к заполнению колонках
                        df[lst_required_filling] = df[lst_required_filling].fillna('Ошибка: Ячейка не заполнена')
                        # Находим ячейки состоящие только из пробельных символов
                        # Регулярное выражение для поиска только пробелов
                        pattern_space = r'^[\s]*$'
                        # Заменяем ячейки, содержащие только пробельные символы, на нан
                        df[lst_required_filling] = df[lst_required_filling].replace(to_replace=pattern_space, value='Ошибка: Ячейка заполнена только пробельными символами', regex=True)
                        # Очищаем для необязательных колонок
                        df[lst_not_required_filling]=df[lst_not_required_filling].replace(to_replace=pattern_space, value=np.nan, regex=True)


                        """
                        Начинаем проверять каждую колонку
                        """
                        # RecType
                        df['RecType'] = df['RecType'].apply(processing_rectype)

                        # assignmentFactUuid
                        df['assignmentFactUuid'] = df[['RecType','assignmentFactUuid']].apply(lambda x: processing_assignment_fact_uuid(x),axis=1)


                        # LMSZID
                        df['LMSZID'] = df['LMSZID'].apply(drop_space_symbols) # убираем все пробельные символы
                        df['LMSZID'] = df['LMSZID'].apply(check_symbols_uuid) # проверяем на допустимые символы и длину
                        df['LMSZID'] = df['LMSZID'].apply(lambda x:check_exists_lmsz(x,dct_lsmz))  # проверяем на вхождение в актуальные ЛМСЗ


                        #categoryID
                        df['categoryID'] = df['categoryID'].apply(drop_space_symbols) # убираем все пробельные символы
                        df['categoryID'] = df['categoryID'].apply(check_symbols_uuid) # проверяем на допустимые символы и длину
                        df['categoryID'] = df['categoryID'].apply(lambda x:check_exists_cat_lmsz(x,dct_lsmz)) # проверяем на вхождение в актуальный список ЛМСЗ

                        df['categoryID'] = df[['LMSZID','categoryID']].apply(lambda x:check_correct_cat_lmsz(x,dct_lsmz),axis=1) # проверяем на соответствие ЛМСЗ

                        #ONMSZCode
                        df['ONMSZCode'] = df['ONMSZCode'].apply(drop_space_symbols) # убираем все пробельные символы
                        df['ONMSZCode'] = df['ONMSZCode'].apply(check_symbols_onmsz)  # проверяем на допустимые символы и длину

                        # SNILS_recip
                        df['SNILS_recip'] = df['SNILS_recip'].apply(processing_snils)

                        # FamilyName_recip
                        df['FamilyName_recip'] = df['FamilyName_recip'].apply(processing_fio)

                        # Name_recip
                        df['Name_recip'] = df['Name_recip'].apply(processing_fio)

                        # Patronymic_recip
                        df['Patronymic_recip'] = df['Patronymic_recip'].apply(processing_patronymic)

                        # Gender_recip
                        df['Gender_recip'] = df['Gender_recip'].apply(processing_gender)

                        current_date = datetime.now().date() # Получаем текущую дату
                        # BirthDate_recip
                        df['BirthDate_recip'] = df['BirthDate_recip'].apply(lambda x:convert_to_date_egisso_cheking(x,current_date))
                        df['BirthDate_recip'] = df['BirthDate_recip'].apply(create_doc_convert_date_egisso_cheking)

                        # doctype_recip
                        df['doctype_recip'] = df['doctype_recip'].apply(processing_doc_type)

                        # doc_Series_recip
                        df['doc_Series_recip'] = df[['doctype_recip','doc_Series_recip']].apply(lambda x:processing_doc_series(x),axis=1)

                        # doc_Number_recip
                        df['doc_Number_recip'] = df[['doctype_recip','doc_Number_recip']].apply(lambda x:processing_doc_number(x),axis=1)

                        # doc_IssueDate_recip
                        df['doc_IssueDate_recip'] = df['doc_IssueDate_recip'].apply(lambda x:convert_to_date_egisso_cheking(x,current_date))
                        df['doc_IssueDate_recip'] = df['doc_IssueDate_recip'].apply(create_doc_convert_date_egisso_cheking)

                        # doc_Issuer_recip
                        df['doc_Issuer_recip'] = df['doc_Issuer_recip'].apply(processing_doc_issuer)

                        """
                        Обработка полей reason
                        """
                        # SNILS_reason
                        df['SNILS_reason'] = df['SNILS_reason'].apply(processing_snils)

                        # FamilyName_reason
                        df['FamilyName_reason'] = df['FamilyName_reason'].apply(processing_fio)

                        # Name_reason
                        df['Name_reason'] = df['Name_reason'].apply(processing_fio)

                        # Patronymic_reason
                        df['Patronymic_reason'] = df['Patronymic_reason'].apply(processing_patronymic)

                        # Gender_reason
                        df['Gender_reason'] = df['Gender_reason'].apply(processing_gender)

                        # BirthDate_reason
                        df['BirthDate_reason'] = df['BirthDate_reason'].apply(
                            lambda x: convert_to_date_egisso_cheking(x, current_date))
                        df['BirthDate_reason'] = df['BirthDate_reason'].apply(create_doc_convert_date_egisso_cheking)

                        # kinshipTypeCode
                        df['kinshipTypeCode'] = df[['Gender_reason', 'kinshipTypeCode']].apply(
                            lambda x: processing_kinship_type_code(x), axis=1)



                        # doctype_reason
                        df['doctype_reason'] = df['doctype_reason'].apply(processing_doc_type)

                        # doc_Series_reason
                        df['doc_Series_reason'] = df[['doctype_reason', 'doc_Series_reason']].apply(
                            lambda x: processing_doc_series(x), axis=1)

                        # doc_Number_reason
                        df['doc_Number_reason'] = df[['doctype_reason','doc_Number_reason']].apply(lambda x:processing_doc_number(x),axis=1)

                        # doc_IssueDate_reason
                        df['doc_IssueDate_reason'] = df['doc_IssueDate_reason'].apply(
                            lambda x: convert_to_date_egisso_cheking(x, current_date))
                        df['doc_IssueDate_reason'] = df['doc_IssueDate_reason'].apply(
                            create_doc_convert_date_egisso_cheking)

                        # doc_Issuer_reason
                        df['doc_Issuer_reason'] = df['doc_Issuer_reason'].apply(processing_doc_issuer)

                        # decision_date
                        df['decision_date'] = df['decision_date'].apply(
                            lambda x: convert_to_date_start_finish_egisso_cheking(x))
                        df['decision_date'] = df['decision_date'].apply(create_doc_convert_date_egisso_cheking)

                        # dateStart
                        df['dateStart'] = df['dateStart'].apply(
                            lambda x: convert_to_date_start_finish_egisso_cheking(x))
                        df['dateStart'] = df['dateStart'].apply(create_doc_convert_date_egisso_cheking)

                        # dateFinish
                        df['dateFinish'] = df['dateFinish'].apply(
                            lambda x: convert_to_date_start_finish_egisso_cheking(x))
                        df['dateFinish'] = df['dateFinish'].apply(create_doc_convert_date_egisso_cheking)

                        # usingSign
                        df['usingSign'] = df['usingSign'].apply(processing_usingsign)

                        # criteria
                        df['criteria'] = df['criteria'].apply(processing_criteria)

                        # criteriaCode
                        df['criteriaCode'] = df['criteriaCode'].apply(processing_criteria_code)

                        # FormCode
                        df['FormCode'] = df['FormCode'].apply(processing_form_code)

                        # amount
                        df['amount'] = df['amount'].apply(processing_amount)

                        # measuryCode
                        df['measuryCode'] = df[['FormCode','measuryCode']].apply(lambda x:processing_measury_code(x),axis=1)

                        # monetization
                        df['monetization'] = df['FormCode'].apply(processing_monetization)

                        # equivalentAmount
                        df['equivalentAmount'] = df[['FormCode','equivalentAmount']].apply(lambda x:processing_equivalent_amount(x),axis=1)

                        # Ищем смешение английских и русских букв
                        df = df.applymap(find_mixing_alphabets)  # ищем смешения

                        # Сохраняем датафрейм с ошибками разделенными по листам в соответсвии с колонками
                        dct_sheet_error_df = dict()  # создаем словарь для хранения названия и датафрейма

                        lst_name_columns = [name_cols for name_cols in df.columns if 'Unnamed' not in name_cols] # получаем список колонок

                        for idx, value in enumerate(lst_name_columns):
                            # получаем ошибки
                            temp_df = df[df[value].astype(str).str.contains('Ошибка')] # фильтруем
                            if temp_df.shape[0] == 0:
                                continue

                            temp_df = temp_df[value].to_frame() # оставляем только одну колонку

                            temp_df.insert(0, '№ строки с ошибкой в исходном файле', list(map(lambda x: x + 2, list(temp_df.index))))
                            dct_sheet_error_df[value] = temp_df


                        # создаем пути для проверки длины файла
                        error_path_file = f'{end_folder}/{name_file}/Базовые ошибки {name_file}.xlsx'
                        fix_path_file = f'{end_folder}/{name_file}/Обработанный {name_file}.xlsx'

                        if len(error_path_file) < 260 or len(fix_path_file) < 260:
                            if not os.path.exists(f'{end_folder}/{name_file}'):
                                os.makedirs(f'{end_folder}/{name_file}')
                                # Сохраняем по папкам
                            if len(dct_sheet_error_df) != 0:
                                file_error_wb = write_df_to_excel_cheking_egisso(dct_sheet_error_df, write_index=False)
                                file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                                file_error_wb.save(f'{end_folder}/{name_file}/Базовые ошибки {name_file}.xlsx')
                            else:
                                file_error_wb = openpyxl.Workbook()
                                file_error_wb.save(f'{end_folder}/{name_file}/Ошибок НЕТ {name_file}.xlsx')

                            file_wb = write_df_error_egisso_to_excel({'Данные': df}, write_index=False)
                            file_wb = del_sheet(file_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                            file_wb.save(f'{end_folder}/{name_file}/Обработанный {name_file}.xlsx')
                        else:
                            if len(dct_sheet_error_df) != 0:
                                file_error_wb = write_df_to_excel_cheking_egisso(dct_sheet_error_df, write_index=False)
                                file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                                file_error_wb.save(f'{end_folder}/Базовые ошибки {name_file}.xlsx')
                            else:
                                file_error_wb = openpyxl.Workbook()
                                file_error_wb.save(f'{end_folder}/Ошибок нет {name_file}.xlsx')

                            file_wb = write_df_error_egisso_to_excel({'Данные': df}, write_index=False)
                            file_wb = del_sheet(file_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                            file_wb.save(f'{end_folder}/Обработанный {name_file}.xlsx')


                        #Сохраняем объединенные файлы
                        df.insert(0,'Название файла',name_file)
                        main_df = pd.concat([main_df,df])

            main_error_wb = write_df_to_excel_cheking_egisso({'Критические ошибки':error_df},write_index=False)
            main_error_wb = del_sheet(main_error_wb,['Sheet', 'Sheet1', 'Для подсчета'])
            main_error_wb.save(f'{end_folder}/Критические ошибки {current_time}.xlsx')

            # добавляем вспомогательные колонки
            main_df.insert(1,'ЛМСЗ',main_df['LMSZID'].apply(lambda x:create_name_lmsz(x,dct_lsmz)))
            main_df.insert(2,'Категория получателей',main_df[['LMSZID','categoryID']].apply(lambda x:create_name_cat_lmsz(x,dct_lsmz),axis=1))

            main_file_wb = write_df_error_egisso_to_excel({'Общий свод': main_df}, write_index=False)
            main_file_wb = del_sheet(main_file_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            main_file_wb.save(f'{end_folder}/Общий свод {current_time}.xlsx')

            # Считаем количество и дубликаты по основным колонкам.
            count_df = pd.DataFrame() # создаем датафрейм чтобы не обрабатывать лишние колонки
            count_df['ФИО получателя'] = main_df['FamilyName_recip'] + ' ' + main_df['Name_recip'] + ' ' + main_df['Patronymic_recip']
            count_df['ФИО причины'] = main_df['FamilyName_reason'] + ' ' + main_df['Name_reason'] + ' ' + main_df['Patronymic_reason']
            count_df['СНИЛС получателя'] = main_df['SNILS_recip']
            count_df['СНИЛС причины'] = main_df['SNILS_reason']
            count_df['Документ получателя'] = main_df['doc_Series_recip'] + ' ' + main_df['doc_Number_recip']
            count_df['Документ причины'] = main_df['doc_Series_reason'] + ' ' + main_df['doc_Number_reason']
            count_df['Тип документа по'] = main_df['doctype_recip']
            count_df['Тип документа пр'] = main_df['doctype_reason']


            count_df['Название файла'] = main_df['Название файла']
            count_df['ЛМСЗ'] = main_df['ЛМСЗ']
            count_df['Категория получателей'] = main_df['Категория получателей']
            count_df['Пол получателя'] = main_df['Gender_recip']
            count_df['Пол причины'] = main_df['Gender_reason']

            count_df['RecType'] = main_df['RecType']
            count_df['LMSZID'] = main_df['LMSZID']
            count_df['categoryID'] = main_df['categoryID']
            count_df['ONMSZCode'] = main_df['ONMSZCode']
            count_df['LMSZProviderCode'] = main_df['LMSZProviderCode']
            count_df['providerCode'] = main_df['providerCode']
            count_df['decision_date'] = main_df['decision_date']
            count_df['dateStart'] = main_df['dateStart']
            count_df['dateFinish'] = main_df['dateFinish']
            count_df['criteria'] = main_df['criteria']
            count_df['criteriaCode'] = main_df['criteriaCode']
            count_df['FormCode'] = main_df['FormCode']
            count_df['amount'] = main_df['amount']
            count_df['measuryCode'] = main_df['measuryCode']
            count_df['monetization'] = main_df['monetization']
            count_df['content'] = main_df['content']
            count_df['comment'] = main_df['comment']
            count_df['equivalentAmount'] = main_df['equivalentAmount']

            """
                            Поиск дубликатов
                            """
            main_dupl_df = count_df[['ФИО получателя','ФИО причины','СНИЛС получателя','СНИЛС причины','Документ получателя',
                             'Документ причины']]
            main_dupl_df = main_dupl_df.reset_index()
            main_dupl_df.drop(columns='index',inplace=True)
            dct_dupl_df = dict()  # создаем словарь для хранения названия и датафрейма
            lst_name_columns = list(main_dupl_df.columns)  # получаем список колонок
            used_name_sheet = []  # список для хранения значений которые уже были использованы
            #
            wb = xlsxwriter.Workbook(f'{end_folder}/Дубликаты в каждой колонке {current_time}.xlsx',
                                     {'constant_memory': True, 'nan_inf_to_errors': True})  # создаем файл
            for idx, value in enumerate(lst_name_columns):
                temp_df = main_dupl_df[main_dupl_df[value].duplicated(keep=False)]  # получаем дубликаты
                if temp_df.shape[0] == 0:
                    continue

                temp_df = temp_df.sort_values(by=value)
                #     # Добавляем +2 к индексу чтобы отобразить точную строку
                temp_df.insert(0, '№ строки дубликата ', list(map(lambda x: x + 2, list(temp_df.index))))
                temp_df.replace(np.nan, None, inplace=True)  # для того чтобы в пустых ячейках ничего не отображалось
                if value == 'Название файла':
                    continue
                dct_dupl_df[value] = temp_df

            for name_sheet, dupl_df in dct_dupl_df.items():
                data_lst = dupl_df.values.tolist()  # преобразуем в список
                wb_name_sheet = wb.add_worksheet(name_sheet)  # создаем лист
                used_name_sheet.append(name_sheet)  # добавляем в список использованных названий
                # Запись заголовков
                headers = list(dupl_df.columns)
                for col, header in enumerate(headers):
                    wb_name_sheet.write(0, col, header)

                # Запись данных
                for row, data_row in enumerate(data_lst):
                    for col, cell_value in enumerate(data_row):
                        wb_name_sheet.write(row + 1, col, cell_value)

            # закрываем
            wb.close()

            """
            Делаем файл для статистики
            """
            # Добавляем столбец для облегчения подсчета по категориям
            # Создаем файл excel
            wb_stat = openpyxl.Workbook()
            count_df['Для подсчета'] = 1
            # Создаем листы
            for idx, name_column in enumerate(count_df.columns):
                # Делаем короткое название не более 30 символов
                if name_column == 'Для подсчета':
                    continue
                wb_stat.create_sheet(title=name_column, index=idx)

            for idx, name_column in enumerate(count_df.columns):
                group_df = count_df.groupby([name_column]).agg({'Для подсчета': 'sum'})
                group_df.columns = ['Количество']

                # Сортируем по убыванию
                group_df.sort_values(by=['Количество'], inplace=True, ascending=False)
                group_df.loc['Итого'] = group_df['Количество'].sum()
                if name_column == 'Для подсчета':
                    continue

                for r in dataframe_to_rows(group_df, index=True, header=True):
                    if len(r) != 1:
                        wb_stat[name_column].append(r)
                wb_stat[name_column].column_dimensions['A'].width = 50

            # Удаляем листы
            del_sheet(wb_stat, ['Sheet', 'Для подсчета'])
            wb_stat.save(f'{end_folder}/Количество {current_time}.xlsx')

            """
                            Смешение русских и английских букв
                            """
            dct_mix_df = dict()
            check_word = 'найдено смешение русского и английского:'  # фраза по которой будет производится отбор
            lst_name_columns = list(main_df.columns)  # получаем список колонок
            used_name_sheet = []  # список для хранения значений которые уже были использованы
            #
            wb_mix = xlsxwriter.Workbook(
                f'{end_folder}/Смешения русских и английских букв в словах {current_time}.xlsx',
                {'constant_memory': True, 'nan_inf_to_errors': True})  # создаем файл

            for idx, value in enumerate(lst_name_columns):
                temp_df = main_df[
                    main_df[value].astype(str).str.contains(check_word)]  # получаем строки где есть сочетание
                if temp_df.shape[0] == 0:
                    continue

                short_value = value[:20]  # получаем обрезанное значение
                short_value = re.sub(r'[\r\b\n\t\[\]\'+()<> :"?*|\\/]', '_', short_value)

                if short_value in used_name_sheet:
                    short_value = f'{short_value}_{idx}'  # добавляем окончание

                temp_df = temp_df.sort_values(by=value)
                #     # Добавляем +2 к индексу чтобы отобразить точную строку
                temp_df.insert(0, '№ строки смешения ', list(map(lambda x: x + 2, list(temp_df.index))))
                temp_df.replace(np.nan, None, inplace=True)  # для того чтобы в пустых ячейках ничего не отображалось
                dct_mix_df[short_value] = temp_df

            for name_sheet, mix_df in dct_mix_df.items():
                data_lst = mix_df.values.tolist()  # преобразуем в список
                wb_name_sheet = wb_mix.add_worksheet(name_sheet)  # создаем лист
                used_name_sheet.append(name_sheet)  # добавляем в список использованных названий
                # Запись заголовков
                headers = list(mix_df.columns)
                for col, header in enumerate(headers):
                    wb_name_sheet.write(0, col, header)

                # Запись данных
                for row, data_row in enumerate(data_lst):
                    for col, cell_value in enumerate(data_row):
                        wb_name_sheet.write(row + 1, col, cell_value)

            wb_mix.close()
















    except NotFile:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'В исходной папке отсутствуют файлы Excel (с расширением xlsx)')
    except BadOrderCols:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Неправильный порядок колонок {error_order_message}')
    except NameError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Выберите папку с файлами и папку куда будет генерироваться результат')
    except FileNotFoundError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                         f'Слишком длинный путь. Выберите в качестве конечной папку в корне диска или на рабочем столе')
    except PermissionError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                         f'Закройте все файлы созданные Деметрой')
    else:
        messagebox.showinfo('Деметра Отчеты социальный паспорт студента', f'Обработка завершена.')


if __name__ == '__main__':
    main_data_folder = 'c:/Users/1/PycharmProjects/Demetra/data/ЕГИССО'
    main_end_folder = 'c:/Users/1/PycharmProjects/Demetra/data/СБОР результат'
    main_lsmz = 'c:/Users/1/PycharmProjects/Demetra/data/Реестр ЛМСЗ Бюджетные ПОО.xlsx'

    start_time = time.time()
    fix_files_egisso(main_data_folder, main_end_folder,main_lsmz)
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Время выполнения: {elapsed_time:.6f} сек.")


    print('Lindy Booth')
