"""
Скрипт для массовой проверки и исправления файлов с данными ГИР ВУ
"""
from demetra_support_functions import (write_df_to_excel_cheking_egisso, del_sheet,write_df_error_egisso_to_excel,
                                       convert_to_date_gir_vu_cheking,create_doc_convert_date_egisso_cheking,convert_to_date_egisso_cheking,convert_to_date_future_cheking)
import pandas as pd
import numpy as np
import openpyxl
import time
import os
import re
from datetime import datetime
import xlsxwriter

class NotFile(Exception):
    """
    Обработка случаев когда нет файлов в папке
    """
    pass

class BadOrderCols(Exception):
    """
    Исключение для обработки случая когда колонки не совпадают
    """
    pass


def strip_if_string(value):
    """Убирает пробелы вокруг строк"""
    if isinstance(value, str):
        return value.strip()
    return value

def delete_semicolon(value):
    """Убирает точку с запятой"""
    if isinstance(value, str):
        return value.replace(';',' ')
    return value


def drop_space_symbols(value:str):
    """
    Функция для замены пробельных символов в строке
    :param value:
    """
    if 'Ошибка' in value:
        return value

    result = re.sub(r'\s','',value)
    return result


def processing_fio(value):
    """
    Функция для обработки Фамилии или имени
    :param value:
    """
    if 'Ошибка' in value:
        return value

    value = re.sub(r'[^\s\w-]', '', value) # очищаем от всего кроме русских букв, пробела и тире
    value = re.sub(r'\s+', ' ', value)  # заменяем пробельные символы на один пробел

    # pattern = r'^[А-ЯЁа-яё]{0,30}(( |-)([А-ЯЁ][а-яё]{0,30})){0,2}$'
    pattern = r'^[А-ЯЁа-яё]{0,30}( |-)?([А-ЯЁа-яё]{0,30})?$'
    result = re.fullmatch(pattern, value)
    if result:
        if len(value) >100:
            return f'Ошибка: в значении {value} больше 100 символов'
        return f'{value[0].upper()}{value[1:].lower()}'
    else:
        return f'Ошибка: в значение {value}. Допустимы только буквы русского алфавита,дефис, пробел. Возможно лишний пробел рядом с дефисом или вместо русской буквы случайно записана английская. Например c-с или o-о'


def processing_patronymic(value):
    """
    Для обработки отчества
    :param value:
    """
    if isinstance(value, str):
        value = re.sub(r'[^\s\w-]', '', value)  # очищаем от всего кроме русских букв, пробела и тире
        value = re.sub(r'\s+', ' ', value)  # заменяем пробельные символы на один пробел


        # pattern = r'^[А-ЯЁа-яё]{0,30}(( |-)([А-ЯЁ][а-яё]{0,30})){0,2}$'
        pattern = r'^[А-ЯЁа-яё]{0,30}( |-)?([А-ЯЁа-яё]{0,30})?$'

        result = re.fullmatch(pattern, value)
        if result:
            if len(value) > 100:
                return f'Ошибка: в значении {value} больше 100 символов'
            return f'{value[0].upper()}{value[1:].lower()}'
        else:
            return f'Ошибка: в значение {value}. Допустимы только буквы русского алфавита,дефис, пробел. Возможно лишний пробел рядом с дефисом или вместо русской буквы случайно записана английская. Например c-с или o-о'

    else:
        return value

def processing_gender(value:str):
    """
    Функция для обработки колонки с полом
    :param value:
    """
    if 'Ошибка' in value:
        return value

    if value in ('0','1','2'):
        return int(value)

    if value[0].upper() == 'М':
        return 1
    elif value[0].upper() == 'Ж':
        return 2
    elif value.upper() == 'НЕ ОПРЕДЕЛЕНО':
        return 0
    else:
        return f'Ошибка: {value} неправильное значение'


def processing_passport_series(value):
    """
    Функция для обработки серии документа
    :param row: значение doctype_recip и doc_series
    """
    if 'Ошибка' in value:
        return value

    if pd.isna(value):
        return f'Ошибка: не заполнена серия паспорта гражданина РФ'
    else:
        series_doc = value.replace(' ','') # очищаем от пробелов
        # если свидетельство о рождении
        result = re.findall(r'\d',series_doc)
        if result:
            if len(result) == 4:
                out_str = ''.join(result)
                return out_str
            else:
                return f'Ошибка: {series_doc} серия паспорта должна состоять из 4 цифр'
        else:
            return f'Ошибка: {series_doc} серия паспорта должна состоять из 4 цифр'


def processing_passport_number(value):
    """
    Функция для обработки номера паспорта

    """
    if 'Ошибка' in value:
        return value

    if pd.isna(value):
        return f'Ошибка: не заполнен номер паспорта гражданина РФ'
    else:
        number_doc = value.replace(' ','') # очищаем от пробелов
        # если свидетельство о рождении
        result = re.findall(r'\d',number_doc)
        if result:
            if len(result) == 6:
                out_str = ''.join(result)
                return out_str
            else:
                return f'Ошибка: {number_doc} номер паспорта должен состоять из 6 цифр'
        else:
            return f'Ошибка: {number_doc} номер паспорта должен состоять из 6 цифр'


def processing_snils(snils):
    """
    Функция для приведения значений снилс в вид ХХХ-ХХХ-ХХХ ХХ
    """
    if pd.isna(snils):
        return f'Ошибка: не заполнен СНИЛС гражданина РФ'
    snils = str(snils)
    result = re.findall(r'\d', snils) # ищем цифры
    if len(result) == 11:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        third_group = ''.join(result[6:9])
        four_group = ''.join(result[9:11])

        out_snils = f'{first_group}-{second_group}-{third_group} {four_group}'
        return out_snils
    else:
        return f'Ошибка: В СНИЛС должно быть 11 цифр - {snils} -{len(result)} цифр'


def processing_name_prof(value):
    """
    Функция для приведения наименований профессии специальности без кода и лишних пробелов
    """
    if 'Ошибка' in value:
        return value

    value = re.sub(r'\d','',value) # очищаем от цифр
    value = re.sub(r'\.','',value) # очищаем от точек
    value = value.strip() # очищаем от пробельных символов в начале и конце
    value = re.sub(r'\s+', ' ', value)  # заменяем пробельные символы на один пробел
    return value


def processing_code_prof(value):
    """
    Функция для приведения наименований профессии специальности без кода и лишних пробелов
    """
    if 'Ошибка' in value:
        return value

    result = re.findall(r'\d', value) # ищем цифры
    if len(result) == 6:
        first_group = ''.join(result[:2])
        second_group = ''.join(result[2:4])
        third_group = ''.join(result[4:6])

        out_value = f'{first_group}.{second_group}.{third_group}'
        return out_value
    else:
        return f'Ошибка: Код профессии, специальности должен состоять из 6 цифр разделенных точкой(например 23.07.02) - {value} -{len(result)} цифр'



def processing_form(value:str):
    """
    Функция для обработки колонки с формой обучения
    :param value:
    """
    if 'Ошибка' in value:
        return value

    if value[0].upper() == 'О':
        return 'Очная'
    elif value[0].upper() == 'З':
        return 'Заочная'
    else:
        return f'Ошибка: {value} неправильное значение. Допустимые значения: Очная, Заочная'


def processing_number_course(value:str):
    """
    Функция для обработки колонки с номером курса
    :param value:
    """
    if 'Ошибка' in value:
        return value

    if value in ('1','2','3','4','5'):
        return int(value)

    else:
        return f'Ошибка: {value} неправильное значение. Допустимые значения: 1,2,3,4,5'



def processing_many_text(value):
    """
    Функция для приведения большого текста от пробельных символов в начале и конце, от лишних пробелов
    """
    if 'Ошибка' in value:
        return value

    value = value.strip() # очищаем от пробельных символов в начале и конце
    value = re.sub(r'\s+', ' ', value)  # заменяем пробельные символы на один пробел
    return value


def check_mixing(value:str):
    """
    Функция для проверки слова на смешение алфавитов
    """
    # ищем буквы русского и английского алфавита
    russian_letters = re.findall(r'[а-яА-ЯёЁ]',value)
    english_letters = re.findall(r'[a-zA-Z]',value)
    # если найдены и те и те
    if russian_letters and english_letters:
        # если русских букв больше то указываем что в русском слове встречаются английские буквы
        if len(russian_letters) > len(english_letters):
            return (f'Ошибка: в слове {value} найдены английские буквы: {",".join(english_letters)}')
        elif len(russian_letters) < len(english_letters):
            # если английских букв больше то указываем что в английском слове встречаются русские буквы
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)}')
        else:
            # если букв поровну то просто выводим их список
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)} и английские буквы: {";".join(english_letters)}')
    else:
        # если слово состоит из букв одного алфавита
        return False


def find_mixing_alphabets(cell):
    """
    Функция для нахождения случаев смешения когда английские буквы используются в русском слове и наоборот
    """
    if isinstance(cell,str):
        lst_word = re.split(r'\W',cell) # делим по не буквенным символам
        lst_result = list(map(check_mixing,lst_word)) # ищем смешения
        lst_result = [value for value in lst_result if value] # отбираем найденые смешения если они есть
        if lst_result:
            return f'Ошибка: в тексте {cell} найдено смешение русского и английского: {"; ".join(lst_result)}'
        else:
            return cell
    else:
        return cell




def fixfiles_girvu(data_folder:str, end_folder:str):
    """
    Функция для проверки и исправления файлов ГИР ВУ
    :param data_folder: папка с файлами которые нужно проверить
    :param end_folder: конечная папка
    """
    # Словарь для замен названий листов
    dct_name_sheet = {'Фамилия': 'Фамилия',
                      'Имя': 'Имя',
                      'Отчество': 'Отчество',
                      'Пол (0-не определено, 1-мужской, 2-женский)': 'Пол',
                      'Дата рождения (ДД.ММ.ГГГГ.)': 'Дата рождения',
                      'Серия паспорта гражданина РФ': 'Серия паспорта',
                      'Номер паспорта гражданина РФ': 'Номер паспорта',
                      'Дата выдачи паспорта гражданина РФ': 'Дата выдачи',
                      'СНИЛС гражданина (при наличии)': 'СНИЛС',
                      'Наименование профессии, специальности, по которой проводится обучение (для программ СПО)': 'Наименование',
                      'Код профессии, специальности, по которой проводится обучения (для программ СПО': 'Код',
                      'Форма обучения': 'Форма обучения',
                      'Номер курса': 'Номер курса',
                      'Полное наименование образовательной организации': 'Наименование',
                      'Адрес образовательной организации': 'Адрес',
                      'Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)': 'Дата поступления',
                      'Дата завершения обучения или отчисления из образовательной организации (ДД.ММ.ГГГГ.)': 'Дата завершения',
                      'ФИО':'ФИО',
                      'Паспорт':'Паспорт',
                      }


    count_errors = 0
    error_df = pd.DataFrame(
        columns=['Название файла', 'Описание ошибки'])  # датафрейм для ошибок

    lst_files = []  # список для файлов
    for dirpath, dirnames, filenames in os.walk(data_folder):
        lst_files.extend(filenames)
    # отбираем файлы
    lst_xlsx = [file for file in lst_files if not file.startswith('~$') and file.endswith('.xlsx')]
    quantity_files = len(lst_xlsx)  # считаем сколько xlsx файлов в папке

    # Обрабатываем в зависимости от количества файлов в папке
    if quantity_files == 0:
        raise NotFile
    else:
        lst_check_cols = ['Фамилия','Имя','Отчество',
                          'Пол (0-не определено, 1-мужской, 2-женский)', 'Дата рождения (ДД.ММ.ГГГГ.)',
                          'Серия паспорта гражданина РФ', 'Номер паспорта гражданина РФ', 'Дата выдачи паспорта гражданина РФ',
                          'СНИЛС гражданина (при наличии)', 'Наименование профессии, специальности, по которой проводится обучение (для программ СПО)',
                          'Код профессии, специальности, по которой проводится обучения (для программ СПО', 'Форма обучения', 'Номер курса',
                          'Полное наименование образовательной организации', 'Адрес образовательной организации', 'Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)',
                          'Дата завершения обучения или отчисления из образовательной организации (ДД.ММ.ГГГГ.)'
                          ]

        # список колонок которые обязательно должны быть заполнены
        lst_required_filling = ['Фамилия','Имя','Отчество',
                          'Пол (0-не определено, 1-мужской, 2-женский)', 'Дата рождения (ДД.ММ.ГГГГ.)',
                          'Серия паспорта гражданина РФ', 'Номер паспорта гражданина РФ', 'Дата выдачи паспорта гражданина РФ',
                          'СНИЛС гражданина (при наличии)', 'Наименование профессии, специальности, по которой проводится обучение (для программ СПО)',
                          'Код профессии, специальности, по которой проводится обучения (для программ СПО', 'Форма обучения', 'Номер курса',
                          'Полное наименование образовательной организации', 'Адрес образовательной организации', 'Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)',
                          'Дата завершения обучения или отчисления из образовательной организации (ДД.ММ.ГГГГ.)'
                                ]
        # lst_not_required_filling = [] # не требующие обязательного заполнения

        main_df = pd.DataFrame(columns=lst_check_cols)
        main_df.insert(0, 'Название файла', '')

        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        for dirpath, dirnames, filenames in os.walk(data_folder):
            for file in filenames:
                if not file.startswith('~$') and file.endswith('.xlsx'):
                    try:
                        name_file = file.split('.xlsx')[0].strip()
                        print(name_file)  # обрабатываемый файл
                        df = pd.read_excel(f'{dirpath}/{file}', dtype=str)  # открываем файл
                    except:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{name_file}',
                                   f'Не удалось обработать файл. Возможно файл поврежден'
                                   ]],
                            columns=['Название файла',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0,
                                             ignore_index=True)
                        count_errors += 1
                        continue

                    # Проверяем на обязательные колонки
                    always_cols = set(lst_check_cols).difference(set(df.columns))
                    if len(always_cols) != 0:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{name_file}', f'{";".join(always_cols)}',
                                   'В файле на листе с данными не найдены указанные обязательные колонки. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                            columns=['Название файла', 'Значение ошибки',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        continue  # не обрабатываем лист, где найдены ошибки

                    df = df[lst_check_cols]  # отбираем только обязательные колонки

                    if len(df) == 0:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{name_file}',
                                   f'Файл пустой. Лист с данными должен быть первым по порядку'
                                   ]],
                            columns=['Название файла',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0,
                                             ignore_index=True)
                        count_errors += 1
                        continue

                    # для строковых значений очищаем от пробельных символов в начале и конце
                    df = df.applymap(strip_if_string)
                    # очищаем от символа точка с запятой
                    df = df.applymap(delete_semicolon)

                    # Находим пропущенные значения в обязательных к заполнению колонках
                    df[lst_required_filling] = df[lst_required_filling].fillna('Ошибка: Ячейка не заполнена')
                    # Находим ячейки состоящие только из пробельных символов
                    # Регулярное выражение для поиска только пробелов
                    pattern_space = r'^[\s]*$'
                    # Заменяем ячейки, содержащие только пробельные символы, на нан
                    df[lst_required_filling] = df[lst_required_filling].replace(to_replace=pattern_space,
                                                                                value='Ошибка: Ячейка заполнена только пробельными символами',
                                                                                regex=True)

                    """
                    Начинаем проверять каждую колонку
                    """
                    df['Фамилия'] = df['Фамилия'].apply(processing_fio) # Фамилия
                    df['Имя'] = df['Имя'].apply(processing_fio) # Имя
                    df['Фамилия'] = df['Фамилия'].apply(processing_patronymic) # Отчество
                    # Пол
                    df['Пол (0-не определено, 1-мужской, 2-женский)'] = df['Пол (0-не определено, 1-мужской, 2-женский)'].apply(processing_gender)
                    # Дата рождения

                    current_date = datetime.now().date()  # Получаем текущую дату
                    # Дата рождения
                    df['Дата рождения (ДД.ММ.ГГГГ.)'] = df['Дата рождения (ДД.ММ.ГГГГ.)'].apply(
                        lambda x: convert_to_date_gir_vu_cheking(x, current_date))
                    df['Дата рождения (ДД.ММ.ГГГГ.)'] = df['Дата рождения (ДД.ММ.ГГГГ.)'].apply(create_doc_convert_date_egisso_cheking)

                    # Серия паспорта
                    df['Серия паспорта гражданина РФ'] = df['Серия паспорта гражданина РФ'].apply(processing_passport_series)
                    df['Номер паспорта гражданина РФ'] = df['Номер паспорта гражданина РФ'].apply(processing_passport_number)
                    # Дата выдачи паспорта
                    df['Дата выдачи паспорта гражданина РФ'] = df['Дата выдачи паспорта гражданина РФ'].apply(
                        lambda x: convert_to_date_egisso_cheking(x, current_date))
                    df['Дата выдачи паспорта гражданина РФ'] = df['Дата выдачи паспорта гражданина РФ'].apply(create_doc_convert_date_egisso_cheking)

                    # СНИЛС
                    df['СНИЛС гражданина (при наличии)'] = df['СНИЛС гражданина (при наличии)'].apply(processing_snils)
                    # Наименование
                    df['Наименование профессии, специальности, по которой проводится обучение (для программ СПО)'] = df['Наименование профессии, специальности, по которой проводится обучение (для программ СПО)'].apply(processing_name_prof)
                    # Код профессии
                    df['Код профессии, специальности, по которой проводится обучения (для программ СПО'] = df['Код профессии, специальности, по которой проводится обучения (для программ СПО'].apply(processing_code_prof)
                    # Форма обучения
                    df['Форма обучения'] = df['Форма обучения'].apply(processing_form)
                    # Курс
                    df['Номер курса'] = df['Номер курса'].apply(processing_number_course)
                    # Полное наименование
                    df['Полное наименование образовательной организации'] = df['Полное наименование образовательной организации'].apply(processing_many_text)
                    # Адрес образовательной организации
                    df['Адрес образовательной организации'] = df['Адрес образовательной организации'].apply(processing_many_text)

                    # Дата поступления
                    df['Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)'] = df['Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)'].apply(
                        lambda x: convert_to_date_egisso_cheking(x, current_date))
                    df['Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)'] = df['Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)'].apply(create_doc_convert_date_egisso_cheking)

                    # Дата завершения
                    df['Дата завершения обучения или отчисления из образовательной организации (ДД.ММ.ГГГГ.)'] = df[['Дата завершения обучения или отчисления из образовательной организации (ДД.ММ.ГГГГ.)'
                                                                                                                     ,'Дата поступления в образовательную организацию (ДД.ММ.ГГГГ)']].apply(convert_to_date_future_cheking,axis=1)

                    df['Дата завершения обучения или отчисления из образовательной организации (ДД.ММ.ГГГГ.)'] = df['Дата завершения обучения или отчисления из образовательной организации (ДД.ММ.ГГГГ.)'].apply(create_doc_convert_date_egisso_cheking)

                    # Ищем смешение английских и русских букв
                    df = df.applymap(find_mixing_alphabets)  # ищем смешения

                    # Сохраняем датафрейм с ошибками разделенными по листам в соответсвии с колонками
                    dct_sheet_error_df = dict()  # создаем словарь для хранения названия и датафрейма

                    lst_name_columns = [name_cols for name_cols in df.columns if
                                        'Unnamed' not in name_cols]  # получаем список колонок



                    for idx, value in enumerate(lst_name_columns):
                        # получаем ошибки
                        temp_df = df[df[value].astype(str).str.contains('Ошибка')]  # фильтруем
                        if temp_df.shape[0] == 0:
                            continue

                        temp_df = temp_df[value].to_frame()  # оставляем только одну колонку

                        temp_df.insert(0, '№ строки с ошибкой в исходном файле',
                                       list(map(lambda x: x + 2, list(temp_df.index))))
                        dct_sheet_error_df[dct_name_sheet[value]] = temp_df

                    # создаем пути для проверки длины файла
                    error_path_file = f'{end_folder}/{name_file}/Базовые ошибки {name_file}.xlsx'
                    fix_path_file = f'{end_folder}/{name_file}/Обработанный {name_file}.xlsx'

                    if len(error_path_file) < 260 or len(fix_path_file) < 260:
                        if not os.path.exists(f'{end_folder}/{name_file}'):
                            os.makedirs(f'{end_folder}/{name_file}')
                            # Сохраняем по папкам
                        if len(dct_sheet_error_df) != 0:
                            file_error_wb = write_df_to_excel_cheking_egisso(dct_sheet_error_df, write_index=False)
                            file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                            file_error_wb.save(f'{end_folder}/{name_file}/Базовые ошибки {name_file}.xlsx')
                        else:
                            file_error_wb = openpyxl.Workbook()
                            file_error_wb.save(f'{end_folder}/{name_file}/Ошибок НЕТ {name_file}.xlsx')

                        file_wb = write_df_error_egisso_to_excel({'Данные': df}, write_index=False)
                        file_wb = del_sheet(file_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                        file_wb.save(f'{end_folder}/{name_file}/Обработанный {name_file}.xlsx')
                    else:
                        if len(dct_sheet_error_df) != 0:
                            file_error_wb = write_df_to_excel_cheking_egisso(dct_sheet_error_df, write_index=False)
                            file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                            file_error_wb.save(f'{end_folder}/Базовые ошибки {name_file}.xlsx')
                        else:
                            file_error_wb = openpyxl.Workbook()
                            file_error_wb.save(f'{end_folder}/Ошибок нет {name_file}.xlsx')

                        file_wb = write_df_error_egisso_to_excel({'Данные': df}, write_index=False)
                        file_wb = del_sheet(file_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                        file_wb.save(f'{end_folder}/Обработанный {name_file}.xlsx')

                    # Сохраняем объединенные файлы
                    df.insert(0, 'Название файла', name_file)
                    main_df = pd.concat([main_df, df])



            main_error_wb = write_df_to_excel_cheking_egisso({'Критические ошибки':error_df},write_index=False)
            main_error_wb = del_sheet(main_error_wb,['Sheet', 'Sheet1', 'Для подсчета'])
            main_error_wb.save(f'{end_folder}/Критические ошибки {current_time}.xlsx')

            main_file_wb = write_df_error_egisso_to_excel({'Общий свод': main_df}, write_index=False)
            main_file_wb = del_sheet(main_file_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            main_file_wb.save(f'{end_folder}/Общий свод {current_time}.xlsx')


            """
            Поиск дубликатов
            """
            # Делаем общую колонку ФИО
            main_df.insert(1,'ФИО',main_df['Фамилия'] + ' ' + main_df['Имя'] + ' '+ main_df['Отчество'])
            # Делаем общую колонку серия и номер паспорта
            main_df.insert(2,'Паспорт',main_df['Серия паспорта гражданина РФ'] + ' ' + main_df['Номер паспорта гражданина РФ'])

            dct_dupl_df = dict()  # создаем словарь для хранения названия и датафрейма
            lst_name_columns = list(main_df.columns)  # получаем список колонок
            used_name_sheet = []  # список для хранения значений которые уже были использованы
            #
            wb = xlsxwriter.Workbook(f'{end_folder}/Дубликаты в каждой колонке {current_time}.xlsx',
                                     {'constant_memory': True, 'nan_inf_to_errors': True})  # создаем файл
            for idx, value in enumerate(lst_name_columns):
                temp_df = main_df[main_df[value].duplicated(keep=False)]  # получаем дубликаты
                if temp_df.shape[0] == 0:
                    continue

                temp_df = temp_df.sort_values(by=value)
                #     # Добавляем +2 к индексу чтобы отобразить точную строку
                temp_df.insert(0, '№ строки дубликата ', list(map(lambda x: x + 2, list(temp_df.index))))
                temp_df.replace(np.nan, None, inplace=True)  # для того чтобы в пустых ячейках ничего не отображалось
                if value == 'Название файла':
                    continue
                dct_dupl_df[dct_name_sheet[value]] = temp_df

            for name_sheet, dupl_df in dct_dupl_df.items():
                data_lst = dupl_df.values.tolist()  # преобразуем в список
                wb_name_sheet = wb.add_worksheet(name_sheet)  # создаем лист
                used_name_sheet.append(name_sheet)  # добавляем в список использованных названий
                # Запись заголовков
                headers = list(dupl_df.columns)
                for col, header in enumerate(headers):
                    wb_name_sheet.write(0, col, header)

                # Запись данных
                for row, data_row in enumerate(data_lst):
                    for col, cell_value in enumerate(data_row):
                        wb_name_sheet.write(row + 1, col, cell_value)

            # закрываем
            wb.close()


            """
            Смешение русских и английских букв
            """
            dct_mix_df = dict()
            check_word = 'найдено смешение русского и английского:' # фраза по которой будет производится отбор
            lst_name_columns = list(main_df.columns)  # получаем список колонок
            used_name_sheet = []  # список для хранения значений которые уже были использованы
            #
            wb_mix = xlsxwriter.Workbook(f'{end_folder}/Смешения русских и английских букв в словах {current_time}.xlsx',{'constant_memory': True,'nan_inf_to_errors': True})  # создаем файл

            for idx, value in enumerate(lst_name_columns):
                temp_df = main_df[main_df[value].astype(str).str.contains(check_word)]  # получаем строки где есть сочетание
                if temp_df.shape[0] == 0:
                    continue

                short_value = value[:20]  # получаем обрезанное значение
                short_value = re.sub(r'[\r\b\n\t\[\]\'+()<> :"?*|\\/]', '_', short_value)

                if short_value in used_name_sheet:
                    short_value = f'{short_value}_{idx}'  # добавляем окончание

                temp_df = temp_df.sort_values(by=value)
                #     # Добавляем +2 к индексу чтобы отобразить точную строку
                temp_df.insert(0, '№ строки смешения ', list(map(lambda x: x + 2, list(temp_df.index))))
                temp_df.replace(np.nan, None,inplace=True) # для того чтобы в пустых ячейках ничего не отображалось
                dct_mix_df[short_value] = temp_df

            for name_sheet, mix_df in dct_mix_df.items():
                data_lst = mix_df.values.tolist() # преобразуем в список
                wb_name_sheet = wb_mix.add_worksheet(name_sheet) # создаем лист
                used_name_sheet.append(name_sheet) # добавляем в список использованных названий
                # Запись заголовков
                headers = list(mix_df.columns)
                for col, header in enumerate(headers):
                    wb_name_sheet.write(0, col, header)

                # Запись данных
                for row, data_row in enumerate(data_lst):
                    for col, cell_value in enumerate(data_row):
                        wb_name_sheet.write(row + 1, col, cell_value)

            wb_mix.close()












if __name__ == '__main__':
    main_data_folder = 'c:/Users/1/PycharmProjects/Demetra/data/ГИР ВУ'
    main_end_folder = 'c:/Users/1/PycharmProjects/Demetra/data/Результат'

    start_time = time.time()
    fixfiles_girvu(main_data_folder, main_end_folder)
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Время выполнения: {elapsed_time:.6f} сек.")


    print('Lindy Booth')
