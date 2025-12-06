"""
Скрипт для финальной выверки данных ЕГИССО
"""
import numpy as np

from demetra_support_functions import write_df_to_excel_cheking_egisso,del_sheet,convert_to_date_egisso_cheking,create_doc_convert_date_egisso_cheking,convert_to_date_start_finish_egisso_cheking,write_df_error_egisso_to_excel # вспомогательные функции
import os
import pandas as pd
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
import time
from datetime import datetime
import re
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',)


class NotFile(Exception):
    """
    Обработка случаев когда нет файлов в папке
    """
    pass

class BadOrderCols(Exception):
    """
    Исключение для обработки случая когда колонки не совпадают
    """
    pass


class NotRecColsLMSZ(Exception):
    """
    Обработка случаев когда нет обязательных колонок в файле
    """
    pass


def processing_snils(value):
    """
    Функция для обработки СНИЛС
    :param value:
    """
    if 'Ошибка' in value:
        return value

    result = re.findall(r'\d', value)
    if len(result) == 11:
        out_str = ''.join(result)
        return out_str
    elif len(result) == 10:
        out_str = ''.join(result)
        return f'0{out_str}'

    else:
        return f'Ошибка: В СНИЛС должно быть 11 цифр а в ячейке {len(result)} цифр(ы). В ячейке указано - {value}'


def drop_space_symbols(value):
    """
    Функция для замены пробельных символов в строке
    :param value:
    """
    value = str(value)
    if 'Ошибка' in value:
        return value

    result = re.sub(r'\s','',value)
    if len(result) == 0:
        return 'Ошибка: ячейка заполнена только пробельными символами'
    else:
        return result




def comparison_snils(row:pd.Series):
    """
    Функция для сравнения двух снилс
    :param row: СНИЛС и СНИЛС лица основания
    :return: Разные СНИЛС
    """
    recip_snils, reason_snils = row
    if pd.isna(reason_snils):
        return 'Не заполнен СНИЛС лица основания'
    # Делаем строковыми
    recip_snils = str(recip_snils)
    reason_snils = str(reason_snils)
    if reason_snils == '{NULL}':
        return 'Не заполнен СНИЛС лица основания'
    # Удаляем фигурные скобки
    reason_snils = reason_snils.replace('{','')
    reason_snils = reason_snils.replace('}','')

    recip_snils = recip_snils.replace('{','')
    recip_snils = recip_snils.replace('}','')





    if recip_snils != reason_snils:
        return 'Отличаются СНИЛС'
    else:
        return 'СНИЛС одинаковые'


def final_checking_files_egisso(data_folder:str, end_folder:str):
    """
    Функция для выверки данных ЕГИССО
    :param data_folder: папка с данными
    :param end_folder: конечная папка
    """
    try:
        count_errors = 0
        error_df = pd.DataFrame(
            columns=['Название файла', 'Описание ошибки'])  # датафрейм для ошибок


        lst_files = []  # список для файлов
        for dirpath, dirnames, filenames in os.walk(data_folder):
            lst_files.extend(filenames)
        # отбираем файлы
        lst_xlsx = [file for file in lst_files if
                    not file.startswith('~$') and (file.endswith('.xlsx') or file.endswith('.xlsm'))]
        quantity_files = len(lst_xlsx)  # считаем сколько xlsx файлов в папке

        # Обрабатываем в зависимости от количества файлов в папке
        if quantity_files == 0:
            raise NotFile

        else:
            lst_check_cols = ['Наименование региона','Код региона','Код ПИ',
                              'Наименование ПИ','Код ОНМСЗ','Наименование ОНМСЗ',
                              'Код МСЗ по Классификатору','Наименование МСЗ','Код ЛМСЗ',
                              'Наименование ЛМСЗ','Период назначения С','Период назначения ПО',
                              'Дата решения','Сумма','Внутренний UUID',
                              'Внешний UUID','СНИЛС','СНИЛС лица основания',
                              ]

            lst_required_filling = ['Наименование региона','Код региона','Код ПИ',
                              'Наименование ПИ','Код ОНМСЗ','Наименование ОНМСЗ',
                              'Код МСЗ по Классификатору','Наименование МСЗ','Код ЛМСЗ',
                              'Наименование ЛМСЗ','Период назначения С','Период назначения ПО',
                              'Дата решения','Сумма','Внутренний UUID',
                              'Внешний UUID','СНИЛС',
                              ]
            # Создаем общий файл

            # Датафрейм для хранения ошибок в колонке Период назначения ПО
            date_po_df = pd.DataFrame(columns=['Название файла','Название листа','№ строки с ошибкой','Период назначения ПО'])

            # Датафрейм для хранения ошибок в колонке СНИЛС
            snils_df = pd.DataFrame(columns=['Название файла','Название листа','№ строки с ошибкой','СНИЛС'])





            t = time.localtime()
            current_time = time.strftime('%H_%M_%S', t)

            for dirpath, dirnames, filenames in os.walk(data_folder):
                for file in filenames:
                    if not file.startswith('~$') and (file.endswith('.xlsx') or file.endswith('.xlsm')):
                        if file.endswith('.xlsx'):
                            name_file = file.split('.xlsx')[0].strip()
                        else:
                            name_file = file.split('.xlsm')[0].strip()
                        print(name_file)  # обрабатываемый файл
                        # получаем список листов в файле
                        try:
                            wb_lst_sheet = openpyxl.load_workbook(f'{dirpath}/{file}',read_only=True)
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}',
                                       f'Не удалось обработать файл. Возможно файл поврежден'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            count_errors += 1
                            continue
                        lst_wb_sheets = wb_lst_sheet.sheetnames
                        wb_lst_sheet.close()

                        # Создаем выходной файл для дублей
                        out_wb = openpyxl.Workbook()
                        hand_check_wb = openpyxl.Workbook() # Файл для значений которые нужно проверить

                        wb_snils = openpyxl.Workbook() # файл для СНИЛС

                        dct_svod = dict() # словарь для хранения сводов по колонке Период ПО

                        dct_snils = dict() # словарь для хранения сводов по колонке СНИЛС


                        for idx,name_sheet in enumerate(lst_wb_sheets,1):
                            print(name_sheet)
                            df = pd.read_excel(f'{dirpath}/{file}',sheet_name=name_sheet,dtype=str)  # открываем файл

                            # Проверяем на обязательные колонки
                            always_cols = set(lst_check_cols).difference(set(df.columns))
                            if len(always_cols) != 0:
                                temp_error_df = pd.DataFrame(
                                    data=[[f'{name_file}', f'{";".join(always_cols)}',
                                           f'В файле на листе {name_sheet} не найдены указанные обязательные колонки. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                    columns=['Название файла', 'Значение ошибки',
                                             'Описание ошибки'])
                                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                                continue  # не обрабатываем лист, где найдены ошибки

                            # убедившись что все колонки есть заново считываем файл с контролируемыми типами
                            df = pd.read_excel(f'{dirpath}/{file}',sheet_name=name_sheet,
                                               dtype={'СНИЛС':'str','СНИЛС лица основания':'str',
                                                      'Сумма':'float64'})  # открываем файл
                            df = df[lst_check_cols]  # отбираем только обязательные колонки

                            if len(df) == 0:
                                temp_error_df = pd.DataFrame(
                                    data=[[f'{name_file} Лист {name_sheet}',
                                           f'Лист пустой.'
                                           ]],
                                    columns=['Название файла',
                                             'Описание ошибки'])
                                error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                     ignore_index=True)
                                count_errors += 1
                                continue

                            # Создаем датафрейм для собирания дубликатов листа
                            lst_cols_main_dupl = ['№ строки дубликата']
                            lst_cols_main_dupl.extend(lst_check_cols)

                            # Создаем датафрейм для собирания дубликатов листа
                            lst_cols_hand_check = ['№ строки ручная проверка']
                            lst_cols_hand_check.extend(lst_check_cols)

                            # Дубликаты
                            main_dupl_df = pd.DataFrame(columns=lst_cols_main_dupl)
                            out_wb.create_sheet(name_sheet,index=idx) # создаем лист
                            # Ручная проверка
                            main_hand_check_df = pd.DataFrame(columns=lst_cols_hand_check)
                            hand_check_wb.create_sheet(name_sheet,index=idx) # создаем лист


                            # Находим пропущенные значения в обязательных к заполнению колонках
                            df[lst_required_filling] = df[lst_required_filling].fillna('Ошибка: Ячейка не заполнена')
                            df['Период назначения ПО'] = df['Период назначения ПО'].apply(drop_space_symbols)
                            # Проверяем СНИЛС
                            df['СНИЛС'] = df['СНИЛС'].apply(processing_snils)

                            # Начинаем обработку
                            lst_snils = list(df['СНИЛС'].unique()) # уникальные снилсы
                            # Сохраняем датафрейм с ошибками разделенными по листам в соответсвии с колонками




                            for idx, value in enumerate(['Период назначения ПО','СНИЛС']):
                                # получаем ошибки
                                temp_df = df[df[value].astype(str).str.contains('Ошибка')]  # фильтруем
                                if temp_df.shape[0] == 0:
                                    continue

                                temp_df = temp_df[value].to_frame()  # оставляем только одну колонку

                                temp_df.insert(0, 'Название файла', name_file)
                                temp_df.insert(1, 'Название листа', name_sheet)
                                temp_df.insert(2, '№ строки с ошибкой',
                                               list(map(lambda x: x + 2, list(temp_df.index))))

                                if value == 'Период назначения ПО':
                                    date_po_df = pd.concat([date_po_df,temp_df])
                                else:
                                    snils_df = pd.concat([snils_df,temp_df])


                            """
                            Дубликаты
                            """
                            for snils in lst_snils:
                                temp_df = df[df['СНИЛС'] == snils]
                                dupl_df = temp_df[temp_df[['Код ЛМСЗ','Период назначения С','Период назначения ПО','Сумма']].duplicated(keep=False)]
                                dupl_df = dupl_df.sort_values(by='Период назначения С')
                                dupl_df.insert(0, '№ строки дубликата', list(map(lambda x: x + 2, list(dupl_df.index))))
                                if len(dupl_df) != 0:
                                    dupl_df.loc['Граница'] = ''
                                    main_dupl_df = pd.concat([main_dupl_df,dupl_df])

                                    # Ручная проверка
                                    temp_df.drop_duplicates(subset=['Код ЛМСЗ','Период назначения С','Период назначения ПО','Сумма'],inplace=True)
                                    hand_check_df = temp_df[
                                        temp_df[['Код ЛМСЗ', 'Период назначения С', 'Период назначения ПО']].duplicated(
                                            keep=False)]
                                    hand_check_df = hand_check_df.sort_values(by='Период назначения С')
                                    hand_check_df.insert(0, '№ строки ручная проверка',
                                                         list(map(lambda x: x + 2, list(hand_check_df.index))))
                                    if len(hand_check_df) != 0:
                                        hand_check_df.loc['Граница'] = ''
                                        main_hand_check_df = pd.concat([main_hand_check_df, hand_check_df])



                            df['Разные СНИЛС'] = df[['СНИЛС','СНИЛС лица основания']].apply(comparison_snils,axis=1)
                            df.insert(0, '№ строки в исходном файле',
                                                         list(map(lambda x: x + 2, list(df.index))))


                            svod_df = pd.pivot_table(df,values='Сумма',
                                                     index=['Наименование ЛМСЗ','Период назначения С'],
                                                     aggfunc='sum',
                                                     fill_value=0,
                                                     margins=True,
                                                     margins_name='Итого')
                            svod_df = svod_df.reset_index()
                            dct_svod[name_sheet] = svod_df # добавляем в словарь

                            svod_snils_df = pd.pivot_table(df,values='Сумма',
                                                     index=['СНИЛС','Период назначения С','Наименование ЛМСЗ'],
                                                     aggfunc='sum',
                                                     fill_value=0,
                                                     margins=True,
                                                     margins_name='Итого')
                            svod_snils_df = svod_snils_df.reset_index()
                            dct_snils[name_sheet] = svod_snils_df # добавляем в словарь


                            # Создаем листы для датафрейма СНИЛС
                            wb_snils.create_sheet(name_sheet,idx) # создаем лист
                            for r in dataframe_to_rows(df, index=False, header=True):
                                if len(r) != 1:
                                    wb_snils[name_sheet].append(r)





                            for r in dataframe_to_rows(main_dupl_df, index=False, header=True):
                                if len(r) != 1:
                                    out_wb[name_sheet].append(r)
                            out_wb[name_sheet].column_dimensions['A'].width = 10
                            out_wb[name_sheet].column_dimensions['K'].width = 20
                            out_wb[name_sheet].column_dimensions['M'].width = 10
                            out_wb[name_sheet].column_dimensions['N'].width = 10
                            out_wb[name_sheet].column_dimensions['P'].width = 37
                            out_wb[name_sheet].column_dimensions['Q'].width = 37
                            out_wb[name_sheet].column_dimensions['R'].width = 15


                            for r in dataframe_to_rows(main_hand_check_df, index=False, header=True):
                                if len(r) != 1:
                                    hand_check_wb[name_sheet].append(r)
                            hand_check_wb[name_sheet].column_dimensions['A'].width = 10
                            hand_check_wb[name_sheet].column_dimensions['K'].width = 20
                            hand_check_wb[name_sheet].column_dimensions['M'].width = 10
                            hand_check_wb[name_sheet].column_dimensions['N'].width = 10
                            hand_check_wb[name_sheet].column_dimensions['P'].width = 37
                            hand_check_wb[name_sheet].column_dimensions['Q'].width = 37
                            hand_check_wb[name_sheet].column_dimensions['R'].width = 15

                        # Удаляем листы
                        del_sheet(out_wb, ['Sheet'])
                        del_sheet(hand_check_wb,['Sheet'])

                        # создаем пути для проверки длины файла
                        dupl_path_file = f'{end_folder}/Дубли {name_file} {current_time}.xlsx'
                        hand_check_path_file = f'{end_folder}/Ручная проверка {name_file} {current_time}.xlsx'

                        out_wb.save(dupl_path_file)
                        hand_check_wb.save(hand_check_path_file)

                        wb_snils = del_sheet(wb_snils, ['Sheet', 'Sheet1', 'Для подсчета'])
                        wb_snils.save(f'{end_folder}/СНИЛС {name_file} {current_time}.xlsx')

                        # Сохраняем файл со сводами периоду ПО
                        with pd.ExcelWriter(f'{end_folder}/Свод Период С {name_file}_{current_time}.xlsx',engine='xlsxwriter') as writer:
                            for name_sheet, out_df in dct_svod.items():
                                out_df.to_excel(writer, sheet_name=str(name_sheet), index=False)

                        # Сохраняем файл со сводами по СНИЛС
                        with pd.ExcelWriter(f'{end_folder}/Свод СНИЛС {name_file}_{current_time}.xlsx',engine='xlsxwriter') as writer:
                            for name_sheet, out_df in dct_snils.items():
                                out_df.to_excel(writer, sheet_name=str(name_sheet), index=False)





        err_dct = {'Ошибки в структуре': error_df,'Период ПО':date_po_df,'СНИЛС':snils_df}


        #
        main_error_wb = write_df_to_excel_cheking_egisso(err_dct, write_index=False)
        main_error_wb = del_sheet(main_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        main_error_wb.save(f'{end_folder}/Критические ошибки {current_time}.xlsx')





    except NotFile:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'В исходной папке отсутствуют файлы Excel (с расширением xlsx, xlsm)')
    except NameError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Выберите папку с файлами и папку куда будет генерироваться результат')
    except FileNotFoundError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                         f'Слишком длинный путь. Выберите в качестве конечной папку в корне диска или на рабочем столе')
    except PermissionError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                         f'Закройте все файлы созданные Деметрой')
    else:
        messagebox.showinfo('Деметра Отчеты социальный паспорт студента', f'Обработка завершена.')















if __name__ == '__main__':
    main_data_folder = 'c:/Users/1/PycharmProjects/Demetra/data/Выверка ЕГИССО'
    main_end_folder = 'c:/Users/1/PycharmProjects/Demetra/data/СБОР результат'

    start_time = time.time()
    final_checking_files_egisso(main_data_folder, main_end_folder)
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Время выполнения: {elapsed_time:.6f} сек.")


    print('Lindy Booth')
