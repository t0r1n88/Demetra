"""
Вспомогательные функции
"""
import pandas as pd
import numpy as np
import datetime
import re
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

from pytrovich.detector import PetrovichGenderDetector
from pytrovich.enums import NamePart, Gender, Case
from pytrovich.maker import PetrovichDeclinationMaker

import time

class ExceedingQuantity(Exception):
    """
    Исключение для случаев когда числа уникальных значений больше 255
    """
    pass

def convert_to_date(value):
    """
    Функция для конвертации строки в текст
    :param value: значение для конвертации
    :return:
    """
    try:
        if value == 'Нет статуса':
            return None
        else:
            date_value  = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
            return date_value
    except ValueError:
        result = re.search(r'^\d{2}\.\d{2}\.\d{4}$',value)
        if result:
            return datetime.datetime.strptime(result.group(0), '%d.%m.%Y')
        else:
            return f'Некорректный формат даты - {value}'
    except:
        return None

def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        if cell is np.nan:
            return 'Не заполнено'
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except:
        return None


def capitalize_double_name(word):
    """
    Функция для того чтобы в двойных именах и фамилиях вторая часть была также с большой буквы
    """
    lst_word = word.split('-')  # сплитим по дефису
    if len(lst_word) == 1:  # если длина списка равна 1 то это не двойная фамилия и просто возвращаем слово

        return word
    elif len(lst_word) == 2:
        first_word = lst_word[0].capitalize()  # делаем первую букву слова заглавной а остальные строчными
        second_word = lst_word[1].capitalize()
        return f'{first_word}-{second_word}'
    else:
        return 'Не удалось просклонять'


def case_lastname(maker, lastname, gender, case: Case):
    """
    Функция для обработки и склонения фамилии. Это нужно для обработки случаев двойной фамилии
    """

    lst_lastname = lastname.split('-')  # сплитим по дефису

    if len(lst_lastname) == 1:  # если длина списка равна 1 то это не двойная фамилия и просто обрабатываем слово
        case_result_lastname = maker.make(NamePart.LASTNAME, gender, case, lastname)
        return case_result_lastname
    elif len(lst_lastname) == 2:
        first_lastname = lst_lastname[0].capitalize()  # делаем первую букву слова заглавной а остальные строчными
        second_lastname = lst_lastname[1].capitalize()
        # Склоняем по отдельности
        first_lastname = maker.make(NamePart.LASTNAME, gender, case, first_lastname)
        second_lastname = maker.make(NamePart.LASTNAME, gender, case, second_lastname)

        return f'{first_lastname}-{second_lastname}'


def detect_gender(detector, lastname, firstname, middlename):
    """
    Функция для определения гендера слова
    """
    #     detector = PetrovichGenderDetector() # создаем объект детектора
    try:
        gender_result = detector.detect(lastname=lastname, firstname=firstname, middlename=middlename)
        return gender_result
    except StopIteration:  # если не удалось определить то считаем что гендер андрогинный
        return Gender.ANDROGYNOUS


def decl_on_case(fio: str, case: Case) -> str:
    """
    Функция для склонения ФИО по падежам
    """
    fio = fio.strip()  # очищаем строку от пробельных символов с начала и конца
    part_fio = fio.split()  # разбиваем по пробелам создавая список где [0] это Фамилия,[1]-Имя,[2]-Отчество

    if len(part_fio) == 3:  # проверяем на длину и обрабатываем только те что имеют длину 3 во всех остальных случаях просим просклонять самостоятельно
        maker = PetrovichDeclinationMaker()  # создаем объект класса
        lastname = part_fio[0].capitalize()  # Фамилия
        firstname = part_fio[1].capitalize()  # Имя
        middlename = part_fio[2].capitalize()  # Отчество

        # Определяем гендер для корректного склонения
        detector = PetrovichGenderDetector()  # создаем объект детектора
        gender = detect_gender(detector, lastname, firstname, middlename)
        # Склоняем

        case_result_lastname = case_lastname(maker, lastname, gender, case)  # обрабатываем фамилию
        case_result_firstname = maker.make(NamePart.FIRSTNAME, gender, case, firstname)
        case_result_firstname = capitalize_double_name(case_result_firstname)  # обрабатываем случаи двойного имени
        case_result_middlename = maker.make(NamePart.MIDDLENAME, gender, case, middlename)
        # Возвращаем результат
        result_fio = f'{case_result_lastname} {case_result_firstname} {case_result_middlename}'
        return result_fio

    else:
        return 'Проверьте количество слов, должно быть 3 разделенных пробелами слова'


def create_initials(cell, checkbox, space):
    """
    Функция для создания инициалов
    """
    lst_fio = cell.split(' ')  # сплитим по пробелу
    lst_fio = [value for value in lst_fio if value] # отбрасываем варианты из за лишнего пробела
    lst_fio = list(map(str.strip,lst_fio))
    if len(lst_fio) == 3:  # проверяем на стандартный размер в 3 слова иначе ничего не меняем
        if checkbox == 'ФИ':
            if space == 'без пробела':
                # возвращаем строку вида Иванов И.И.
                return f'{lst_fio[0]} {lst_fio[1][0].upper()}.{lst_fio[2][0].upper()}.'
            else:
                # возвращаем строку с пробелом после имени Иванов И. И.
                return f'{lst_fio[0]} {lst_fio[1][0].upper()}. {lst_fio[2][0].upper()}.'

        else:
            if space == 'без пробела':
                # И.И. Иванов
                return f'{lst_fio[1][0].upper()}.{lst_fio[2][0].upper()}. {lst_fio[0]}'
            else:
                # И. И. Иванов
                return f'{lst_fio[1][0].upper()}. {lst_fio[2][0].upper()}. {lst_fio[0]}'
    else:
        return cell

def extract_fio(fio:str,code_fio:int):
    """
    Функция для получения из ячейки с ФИО отдельных Фамилии, Имени, Отчества
    :param value: ФИО
    :param code_fio: что нужно получить 0-Фамилия, 1- Имя, 2- Отчество
    :return: соответствующее значение
    """
    fio = fio.strip()  # очищаем строку от пробельных символов с начала и конца
    if len(fio) == 0:
        return 'Не заполнено ФИО'
    else:
        part_fio = fio.split()  # разбиваем по пробелам создавая список где [0] это Фамилия,[1]-Имя,[2]-Отчество
        part_fio = [value for value in part_fio if value]  # отбрасываем варианты из за лишнего пробела
        part_fio = list(map(str.strip, part_fio))

        if len(part_fio) == 3:  # проверяем на длину и обрабатываем только те что имеют длину 3
            lastname = part_fio[0].capitalize()  # Фамилия
            firstname = part_fio[1].capitalize()  # Имя
            middlename = part_fio[2].capitalize()  # Отчество

            if code_fio == 0:
                return lastname
            elif code_fio == 1:
                return firstname
            else:
                return middlename
        elif len(part_fio) == 2:  # проверяем на длину и обрабатываем только те что имеют длину 2
            lastname = part_fio[0].capitalize()  # Фамилия
            firstname = part_fio[1].capitalize()  # Имя

            if code_fio == 0:
                return lastname
            elif code_fio == 1:
                return firstname
            else:
                return 'Не найдено отчество'
        elif len(part_fio) == 1:  # проверяем на длину и обрабатываем только те что имеют длину 1
            lastname = part_fio[0].capitalize()  # Фамилия
            if code_fio == 0:
                return lastname
            elif code_fio == 1:
                return 'Не найдено имя'
            else:
                return 'Не найдено отчество'
        else:
            lastname = part_fio[0].capitalize()  # Фамилия
            firstname = part_fio[1].capitalize()  # Имя
            middlename = part_fio[2].capitalize()  # Отчество
            if code_fio == 0:
                return lastname
            elif code_fio == 1:
                return firstname
            else:
                return middlename



def declension_fio_by_case(df:pd.DataFrame):
    """
    Функция для склонения ФИО по падежам и создания инициалов
    :param df: датафрейм в который надо добавить инициалы и падежи

    :return: файл Excel в котором после колонки fio_column добавляется 29 колонок с падежами
    """
    fio_column = 'ФИО'
    df[fio_column] = df[fio_column].astype(str) # делаем на всякий случай строковой

    temp_df = pd.DataFrame()  # временный датафрейм для хранения колонок просклоненных по падежам

    # Получаем номер колонки с фио которые нужно обработать
    lst_columns = list(df.columns)  # Превращаем в список
    index_fio_column = lst_columns.index(fio_column)  # получаем индекс

    # Обрабатываем nan значения и те которые обозначены пробелом
    df[fio_column].fillna('Не заполнено', inplace=True)
    df[fio_column] = df[fio_column].apply(lambda x: x.strip())
    df[fio_column] = df[fio_column].apply(
        lambda x: x if x else 'Не заполнено')  # Если пустая строка то заменяем на значение Не заполнено

    temp_df['Родительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.GENITIVE))
    temp_df['Дательный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.DATIVE))
    temp_df['Винительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.ACCUSATIVE))
    temp_df['Творительный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.INSTRUMENTAL))
    temp_df['Предложный_падеж'] = df[fio_column].apply(lambda x: decl_on_case(x, Case.PREPOSITIONAL))
    temp_df['Фамилия_инициалы'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Инициалы_фамилия'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Фамилия_инициалы_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_пробел'] = df[fio_column].apply(lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами родительный падеж
    temp_df['Фамилия_инициалы_род_падеж'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_род_падеж_пробел'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_род_падеж'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_род_падеж_пробел'] = temp_df['Родительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами дательный падеж
    temp_df['Фамилия_инициалы_дат_падеж'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_дат_падеж_пробел'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_дат_падеж'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_дат_падеж_пробел'] = temp_df['Дательный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами винительный падеж
    temp_df['Фамилия_инициалы_вин_падеж'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_вин_падеж_пробел'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_вин_падеж'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_вин_падеж_пробел'] = temp_df['Винительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Создаем колонки для склонения фамилий с иницалами творительный падеж
    temp_df['Фамилия_инициалы_твор_падеж'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_твор_падеж_пробел'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_твор_падеж'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_твор_падеж_пробел'] = temp_df['Творительный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))
    # Создаем колонки для склонения фамилий с иницалами предложный падеж
    temp_df['Фамилия_инициалы_пред_падеж'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'без пробела'))
    temp_df['Фамилия_инициалы_пред_падеж_пробел'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ФИ', 'пробел'))
    temp_df['Инициалы_фамилия_пред_падеж'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'без пробела'))
    temp_df['Инициалы_фамилия_пред_падеж_пробел'] = temp_df['Предложный_падеж'].apply(
        lambda x: create_initials(x, 'ИФ', 'пробел'))

    # Вставляем получившиеся колонки в конец датафрейма
    # Фамилия Имя Отчество
    df['Фамилия'] = df[fio_column].apply(lambda x:extract_fio(x,0))
    df['Имя'] = df[fio_column].apply(lambda x:extract_fio(x,1))
    df['Отчество'] = df[fio_column].apply(lambda x:extract_fio(x,2))

    # Падежи
    df['Родительный_падеж'] = temp_df['Родительный_падеж']
    df['Дательный_падеж'] = temp_df['Дательный_падеж']
    df['Винительный_падеж'] = temp_df['Винительный_падеж']
    df['Творительный_падеж'] = temp_df['Творительный_падеж']
    df['Предложный_падеж'] = temp_df['Предложный_падеж']
    # Инициалы
    df['Фамилия_инициалы'] = temp_df['Фамилия_инициалы']
    df['Инициалы_фамилия'] = temp_df['Инициалы_фамилия']
    df['Фамилия_инициалы_пробел'] = temp_df['Фамилия_инициалы_пробел']
    df['Инициалы_фамилия_пробел'] = temp_df['Инициалы_фамилия_пробел']

    # Добавляем колонки с склонениями инициалов родительный падеж
    df['Фамилия_инициалы_род_падеж'] = temp_df['Фамилия_инициалы_род_падеж']
    df['Фамилия_инициалы_род_падеж_пробел'] = temp_df['Фамилия_инициалы_род_падеж_пробел']
    df['Инициалы_фамилия_род_падеж'] = temp_df['Инициалы_фамилия_род_падеж']
    df['Инициалы_фамилия_род_падеж_пробел'] = temp_df['Инициалы_фамилия_род_падеж_пробел']

    # Добавляем колонки с склонениями инициалов дательный падеж
    df['Фамилия_инициалы_дат_падеж'] = temp_df['Фамилия_инициалы_дат_падеж']
    df['Фамилия_инициалы_дат_падеж_пробел'] = temp_df['Фамилия_инициалы_дат_падеж_пробел']
    df['Инициалы_фамилия_дат_падеж'] = temp_df['Инициалы_фамилия_дат_падеж']
    df['Инициалы_фамилия_дат_падеж_пробел'] = temp_df['Инициалы_фамилия_дат_падеж_пробел']

    # Добавляем колонки с склонениями инициалов винительный падеж
    df['Фамилия_инициалы_вин_падеж'] = temp_df['Фамилия_инициалы_вин_падеж']
    df['Фамилия_инициалы_вин_падеж_пробел'] = temp_df['Фамилия_инициалы_вин_падеж_пробел']
    df['Инициалы_фамилия_вин_падеж'] = temp_df['Инициалы_фамилия_вин_падеж']
    df['Инициалы_фамилия_вин_падеж_пробел'] = temp_df['Инициалы_фамилия_вин_падеж_пробел']

    # Добавляем колонки с склонениями инициалов творительный падеж
    df['Фамилия_инициалы_твор_падеж'] = temp_df['Фамилия_инициалы_твор_падеж']
    df['Фамилия_инициалы_твор_падеж_пробел'] = temp_df['Фамилия_инициалы_твор_падеж_пробел']
    df['Инициалы_фамилия_твор_падеж'] = temp_df['Инициалы_фамилия_твор_падеж']
    df['Инициалы_фамилия_твор_падеж_пробел'] = temp_df['Инициалы_фамилия_твор_падеж_пробел']

    # Добавляем колонки с склонениями инициалов предложный падеж
    df['Фамилия_инициалы_пред_падеж'] = temp_df['Фамилия_инициалы_пред_падеж']
    df['Фамилия_инициалы_пред_падеж_пробел'] = temp_df['Фамилия_инициалы_пред_падеж_пробел']
    df['Инициалы_фамилия_пред_падеж'] = temp_df['Инициалы_фамилия_пред_падеж']
    df['Инициалы_фамилия_пред_падеж_пробел'] = temp_df['Инициалы_фамилия_пред_падеж_пробел']

    return df

def write_df_to_excel(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    if len(dct_df) >= 253:
        raise ExceedingQuantity # проверяем количество значений
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

        column_number = 0  # номер колонки
        # Создаем стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                            min_col=column_number, max_col=df.shape[1] + 1):  # Перебираем строки
            if 'Итого' in str(row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill

    return wb



def write_df_big_dct_to_excel(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    used_name_sheet = set()  # множество для хранения значений которые уже были использованы
    if len(dct_df) >= 253:
        raise ExceedingQuantity
    for name_sheet, df in dct_df.items():
        short_name_sheet = name_sheet[:20]  # получаем обрезанное значение
        short_name_sheet = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_name_sheet)
        if short_name_sheet.lower() in used_name_sheet:
            short_name_sheet = f'{short_name_sheet}_{count_index}'  # добавляем окончание

        wb.create_sheet(title=short_name_sheet, index=count_index)  # создаем лист
        used_name_sheet.add(short_name_sheet.lower()) # добавляем в список использованных названий
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[short_name_sheet].append(row)
            else:
                wb[short_name_sheet].append(row)
        if none_check:
            wb[short_name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[short_name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[short_name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

        column_number = 0  # номер колонки
        # Создаем стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[short_name_sheet].iter_rows(min_row=1, max_row=wb[short_name_sheet].max_row,
                                            min_col=column_number, max_col=df.shape[1] + 1):  # Перебираем строки
            if 'Итого' in str(row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill

    return wb


def write_df_to_excel_report_brit(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel отчета по стандарту БРИТ
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1
        column_number = 0  # номер колонки в которой ищем слово Статус_
        # Создаем  стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                            min_col=column_number, max_col=df.shape[1] + 1):  # Перебираем строки
            if 'Итого' in str(row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill

    return wb


def write_df_to_excel_expired_docs(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        # Красим в зависимости от цвета
        last_column_number = wb.active.max_column  # номер колонки в которой ищем слово Осталось дней
        # Создаем  стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        # Месяц
        fill_week = PatternFill(fill_type='solid', fgColor='ff0000')  # Красный цвет
        fill_two_week = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        fill_month = PatternFill(fill_type='solid', fgColor='ffff00')  # Желтый цвет
        for row in wb[name_sheet].iter_rows(min_row=2, max_row=wb[name_sheet].max_row,
                                            min_col=0,
                                            max_col=last_column_number):  # Перебираем строки
            if int(row[last_column_number-1].value) <= 7:
                for cell in row:
                    cell.font = font
                    cell.fill = fill_week
            elif int(row[last_column_number-1].value) <= 14:
                for cell in row:
                    cell.font = font
                    cell.fill = fill_two_week
            else:
                for cell in row:
                    cell.font = font
                    cell.fill = fill_month


        count_index += 1

    return wb


def del_sheet(wb: openpyxl.Workbook, lst_name_sheet: list) -> openpyxl.Workbook:
    """
    Функция для удаления лишних листов из файла
    :param wb: объект таблицы
    :param lst_name_sheet: список удаляемых листов
    :return: объект таблицы без удаленных листов
    """
    for del_sheet in lst_name_sheet:
        if del_sheet in wb.sheetnames:
            del wb[del_sheet]

    return wb


def write_group_df_to_excel(wb:openpyxl.Workbook,name_sheet:str,df:pd.DataFrame,write_index:bool,write_header:bool):
    """
    Функция для записи сгруппированных датафреймов так как у них иногда появляются пустые сроки
    :param wb: документ openpyxl
    :param name_sheet: название листа
    :param df: датафрейм для записи
    :param write_index: записывать ли индекс
    :param write_header: записывать ли заголовок
    :return: записаный лист
    """
    # записываем данные в лист
    none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
    for row in dataframe_to_rows(df, index=write_index, header=write_header):
        if len(row) == 1 and not row[0]:  # убираем пустую строку
            none_check = True
            wb[name_sheet].append(row)
        else:
            wb[name_sheet].append(row)
    if none_check:
        wb[name_sheet].delete_rows(2)
        # сохраняем по ширине колонок
    for column in wb[name_sheet].columns:
        max_length = 0
        column_name = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width >= 80:
            wb[name_sheet].column_dimensions[column_name].width = 80
            for cell in wb[name_sheet][column_name]:
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
        else:
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width + 2
    return wb


def write_to_excel_pers_egisso(df:pd.DataFrame, type:str):
    """
    Функция для создания файла openpyxl  с листами по льготам
    :param df: датафрейм с данными
    :param type: для выбора параметров автоширины колонок Чистый или Ошибки
    :return: файл openpyxl WOrkbook
    """
    wb = openpyxl.Workbook()
    name_base = wb.sheetnames[0] # Получаем название листа
    # Записываем общий список
    for row in dataframe_to_rows(df, index=False, header=True):
        wb[name_base].append(row)
    # Устанавливаем автоширину колонок
    if type == 'Чистый':
        for column in wb[name_base].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_base].column_dimensions[column_name].width = adjusted_width
        wb[name_base].column_dimensions['C'].width = 40
    else:
        wb[name_base].column_dimensions['B'].width = 15
        wb[name_base].column_dimensions['F'].width = 15
        wb[name_base].column_dimensions['G'].width = 16
        wb[name_base].column_dimensions['H'].width = 16
        wb[name_base].column_dimensions['I'].width = 16
        wb[name_base].column_dimensions['K'].width = 16
        wb[name_base].column_dimensions['N'].width = 16
        wb[name_base].column_dimensions['O'].width = 16

    # Создаем лист для дубликатов получателей льгот
    wb.create_sheet('Несколько льгот',1)

    dupl_df = df[df['СНИЛС'].duplicated(keep=False)] # получаем дубликаты по СНИЛС
    dupl_df['СНИЛС'] = dupl_df['СНИЛС'].astype(str)
    dupl_df.sort_values(by='СНИЛС',inplace=True)
    for row in dataframe_to_rows(dupl_df, index=False, header=True):
        wb['Несколько льгот'].append(row)
    # Устанавливаем автоширину колонок
    if type == 'Чистый':
        for column in wb['Несколько льгот'].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb['Несколько льгот'].column_dimensions[column_name].width = adjusted_width
        wb['Несколько льгот'].column_dimensions['C'].width = 40
    else:
        wb['Несколько льгот'].column_dimensions['B'].width = 15
        wb['Несколько льгот'].column_dimensions['F'].width = 15
        wb['Несколько льгот'].column_dimensions['G'].width = 16
        wb['Несколько льгот'].column_dimensions['H'].width = 16
        wb['Несколько льгот'].column_dimensions['I'].width = 16
        wb['Несколько льгот'].column_dimensions['K'].width = 16
        wb['Несколько льгот'].column_dimensions['N'].width = 16
        wb['Несколько льгот'].column_dimensions['O'].width = 16

    lst_ben = df['Льгота'].unique() # Список льгот
    for idx,benefit in enumerate(lst_ben,2):
        temp_df = df[df['Льгота'] == benefit] # фильтруем датафрейм по названию льготы
        short_value = benefit[:25]  # получаем обрезанное значение
        short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)
        wb.create_sheet(short_value,idx) # создаем лист
        for row in dataframe_to_rows(temp_df, index=False, header=True):
            wb[short_value].append(row)
        # Устанавливаем автоширину колонок
        if type == 'Чистый':
            for column in wb[short_value].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                wb[short_value].column_dimensions[column_name].width = adjusted_width
            wb[short_value].column_dimensions['C'].width = 40
        else:
            wb[short_value].column_dimensions['B'].width = 15
            wb[short_value].column_dimensions['F'].width = 15
            wb[short_value].column_dimensions['G'].width = 16
            wb[short_value].column_dimensions['H'].width = 16
            wb[short_value].column_dimensions['I'].width = 16
            wb[short_value].column_dimensions['K'].width = 16
            wb[short_value].column_dimensions['N'].width = 16
            wb[short_value].column_dimensions['O'].width = 16

    wb[name_base].title = 'Общий список'
    return wb

def write_to_excel_full_egisso(df:pd.DataFrame, type:str):
    """
    Функция для создания файла openpyxl с листами по льготам
    :param df: датафрейм с данными
    :param type: для выбора параметров автоширины колонок Чистый или Ошибки
    :return: файл openpyxl WOrkbook
    """
    wb = openpyxl.Workbook()
    name_base = wb.sheetnames[0] # Получаем название листа
    # Записываем общий список
    for row in dataframe_to_rows(df, index=False, header=True):
        wb[name_base].append(row)
    # Устанавливаем автоширину колонок
    if type == 'Чистый':
        for column in wb[name_base].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_base].column_dimensions[column_name].width = adjusted_width
        wb[name_base].column_dimensions['C'].width = 40
    else:
        wb[name_base].column_dimensions['A'].width = 15
        wb[name_base].column_dimensions['B'].width = 25
        wb[name_base].column_dimensions['F'].width = 20
        wb[name_base].column_dimensions['G'].width = 16
        wb[name_base].column_dimensions['H'].width = 16
        wb[name_base].column_dimensions['I'].width = 16
        wb[name_base].column_dimensions['K'].width = 16
        wb[name_base].column_dimensions['N'].width = 16
        wb[name_base].column_dimensions['O'].width = 16

    # Создаем лист для дубликатов получателей льгот
    wb.create_sheet('Несколько льгот',1)

    if type == 'Чистый':
        dupl_df = df[df['SNILS_recip'].duplicated(keep=False)] # получаем дубликаты по СНИЛС
        dupl_df['SNILS_recip'] = dupl_df['SNILS_recip'].astype(str)
        dupl_df.sort_values(by='SNILS_recip',inplace=True)
        for row in dataframe_to_rows(dupl_df, index=False, header=True):
            wb['Несколько льгот'].append(row)
    else:
        dupl_df = df[df['СНИЛС'].duplicated(keep=False)] # получаем дубликаты по СНИЛС
        dupl_df['СНИЛС'] = dupl_df['СНИЛС'].astype(str)
        dupl_df.sort_values(by='СНИЛС',inplace=True)
        for row in dataframe_to_rows(dupl_df, index=False, header=True):
            wb['Несколько льгот'].append(row)
    # Устанавливаем автоширину колонок
    if type == 'Чистый':
        for column in wb['Несколько льгот'].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb['Несколько льгот'].column_dimensions[column_name].width = adjusted_width
        wb['Несколько льгот'].column_dimensions['C'].width = 40
    else:
        wb['Несколько льгот'].column_dimensions['B'].width = 15
        wb['Несколько льгот'].column_dimensions['F'].width = 15
        wb['Несколько льгот'].column_dimensions['G'].width = 16
        wb['Несколько льгот'].column_dimensions['H'].width = 16
        wb['Несколько льгот'].column_dimensions['I'].width = 16
        wb['Несколько льгот'].column_dimensions['K'].width = 16
        wb['Несколько льгот'].column_dimensions['N'].width = 16
        wb['Несколько льгот'].column_dimensions['O'].width = 16

    lst_ben = df['Льгота'].unique() # Список льгот
    for idx,benefit in enumerate(lst_ben,2):
        temp_df = df[df['Льгота'] == benefit] # фильтруем датафрейм по названию льготы
        short_value = benefit[:25]  # получаем обрезанное значение
        short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)
        wb.create_sheet(short_value,idx) # создаем лист
        for row in dataframe_to_rows(temp_df, index=False, header=True):
            wb[short_value].append(row)
        # Устанавливаем автоширину колонок
        if type == 'Чистый':
            for column in wb[short_value].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                wb[short_value].column_dimensions[column_name].width = adjusted_width
            wb[short_value].column_dimensions['C'].width = 40
        else:
            wb[short_value].column_dimensions['B'].width = 15
            wb[short_value].column_dimensions['F'].width = 15
            wb[short_value].column_dimensions['G'].width = 16
            wb[short_value].column_dimensions['H'].width = 16
            wb[short_value].column_dimensions['I'].width = 16
            wb[short_value].column_dimensions['K'].width = 16
            wb[short_value].column_dimensions['N'].width = 16
            wb[short_value].column_dimensions['O'].width = 16

    wb[name_base].title = 'Общий список'
    return wb


def write_to_excel_non_find_ben_egisso(df:pd.DataFrame):
    """
    Функция для создания файла openpyxl с листами по льготам для которых не найдены совпадения
    :param df: датафрейм с данными
    :return: файл openpyxl WOrkbook
    """
    wb = openpyxl.Workbook()
    name_base = wb.sheetnames[0] # Получаем название листа
    pers_df = df[df['_merge'] =='left_only'] # получаем получателей льгот у которых нет совпадений
    pers_df.drop(columns=['Название колонки с льготой','Наименование категории','LMSZID','categoryID','ONMSZCode',
                          'LMSZProviderCode','providerCode','usingSign','criteria','criteriaCode',
                          'FormCode','amount','measuryCode','monetization','content','comment','equivalentAmount','_merge'],inplace=True)

    # Записываем льготников
    for row in dataframe_to_rows(pers_df, index=False, header=True):
        wb[name_base].append(row)

    wb[name_base].column_dimensions['A'].width = 15
    wb[name_base].column_dimensions['B'].width = 35
    wb[name_base].column_dimensions['G'].width = 15

    ben_df = df[df['_merge'] == 'right_only']
    ben_df.drop(columns=['Льгота','Статус льготы','Реквизиты','Дата окончания льготы','Файл',
                          'СНИЛС','Фамилия','Имя','Отчество','Пол',
                          'Дата_рождения','Тип документа','Серия_паспорта','Номер_паспорта','Дата_выдачи_паспорта','Кем_выдан','_merge'],inplace=True)
    wb.create_sheet('Льготы без получателей',index=1)
    # Записываем льготы для которых не найдены получатели
    for row in dataframe_to_rows(ben_df, index=False, header=True):
        wb['Льготы без получателей'].append(row)
    wb['Льготы без получателей'].column_dimensions['A'].width = 15
    wb['Льготы без получателей'].column_dimensions['B'].width = 35
    wb['Льготы без получателей'].column_dimensions['G'].width = 15

    wb[name_base].title = 'Получатели без льгот'
    return wb


def write_to_excel_print_group_egisso(df:pd.DataFrame,path_end_folder:str):
    """
    Функция для создания файла openpyxl с листами по льготам для которых не найдены совпадения
    :param df: датафрейм с данными
    :return: файл openpyxl WOrkbook
    """
    lst_group = df['Группа'].unique() # список групп
    used_name_sheet = set()  # множество для хранения значений которые уже были использованы
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)
    if len(lst_group) >= 253:
        raise ExceedingQuantity

    wb = openpyxl.Workbook()  # создаем файл
    name_base = wb.sheetnames[0] # базовый лист
    df['№ п/п'] = range(1,len(df)+1)
    # Записываем общий лист
    for row in dataframe_to_rows(df, index=False, header=True):
        wb[name_base].append(row)
    # Устанавливаем автоширину для каждой колонки
    for column in wb[name_base].columns:
        max_length = 0
        column_name = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        wb[name_base].column_dimensions[column_name].width = adjusted_width




    # Записываем остальные листы
    for idx, value in enumerate(lst_group,1):
        temp_df = df[df['Группа'] == value]  # отфильтровываем по значению
        temp_df['№ п/п'] = range(1,len(temp_df)+1)
        short_value = value[:20]  # получаем обрезанное значение
        short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

        if short_value in used_name_sheet:
            short_value = f'{short_value}_{idx}'  # добавляем окончание
        wb.create_sheet(short_value, index=idx)  # создаем лист
        used_name_sheet.add(short_value)
        for row in dataframe_to_rows(temp_df, index=False, header=True):
            wb[short_value].append(row)

        # Устанавливаем автоширину для каждой колонки
        for column in wb[short_value].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[short_value].column_dimensions[column_name].width = adjusted_width
    wb[name_base].title = 'Общий список льготников'
    wb.save(f'{path_end_folder}/Льготники {current_time}.xlsx')
    wb.close()

    # Сохраняем по отдельным файлам в папку
    path_group_file =f'{path_end_folder}/Льготники по группам'
    if not os.path.exists(path_group_file):
        os.makedirs(path_group_file)
    used_name_file = set()  # множество для уже использованных имен файлов
    for idx, value in enumerate(lst_group):
        wb = openpyxl.Workbook()  # создаем файл
        name_base = wb.sheetnames[0]
        temp_df = df[df['Группа'] == value]  # отфильтровываем по значению
        temp_df['№ п/п'] = range(1, len(temp_df) + 1)
        short_name = value[:40]  # получаем обрезанное значение
        short_name = re.sub(r'[\r\b\n\t\'+()<> :"?*|\\/]', '_', short_name)
        if short_name in used_name_file:
            short_name = f'{short_name}_{idx}'  # добавляем окончание
        for row in dataframe_to_rows(temp_df, index=False, header=True):
            wb[name_base].append(row)

        # Устанавливаем автоширину для каждой колонки
        for column in wb[name_base].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_base].column_dimensions[column_name].width = adjusted_width

        # Создаем листы по льготам
        lst_ben = temp_df['Льгота'].unique() # список имеющихся льгот
        for idx, value in enumerate(lst_ben, 1):
            ben_df = temp_df[temp_df['Льгота'] == value]  # отфильтровываем по значению
            ben_df['№ п/п'] = range(1, len(ben_df) + 1)
            short_value = value[:20]  # получаем обрезанное значение
            short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

            wb.create_sheet(short_value, index=idx)  # создаем лист
            used_name_sheet.add(short_value)
            for row in dataframe_to_rows(ben_df, index=False, header=True):
                wb[short_value].append(row)

            # Устанавливаем автоширину для каждой колонки
            for column in wb[short_value].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                wb[short_value].column_dimensions[column_name].width = adjusted_width

        wb[name_base].title = 'Общий список'
        wb.save(f'{path_group_file}/{short_name}.xlsx')
        used_name_file.add(short_name)
        wb.close()



def replace_point(value):
    if value is np.nan:
        return None
    try:
        return float(value)
    except:
        return None


def extract_parameters_egisso(path_egisso_params: str, df_cols:list):
    """
    Функция для извлечения параметров из файла егиссо
    """
    dct_params = {} # словарь для хранения параметров ЕГИССО
    # датафрейм для ошибок
    error_df = pd.DataFrame(
        columns=['Название файла', 'Название листа', 'Значение ошибки', 'Описание ошибки'])  # датафрейм для ошибок
    df_params = pd.read_excel(path_egisso_params, dtype=str)
    required_cols_set = {'Название колонки с льготой','Наименование категории','LMSZID','categoryID','ONMSZCode',
                         'LMSZProviderCode','providerCode','decision_date','dateStart','dateFinish','usingSign','criteria','criteriaCode','FormCode','amount',
                         'measuryCode','monetization','content','comment','equivalentAmount'}

    diff_cols = required_cols_set.difference(set(df_params.columns))
    # проверяем на наличие обязательных колонок
    if len(diff_cols) != 0:
        temp_error_df = pd.DataFrame(
            data=[[f'{path_egisso_params}', f'Первый лист по порядку', f'{";".join(diff_cols)}',
                   'В файле на указанном листе не найдены указанные обязательные колонки. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
            columns=['Название файла', 'Название листа', 'Значение ошибки',
                     'Описание ошибки'])
        error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
        return dct_params,error_df
    else:
        df_params.dropna(thresh=3,inplace=True) # очищаем от пустых строк где менее 3 заполненных колонок
        # Обрабатываем незаполненные значения в главных колонках
        df_params[['Название колонки с льготой','Наименование категории']] = df_params[['Название колонки с льготой','Наименование категории']].fillna('Не заполнено')
        df_params[['decision_date','dateStart','dateFinish']] = df_params[['decision_date','dateStart','dateFinish']].applymap(convert_to_date)
        df_params[['decision_date', 'dateStart', 'dateFinish']] = df_params[
            ['decision_date', 'dateStart', 'dateFinish']].applymap(create_doc_convert_date)
        df_params[['amount', 'equivalentAmount']] = df_params[['amount','equivalentAmount']].applymap(replace_point)
        # Проверяем наличие колонок в датафрейме
        for idx,ben_col in enumerate(df_params['Название колонки с льготой'].tolist(),2):
            if ben_col not in df_cols:
                temp_error_df = pd.DataFrame(
                    data=[[f'{path_egisso_params}', f'Первый лист по порядку', f'{ben_col}',
                           f'Колонка {ben_col} на строке {idx} в файле с параметрами ЕГИССО отсутствует в эталонном файле. Параметры в данной строке не будут обрабатываться']],
                    columns=['Название файла', 'Название листа', 'Значение ошибки',
                             'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)

        df_params = df_params[df_params['Название колонки с льготой'].isin(df_cols)]
        df_params['Название колонки с льготой'] = df_params['Название колонки с льготой'].apply(lambda x:str.replace(x,'Статус_',''))


        return df_params,error_df

"""
Функции для проверки персональных данных
"""

def check_simple_str_column(value, error_str: str):
    """
    Функция для проверки на заполнение ячейки для простой колонки с текстом не требующим дополнительной проверки
    :param value: значение ячейки
    :param error_str: сообщение об ошибки
    """
    if pd.isna(value):
        return error_str
    elif value == 'Нет статуса':
        return 'Ошибка Не заполнено!'
    else:
        return value

def processing_snils(value):
    """
    Функция для проверки и обработки СНИЛС
    :param value:
    :return:
    """
    if value == 'Нет статуса':
        return 'Ошибка СНИЛС не заполнен!'
    result = re.findall(r'\d',value)
    if len(result) == 11:
        # проверяем на лидирующий ноль
        out_str = ''.join(result)
        if out_str.startswith('0'):
            return out_str
        else:
            return int(out_str)
    else:
        return f'Ошибка: В СНИЛС должно быть 11 цифр а в ячейке {len(result)} цифр(ы). В ячейке указано - {value}'

def processing_fio(value,pattern):
    """
    Функция для проверки соответствия
    :param value:значение
    :param pattern: объект re.compile
    :return:
    """
    if re.fullmatch(pattern,value):
        return value
    else:
        return f'Ошибка: ФИО должно начинаться с большой буквы и содержать только буквы кириллицы и дефис. В ячейке указано - {value}'

def comparison_date(value, pattern):
    """
    Функция для проверки соответсвия формата даты
    :param value:значение
    :param pattern: объект re.compile
    :return:
    """
    if re.fullmatch(pattern,value):
        return value
    else:
        return f'Ошибка: В ячейке указано - {value}'


def processing_series(value, pattern):
    """
    Функция для проверки соответсвия формата серии паспорта
    :param value:значение
    :param pattern: объект re.compile
    :return:
    """
    if value == 'Нет статуса':
        return 'Ошибка Серия паспорта не заполнена!'
    if re.fullmatch(pattern,value):
        if value.startswith('0'):
            return value
        else:
            return int(value)
    else:
        error_str = re.sub(r'\s','Пробельный символ',value)
        return f'Ошибка: Серия паспорта должна состоять из 4 цифр без пробелов, например 0343. В ячейке указано - {error_str}'

def processing_number(value, pattern):
    """
    Функция для проверки соответсвия формата номера паспорта
    :param value:значение
    :param pattern: объект re.compile
    :return:
    """
    if value == 'Нет статуса':
        return 'Ошибка номер паспорта не заполнен!'

    if re.fullmatch(pattern,value):
        if value.startswith('0'):
            return value
        else:
            return int(value)
    else:
        error_str = re.sub(r'\s','Пробельный символ',value)
        return f'Ошибка: Номер паспорта должен состоять из 6 цифр без пробелов, например 420343. В ячейке указано - {error_str}'

def find_error_in_row(row):
    """
    Функция для поиска в каждой колонке строки слова Ошибка
    :param row:
    :return:
    """
    value_lst = row.tolist()
    error_lst = [value for value in value_lst if isinstance(value,str) and 'Ошибка' in value]
    if len(error_lst) !=0:
        return 'Ошибка'
    else:
        return 'Нет ошибок'


def check_inn(inn):
    """
    Функция для приведения значений снилс в вид 12 цифр
    """
    if inn is np.nan:
        return 'Ошибка'
    inn = str(inn)
    if inn == 'Нет статуса':
        return 'Ошибка ИНН не заполнен!'
    result = re.findall(r'\d', inn) # ищем цифры
    if len(result) == 12:
        return ''.join(result)
    else:
        return f'Ошибка ИНН (ИНН физлица должно состоять из 12 цифр)- {inn} -{len(result)} цифр'

def contains_word(s):
    s = str(s)
    return 'Ошибка' in s


def check_error_in_pers_data(df:pd.DataFrame,path_end_folder:str,current_time):
    """
    Функция для проверки правильности данных
    :param df:датафрейм который нужно проверить на корректность заполнения
    :param path_end_folder: папка куда будет сохранен результат
    :param current_time: время сохранения
    """

    out_df = df[['Файл','Группа','ФИО']]
    # Проверяем М и Ж
    out_df['Пол'] = df['Пол'].apply(lambda x:x if x in ('М','Ж') else f'Ошибка: Допустимые значения М и Ж. В ячейке указано {x}')
    # СНИЛС
    df['СНИЛС'] = df['СНИЛС'].astype(str)
    out_df['СНИЛС'] = df['СНИЛС'].apply(processing_snils) # проверяем снилс и конвертируем снилс
    # проверяем колонку ИНН
    out_df['ИНН'] = df['ИНН'].apply(check_inn) # проверяем ИНН


    date_pattern = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')  # созадем паттерн
    df['Дата_рождения'] = df['Дата_рождения'].astype(str)
    out_df['Дата_рождения'] = df['Дата_рождения'].apply(lambda x: comparison_date(x, date_pattern))
    # Проверяем колонку серия паспорта
    series_pattern = re.compile(r'^\d{4}$')
    df['Серия_паспорта'] = df['Серия_паспорта'].astype(str)
    out_df['Серия_паспорта'] = df['Серия_паспорта'].apply(lambda x: processing_series(x, series_pattern))
    # проверяем номер паспорта
    number_pattern = re.compile(r'^\d{6}$')
    df['Номер_паспорта'] = df['Номер_паспорта'].astype(str)
    out_df['Номер_паспорта'] = df['Номер_паспорта'].apply(lambda x: processing_number(x, number_pattern))
    # проверяем колонку дата выдачи паспорта
    date_pattern = re.compile(r'^\d{2}\.\d{2}.\d{4}$')  # созадем паттерн
    df['Дата_выдачи_паспорта'] = df['Дата_выдачи_паспорта'].astype(str)
    out_df['Дата_выдачи_паспорта'] = df['Дата_выдачи_паспорта'].apply(lambda x: comparison_date(x, date_pattern))
    out_df['Код_подразделения'] = df['Код_подразделения'].apply(lambda x: check_simple_str_column(x, 'Ошибка: не заполнено'))
    # Проверяем колонку Кем выдано
    out_df['Кем_выдан'] = df['Кем_выдан'].apply(lambda x: check_simple_str_column(x, 'Ошибка: не заполнено'))
    # Проверяем адреса регистрации и фактический адрес
    out_df['Адрес_регистрации'] = df['Адрес_регистрации'].apply(lambda x: check_simple_str_column(x, 'Ошибка: не заполнено'))
    out_df['Фактический_адрес'] = df['Фактический_адрес'].apply(lambda x: check_simple_str_column(x, 'Ошибка: не заполнено'))




    # првоеряем ФИО
    fio_pattern = re.compile(r'^[ЁА-Я][ёЁа-яА-Я-]+$')
    out_df['Фамилия'] = df['Фамилия'].apply(lambda x:processing_fio(x,fio_pattern)) # проверяем фамилию
    out_df['Имя'] = df['Имя'].apply(lambda x:processing_fio(x,fio_pattern)) # проверяем имя
    out_df['Отчество'] = df['Отчество'].apply(lambda x:processing_fio(x,fio_pattern)) # проверяем отчество




    # определяем строки где встречается слово ошибка
    out_df['Ошибка'] = out_df.apply(find_error_in_row, axis=1)

    error_df = out_df[out_df['Ошибка'] == 'Ошибка']
    # Убираем лишнюю колонку
    error_df.drop(columns=['Ошибка'],inplace=True)
    # Убираем колонки в которых нет ошибок чтобы таблица была компактнее
    lst_check_cols = ['Пол','СНИЛС','ИНН','Дата_рождения','Серия_паспорта','Номер_паспорта','Код_подразделения','Кем_выдан','Дата_выдачи_паспорта','Адрес_регистрации','Фактический_адрес','Фамилия','Имя','Отчество']
    # Проверяем колонки
    columns_to_drop = []
    for column in lst_check_cols:
        if not error_df[column].apply(contains_word).any():
            columns_to_drop.append(column)

    # Удаляем найденные колонки
    error_df.drop(columns_to_drop, axis=1, inplace=True)
    out_dct = {'Общий список ошибок':error_df}
    # Создаем отдельные датафреймы
    for name_col in error_df.columns:
        if name_col not in ('Файл','Группа','ФИО'):
            error_df[name_col] = error_df[name_col].astype(str)
            temp_df = error_df[error_df[name_col].str.contains('Ошибка')]
            temp_df = temp_df[['Файл','Группа','ФИО',name_col]]
            out_dct[name_col] = temp_df





    error_pers_wb = write_df_to_excel(out_dct, write_index=False)
    error_pers_wb = del_sheet(error_pers_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
    error_pers_wb.save(f'{path_end_folder}/Ошибки в персональных данных от {current_time}.xlsx')



def convert_snils_dash(snils):
    """
    Функция для приведения значений снилс в вид ХХХ-ХХХ-ХХХ ХХ
    """
    if snils is np.nan:
        return 'Не заполнено'
    snils = str(snils)
    result = re.findall(r'\d', snils) # ищем цифры
    if len(result) == 11:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        third_group = ''.join(result[6:9])
        four_group = ''.join(result[9:11])

        out_snils = f'{first_group}-{second_group}-{third_group} {four_group}'
        return out_snils
    else:
        return f'Неправильное значение!В СНИЛС должно быть 11 цифр - {snils} -{len(result)} цифр'

def convert_snils_not_dash(snils):
    """
    Функция для приведения значений снилс в вид XXXXXXXXXXX 11 цифр подряд
    """
    if snils is np.nan:
        return 'Не заполнено'
    snils = str(snils)
    result = re.findall(r'\d', snils) # ищем цифры
    if len(result) == 11:
        out_snils = ''.join(result)

        return out_snils
    else:
        return f'Неправильное значение!В СНИЛС должно быть 11 цифр - {snils} -{len(result)} цифр'