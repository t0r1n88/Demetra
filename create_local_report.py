"""
Скрипт для обработки списка студентов на отделении и создания отчетности по нему
"""
from demetra_support_functions import write_df_to_excel, del_sheet, \
    declension_fio_by_case,extract_parameters_egisso,write_df_big_dct_to_excel,check_error_in_pers_data,convert_snils_dash,convert_snils_not_dash
from demetra_processing_date import proccessing_date
from demetra_egisso import create_part_egisso_data, create_full_egisso_data
from tkinter import messagebox
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import time
import datetime
from collections import Counter
import re
import os
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import sys
import locale


def set_rus_locale():
    """
    Функция чтобы можно было извлечь русские названия месяцев
    """
    locale.setlocale(
        locale.LC_ALL,
        'rus_rus' if sys.platform == 'win32' else 'ru_RU.UTF-8')


class NotColumn(Exception):
    """
    Исключение для обработки случая когда отсутствуют нужные колонки
    """
    pass


class NotCustomColumn(Exception):
    """
    Исключение для обработки случая когда отсутствуют колонки указанные в параметрах
    """
    pass


class NotGoodSheet(Exception):
    """
    Исключение для случая когда ни один лист не подхожит под эталон
    """
    pass

class BadEtalonFile(Exception):
    """
    Исключение для обработки случая когда эталонный файл поврежден и его нельзя открыть
    """
    pass
def convert_number(value):
    """
    Функция для конвертации в float значений колонок содержащих в названии Подсчет_
    :param value: значение
    :return: число в формате float
    """
    try:
        return float(value)
    except:
        return 0

def convert_to_date(value):
    """
    Функция для конвертации строки в текст
    :param value: значение для конвертации
    :return:
    """
    try:
        if value == 'Нет статуса':
            return ''
        else:
            date_value = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
            return date_value
    except ValueError:
        result = re.search(r'^\d{2}\.\d{2}\.\d{4}$', value)
        if result:
            try:
                return datetime.datetime.strptime(result.group(0), '%d.%m.%Y')
            except ValueError:
                # для случаев вида 45.09.2007
                return f'Некорректный формат даты - {value}, проверьте лишние пробелы,наличие точек'
        else:
            return f'Некорректный формат даты - {value}, проверьте лишние пробелы,наличие точек'
    except:
        return f'Некорректный формат даты - {value}, проверьте лишние пробелы,наличие точек'



def create_value_str(df: pd.DataFrame, name_column: str, target_name_column: str, dct_str: dict,checkbox_expelled:int) -> pd.DataFrame:
    """
    Функция для формирования строки нужного формата с использованием переменных
    :param df:датафрейм
    :param name_column:название колонки для значений которой нужно произвести подсчет
    :param target_name_column: название колонки по которой будет производится подсчет
    :param dct_str:словарь с параметрами
    :param checkbox_expelled: как будет отображаться количество людей в академ и отчисленных
    :return:датафрейм
    """
    temp_counts = df[name_column].value_counts()  # делаем подсчет
    new_value_df = temp_counts.to_frame().reset_index()  # создаем датафрейм с данными
    new_value_df.columns = ['Показатель', 'Значение']  # делаем одинаковыми названия колонок
    new_value_df['Показатель'] = new_value_df['Показатель'].astype(str)
    new_value_df.sort_values(by='Показатель', inplace=True)
    for idx, row in enumerate(new_value_df.iterrows()):
        name_op = row[1].values[0]  # получаем название ОП
        temp_df = df[df[name_column] == name_op]  # отфильтровываем по названию ОП
        quantity_study_student = temp_df[temp_df[target_name_column] == dct_str['Обучается']].shape[
            0]  # со статусом Обучается
        quantity_academ_student = temp_df[temp_df[target_name_column].str.contains(dct_str['Академ'])].shape[
            0]
        quantity_not_status_student = \
        temp_df[temp_df[target_name_column].str.contains(dct_str['Не указан статус'])].shape[
            0]

        if checkbox_expelled == 0:
            out_str = f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student},' \
                      f' Не указан статус - {quantity_not_status_student}, Всего {len(temp_df)}'
        elif checkbox_expelled == 1:
            out_str = f'Обучается - {quantity_study_student},' \
                      f' Не указан статус - {quantity_not_status_student}, Всего {len(temp_df)}'
        else:
            quantity_except = temp_df[temp_df[target_name_column].str.contains('Отчислен')].shape[
                0]
            out_str = f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student}, Отчислено - {quantity_except}, Не указан статус - {quantity_not_status_student}, Всего {len(temp_df)}'

        new_value_df.iloc[idx, 1] = out_str  # присваиваем значение

    return new_value_df

def extract_part_status_op(value,part_extract:str):
    """
    Функция для извлечения кода или наименования специальности профессии
    :param value: значение
    :param part_extract: что именно нужно извлекать либо Код_ОП либо Наименование_ОП
    :return:
    """
    lst_part = str(value).split(' ',maxsplit=1) # делим по первому пробелу
    if len(lst_part) == 2:
        if part_extract == 'Код_ОП':
            return lst_part[0]
        else:
            return lst_part[1]
    else:
        return f''




def prepare_file_params(params_file: str):
    """
    Функция для подготовки словаря с параметрами, преобразуюет первую колонку в ключи а вторую колонку в значения
    :param params_file: путь к файлу с параметрами в формате xlsx
    :return: словарь с параметрами
    """
    df = pd.read_excel(params_file, usecols='A:B', dtype=str)
    df.dropna(inplace=True)  # удаляем все строки где есть нан
    lst_unique_name_column = df.iloc[:, 0].unique()  # получаем уникальные значения колонок в виде списка
    temp_dct = {key: {} for key in
                lst_unique_name_column}  # создаем словарь верхнего уровня для хранения сгенерированных названий колонок
    # перебираем датафрейм и получаем
    for row in df.itertuples():
        # заполняем словарь
        temp_dct[row[1]][f'{row[1]}_{row[2]}'] = row[2]
    return temp_dct, lst_unique_name_column, df


def create_for_custom_report(df: pd.DataFrame, params_df: pd.DataFrame) -> openpyxl.Workbook:
    """
    Функция для создания файла в котором в виде списков будут находиться данные использованные для создания настраиваемого отчета
    :param df: основной датафрейм
    :param params_df: датафрейм с параметрами
    :return:
    """
    used_name = set()  # множество для использованных названий листов
    dct_df = dict()  # словарь в котором будут храниться датафреймы
    for idx, row in enumerate(params_df.itertuples()):
        name_column = row[1]  # название колонки
        value_column = row[2]  # значение которое нужно подсчитать
        temp_df = df[df[name_column] == value_column]
        name_sheet = f'{name_column}'[:30]  # для того чтобы не было слишком длинного названия
        if name_sheet.lower() not in used_name:
            dct_df[name_sheet] = temp_df
            used_name.add(name_sheet.lower())
        else:
            dct_df[f'{name_sheet}_{idx}'] = temp_df
            used_name.add(f'{name_sheet}_{idx}')

    lst_custom_wb = write_df_to_excel(dct_df, write_index=False)
    lst_custom_wb = del_sheet(lst_custom_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
    return lst_custom_wb


def create_slice_report(df:pd.DataFrame,slice_column:str,dct_params:dict):
    """
    Функция для создания отчета-среза по выбранной колонке по показателям указанным в словаре
    :param df:датафрейм с объединенными данными
    :param slice_column: колонка по которой нужно произвести срез
    :param dct_params:словарь в котором указаны колонки по которым нужны данные и данные которые надо считать
    :return:датафрейм
    """
    value_for_index = df[slice_column].unique() # создаем индекс из уникальных значений в колонке
    out_df = pd.DataFrame(index=value_for_index) # создаем базовый датафрейм

    # Перебираем словарь с параметрами создавая датафреймы для каждого значения
    for name_column, dct_columns in dct_params.items():
        for name_and_target,target_value in dct_columns.items():
            temp_df = df[df[name_column] == target_value] # получаем датафрейм где есть указанные значения
            svod_df = pd.pivot_table(temp_df,index=slice_column,
                                     values='ФИО',
                                     aggfunc='count').rename(columns={'ФИО':name_and_target})
            # если нет то добавляем колонку с нулями
            if len(svod_df) == 0:
                svod_df[name_and_target] = 0
            out_df = pd.concat([out_df,svod_df],axis=1) # добавляем в общий датафрейм

    out_df.fillna(0,inplace=True) # заполняем наны
    out_df = out_df.astype(int) # приводим к инту
    if slice_column == 'Текущий_возраст' or slice_column== 'Год_рождения':
        # для проведения сортировки с учетом слова ошибочное значение
        out_df.rename(index={'Ошибочное значение!!!': 100000000}, inplace=True)
        out_df.sort_index(inplace=True)
        out_df.rename(index={100000000: 'Ошибочное значение!!!'}, inplace=True)

    out_df.loc['Итого'] = out_df.sum(axis=0) # создаем результирующую строку
    out_df = out_df.reset_index().rename(columns={'index':slice_column})

    return out_df

def create_svod_counting(df: pd.DataFrame, name_column: str, postfix_file: str, lst_counting_name_columns: list,
                         path: str, current_time):
    """
    Функция для создания сводов по колонкам подсчета
    :param df: основной датафрейм
    :param name_column: название колонки
    :param postfix_file: дополнение к имени файла
    :param lst_counting_name_columns: список колонок типа Подсчет
    :param path: путь куда сохранять файлы
    :param current_time: время под которым будет сохраняться файл
    """
    dct_counting_df = {}  # словарь для хранения датафреймов
    for name_counting_column in lst_counting_name_columns:
        temp_svod_df = (pd.pivot_table(df, index=[name_column],
                                       values=[name_counting_column],
                                       aggfunc=[np.mean, np.sum, np.median, np.min, np.max, len]))
        temp_svod_df = temp_svod_df.reset_index()  # убираем мультииндекс
        temp_svod_df = temp_svod_df.droplevel(axis=1, level=0)  # убираем мультиколонки
        temp_svod_df.columns = [name_column, 'Среднее', 'Сумма', 'Медиана', 'Минимум', 'Максимум',
                                'Количество']
        if name_column == 'Текущий_возраст' or name_column == 'Год_рождения':
            # для проведения сортировки
            temp_svod_df.rename(index={'Ошибочное значение!!!': 100000000}, inplace=True)
            temp_svod_df.sort_index(inplace=True)
            temp_svod_df.rename(index={100000000: 'Ошибочное значение!!!'}, inplace=True)

        dct_counting_df[name_counting_column] = temp_svod_df  # сохраняем в словарь

    # Сохраняем контролируя количество листов в файле
    if len(dct_counting_df) >= 252:
        lst_dct = list(dct_counting_df.items())  # превращаем в список кортежей
        start_threshold = 0
        end_threshold = 252
        count = 1
        while end_threshold <= len(lst_dct) + 1:
            big_dct = dict(lst_dct[start_threshold:end_threshold])
            counting_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
            counting_report_wb = del_sheet(counting_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            counting_report_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
            count += 1
            start_threshold += 252
            end_threshold += 252
            # контролируем диапазон
            if end_threshold > len(lst_dct):
                big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                counting_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                counting_report_wb = del_sheet(counting_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                counting_report_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
    else:
        counting_report_wb = write_df_big_dct_to_excel(dct_counting_df, write_index=False)
        counting_report_wb = del_sheet(counting_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        counting_report_wb.save(f'{path}/Свод по {postfix_file} от {current_time}.xlsx')

def create_svod_list(df: pd.DataFrame, name_column: str, postfix_file: str, lst_list_name_columns: list,
                         path: str, current_time, dct_values_in_list_columns:dict):
    """
    Функция для создания сводов по колонкам списков
    :param df: основной датафрейм
    :param name_column: название колонки
    :param postfix_file: дополнение к имени файла
    :param lst_counting_name_columns: список колонок типа Список
    :param path: путь куда сохранять файлы
    :param current_time: время под которым будет сохраняться файл
    """
    dct_svod_list_df = {}  # словарь в котором будут храниться датафреймы по названию колонок
    for name_lst_column in lst_list_name_columns:
        file_cols = dct_values_in_list_columns[name_lst_column]
        file_cols.insert(0, name_column)
        dct_svod_list_df[name_lst_column] = pd.DataFrame(columns=file_cols)

    lst_file = df[name_column].unique()  # список файлов
    for name in lst_file:
        temp_df = df[df[name_column] == name]
        for name_lst_column in lst_list_name_columns:

            temp_col_value_lst = temp_df[name_lst_column].tolist()  # создаем список
            temp_col_value_lst = [value for value in temp_col_value_lst if value]  # отбрасываем пустые значения
            temp_unwrap_lst = []
            temp_col_value_lst = list(map(str, temp_col_value_lst))  # делаем строковым каждый элемент
            for value in temp_col_value_lst:
                temp_unwrap_lst.extend(value.split(','))
            temp_unwrap_lst = list(map(str.strip, temp_unwrap_lst))  # получаем список
            # убираем повторения и сортируем
            dct_values_in_list_columns[name_lst_column] = sorted(list(set(temp_unwrap_lst)))

            dct_value_list = dict(Counter(temp_unwrap_lst))  # Превращаем в словарь
            sorted_dct_value_lst = dict(sorted(dct_value_list.items()))  # сортируем словарь
            # создаем датафрейм
            temp_svod_df = pd.DataFrame(list(sorted_dct_value_lst.items()),
                                        columns=['Показатель', 'Значение']).transpose()
            new_temp_cols = temp_svod_df.iloc[0]  # получаем первую строку для названий
            temp_svod_df = temp_svod_df[1:]  # удаляем первую строку
            temp_svod_df.columns = new_temp_cols
            temp_svod_df.insert(0, name_column, name)
            # добавляем значение в датафрейм
            dct_svod_list_df[name_lst_column] = pd.concat([dct_svod_list_df[name_lst_column], temp_svod_df])

    for key, value_df in dct_svod_list_df.items():
        dct_svod_list_df[key].fillna(0, inplace=True)  # заполняем наны
        dct_svod_list_df[key] = dct_svod_list_df[key].astype(int, errors='ignore')
        if name_column == 'Текущий_возраст' or name_column == 'Год_рождения':
            dct_svod_list_df[key] = dct_svod_list_df[key].astype(str) # делаем строковой
            # заменяем ошибочное значение числом чтобы отсортировать
            dct_svod_list_df[key] = dct_svod_list_df[key].apply(lambda x:x.replace('Ошибочное значение!!!',100000000))
            dct_svod_list_df[key] = dct_svod_list_df[key].astype(float) # делаем числом
            dct_svod_list_df[key] = dct_svod_list_df[key].astype(int) # делаем числом
            if name_column == 'Текущий_возраст':
                dct_svod_list_df[key].sort_values(by='Текущий_возраст',inplace=True)
            else:
                dct_svod_list_df[key].sort_values(by='Год_рождения', inplace=True)


        sum_row = dct_svod_list_df[key].sum(axis=0)  # суммируем колонки
        dct_svod_list_df[key].loc['Итого'] = sum_row  # добавляем суммирующую колонку
        dct_svod_list_df[key].iloc[-1, 0] = 'Итого'
        if name_column == 'Текущий_возраст':
            dct_svod_list_df[key]['Текущий_возраст'] = dct_svod_list_df[key]['Текущий_возраст'] .astype(str)  # делаем строковой
            dct_svod_list_df[key]['Текущий_возраст']  = dct_svod_list_df[key]['Текущий_возраст'] .apply(lambda x: x.replace('100000000','Ошибочное значение!!!'))
        if name_column == 'Год_рождения':
            dct_svod_list_df[key]['Год_рождения'] = dct_svod_list_df[key]['Год_рождения'] .astype(str)  # делаем строковой
            dct_svod_list_df[key]['Год_рождения']  = dct_svod_list_df[key]['Год_рождения'] .apply(lambda x: x.replace('100000000','Ошибочное значение!!!'))
        # Сохраняем
    if len(dct_svod_list_df) >= 252:
        lst_dct = list(dct_svod_list_df.items())  # превращаем в список кортежей
        start_threshold = 0
        end_threshold = 252
        count = 1
        while end_threshold <= len(lst_dct) + 1:
            big_dct = dict(lst_dct[start_threshold:end_threshold])
            list_columns_svod_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
            list_columns_svod_wb = del_sheet(list_columns_svod_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            list_columns_svod_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
            count += 1
            start_threshold += 252
            end_threshold += 252
            # контролируем диапазон
            if end_threshold > len(lst_dct):
                big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                list_columns_svod_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                list_columns_svod_wb = del_sheet(list_columns_svod_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                list_columns_svod_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
    else:
        list_columns_svod_wb = write_df_big_dct_to_excel(dct_svod_list_df, write_index=False)
        list_columns_svod_wb = del_sheet(list_columns_svod_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        list_columns_svod_wb.save(f'{path}/Свод по {postfix_file} {current_time}.xlsx')


def create_svod_status(df: pd.DataFrame, name_column: str, postfix_file: str, lst_status: list,
                         path: str, current_time):
    """
    Функция для создания сводов по колонкам подсчета
    :param df: основной датафрейм
    :param name_column: название колонки
    :param postfix_file: дополнение к имени файла
    :param lst_counting_name_columns: список колонок типа Список
    :param path: путь куда сохранять файлы
    :param current_time: время под которым будет сохраняться файл
    """
    dct_status = {}  # словарь для хранения сводных датафреймов
    lst_status_columns = lst_status.copy()

    if name_column == 'Статус_Пол':
        lst_status_columns.remove('Статус_Пол')
    if name_column == 'Статус_ОП':
        lst_status_columns.remove('Статус_ОП')

    for name_status_column in lst_status_columns:
        svod_df = pd.pivot_table(df, index=name_column, columns=name_status_column, values='ФИО',
                                 aggfunc='count', fill_value=0, margins=True,
                                 margins_name='Итого').reset_index()
        name_sheet = name_status_column.replace('Статус_', '')
        dct_status[name_sheet] = svod_df  # сохраняем в словарь сводную таблицу

    # Сохраняем
    if len(dct_status) >= 252:
        lst_dct = list(dct_status.items())  # превращаем в список кортежей
        start_threshold = 0
        end_threshold = 252
        count = 1
        while end_threshold <= len(lst_dct) + 1:
            big_dct = dict(lst_dct[start_threshold:end_threshold])
            svod_status_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
            svod_status_wb = del_sheet(svod_status_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            svod_status_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
            count += 1
            start_threshold += 252
            end_threshold += 252
            # контролируем диапазон
            if end_threshold > len(lst_dct):
                big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                svod_status_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                svod_status_wb = del_sheet(svod_status_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                svod_status_wb.save(f'{path}/Свод по {postfix_file} {count}.xlsx')
    else:
        svod_status_wb = write_df_big_dct_to_excel(dct_status, write_index=False)
        svod_status_wb = del_sheet(svod_status_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        svod_status_wb.save(f'{path}/Свод по {postfix_file} {current_time}.xlsx')








def create_local_report(etalon_file: str, data_folder: str, path_end_folder: str, params_report: str,path_egisso_params,
                        checkbox_expelled: int, raw_date) -> None:
    """
    Функция для генерации отчетов на основе файла с данными групп
    """
    try:
        set_rus_locale()  # устанавливаем русскую локаль что категоризация по месяцам работала
        # обязательные колонки
        name_columns_set = {'Статус_ОП', 'Статус_Учёба', 'ФИО', 'Дата_рождения','СНИЛС', 'Пол','Адрес_регистрации','Фактический_адрес','Серия_паспорта', 'Номер_паспорта','Дата_выдачи_паспорта','Код_подразделения', 'Кем_выдан'}
        error_df = pd.DataFrame(
            columns=['Название файла', 'Название листа', 'Значение ошибки', 'Описание ошибки'])  # датафрейм для ошибок
        try:
            wb = openpyxl.load_workbook(etalon_file)  # загружаем эталонный файл
            quantity_sheets = 0  # считаем количество групп
            main_sheet = wb.sheetnames[0]  # получаем название первого листа с которым и будем сравнивать новые файлы
            main_df = pd.read_excel(etalon_file, sheet_name=main_sheet,
                                    nrows=0)  # загружаем датафрейм чтобы получить эталонные колонки
        except:
            raise BadEtalonFile

        # Проверяем на обязательные колонки
        always_cols = name_columns_set.difference(set(main_df.columns))
        if len(always_cols) != 0:
            raise NotColumn
        etalon_cols = set(main_df.columns)  # эталонные колонки
        # словарь для основных параметров по которым нужно построить отчет
        # список уникальных названий колонок для проверки эталонного файла
        # датарфейм с параметрами, нужен для создания списков
        dct_params, lst_unique_params, params_df = prepare_file_params(
            params_report)  # получаем значения по которым нужно подсчитать данные и уникальные названия колонок
        # проверяем наличие колонок из файла параметров в эталонном файле
        custom_always_cols = set(lst_unique_params).difference(set(main_df.columns))
        if len(custom_always_cols) != 0:
            raise NotCustomColumn
        lst_generate_name_columns = []  # создаем список для хранения значений сгенерированных колонок
        for key, value in dct_params.items():
            for name_gen_column in value.keys():
                lst_generate_name_columns.append(name_gen_column)
        # Создаем свод по файлам
        custom_report_df = pd.DataFrame(columns=lst_generate_name_columns)
        custom_report_df.insert(0, 'Файл', None)
        custom_report_df.insert(1, 'Лист', None)


        for idx, file in enumerate(os.listdir(data_folder)):
            if file.endswith('.xls') or file.endswith('.ods'):
                name_file = file.split('.xls')[0]
                temp_error_df = pd.DataFrame(data=[[f'{name_file}', '', '',
                                                    'Программа обрабатывает только XLSX и XLSM файлы ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!!']],
                                             columns=['Название файла', 'Строка или колонка с ошибкой',
                                                      'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                continue
            if not file.startswith('~$') and (file.endswith('.xlsx') or file.endswith('.xlsm')):
                if file.endswith('.xlsx'):
                    name_file = file.split('.xlsx')[0].strip()
                else:
                    name_file = file.split('.xlsm')[0].strip()
                print(f'Файл: {name_file}')
                try:
                    temp_wb = openpyxl.load_workbook(f'{data_folder}/{file}')  # открываем
                except:
                    temp_error_df = pd.DataFrame(
                        data=[[f'{name_file}', f'', f'',
                               'Не удалось обработать файл.']],
                        columns=['Название файла', 'Название листа', 'Значение ошибки',
                                 'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                    continue  # не обрабатываем лист, где найдены ошибки
                lst_sheets_temp_wb = temp_wb.sheetnames  # получаем список листов в файле
                for name_sheet in lst_sheets_temp_wb:
                    if name_sheet != 'Данные для выпадающих списков':  # отбрасываем лист с даннными выпадающих списков
                        try:
                            temp_df = pd.read_excel(f'{data_folder}/{file}',
                                                sheet_name=name_sheet,dtype=str)  # получаем колонки которые есть на листе
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'', f'',
                                       'Не удалось обработать файл.']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки

                        # Проверяем порядок колонок
                        order_main_columns = list(main_df.columns)  # порядок колонок эталонного файла
                        order_temp_df_columns = list(temp_df.columns)  # порядок колонок проверяемого файла
                        error_order_lst = []  # список для несовпадающих пар
                        # Сравниваем попарно колонки
                        for main, temp in zip(order_main_columns, order_temp_df_columns):
                            if main != temp:
                                error_order_lst.append(f'На месте колонки {main} находится колонка {temp}')
                        if len(error_order_lst) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(error_order_lst)}',
                                       'Неправильный порядок колонок по сравнению с эталоном. Приведите порядок колонок в соответствии с порядком в эталоне. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист, где найдены ошибки


                        # проверяем на соответствие эталонному файлу
                        diff_cols = etalon_cols.symmetric_difference(set(temp_df.columns))
                        if len(diff_cols) != 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'{";".join(diff_cols)}',
                                       'В файле на указанном листе найдены лишние или отличающиеся колонки по сравнению с эталоном. ДАННЫЕ ФАЙЛА НЕ ОБРАБОТАНЫ !!! ']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                            continue  # не обрабатываем лист где найдены ошибки

                        # указываем где есть ФИО в которых пробельные символы
                        lst_fio = temp_df['ФИО'].tolist()  # делаем список
                        lst_spaces = []  # список в котором будут храниться индексы
                        for idx, fio in enumerate(lst_fio):
                            lst_spaces.append(fio)
                            if str(fio).strip() == '':
                                lst_spaces[idx] = f'Строка {idx + 2} состоит из только из пробельных символов'

                        temp_df['ФИО'] = lst_spaces  #
                        temp_df.dropna(how='all', inplace=True)  # удаляем пустые строки
                        temp_df = temp_df[temp_df['ФИО'].notna()] # отсекаем строки где не заполнено ФИО

                        if 'Файл' not in temp_df.columns:
                            temp_df.insert(0, 'Файл', name_file)

                        if 'Группа' not in temp_df.columns:
                            temp_df.insert(0, 'Группа', name_sheet)  # вставляем колонку с именем листа

                        if checkbox_expelled == 0:
                            temp_df = temp_df[
                                temp_df['Статус_Учёба'] != 'Отчислен']  # отбрасываем отчисленных если поставлен чекбокс
                        elif checkbox_expelled == 1:
                            temp_df = temp_df[
                                temp_df['Статус_Учёба'] != 'Отчислен']  # отбрасываем отчисленных и академистов
                            temp_df = temp_df[~temp_df['Статус_Учёба'].isin(['Академический отпуск(декрет)','Академический отпуск(служба в РА)','Академический отпуск(по болезни)','Академический отпуск(ученич. договор)'])]

                        else:
                            temp_df = temp_df

                        main_df = pd.concat([main_df, temp_df], axis=0, ignore_index=True)  # добавляем в общий файл
                        if len(temp_df) == 0:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}', f'{name_sheet}', f'',
                                       'Пустой файл']],
                                columns=['Название файла', 'Название листа', 'Значение ошибки',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0, ignore_index=True)
                        else:
                            quantity_sheets += 1

                        # Создаем свод по файлам
                        row_dct = {key: 0 for key in lst_generate_name_columns}  # создаем словарь для хранения данных
                        row_dct['Файл'] = name_file
                        row_dct['Лист'] = name_sheet  # добавляем колонки для листа
                        for name_column, dct_value_column in dct_params.items():
                            for key, value in dct_value_column.items():
                                row_dct[key] = temp_df[temp_df[name_column] == value].shape[0]
                        new_row = pd.DataFrame(row_dct, index=[0])
                        custom_report_df = pd.concat([custom_report_df, new_row], axis=0)




        # получаем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Проверяем есть ли данные в общем файле, если нет то вызываем исключение
        if len(main_df) == 0:
            error_df.to_excel(f'{path_end_folder}/Ошибки от {current_time}.xlsx', index=False)
            raise NotGoodSheet

        main_df.rename(columns={'Группа': 'Для переноса', 'Файл': 'файл для переноса'},
                       inplace=True)  # переименовываем группу чтобы перенести ее в начало таблицы
        main_df.insert(0, 'Файл', main_df['файл для переноса'])
        main_df.insert(1, 'Группа', main_df['Для переноса'])
        main_df.drop(columns=['Для переноса', 'файл для переноса'], inplace=True)
        main_df.fillna('Нет статуса', inplace=True)  # заполняем пустые ячейки

        # Сохраняем лист со всеми данными
        lst_date_columns = []  # список для колонок с датами
        for column in main_df.columns:
            if 'дата' in column.lower():
                lst_date_columns.append(column)
        main_df[lst_date_columns] = main_df[lst_date_columns].applymap(convert_to_date)  # Приводим к типу
        main_df[lst_date_columns] = main_df[lst_date_columns].applymap(
            lambda x: x.strftime('%d.%m.%Y') if isinstance(x, (pd.Timestamp,datetime.datetime)) and pd.notna(x) else x)

        main_df.replace('Нет статуса', '', inplace=True)
        # Добавляем 2 колонки с вариантами СНИЛС
        # классический вариант
        main_df['СНИЛС_дефис'] = main_df.apply(lambda row: convert_snils_dash(row['СНИЛС']) if row['ФИО'] != 'Группа' else '',
                                     axis=1)
        # просто цифры
        main_df['СНИЛС_без_дефиса'] = main_df.apply(
            lambda row: convert_snils_not_dash(row['СНИЛС']) if row['ФИО'] != 'Группа' else '', axis=1)
        # Добавляем раздельные колонки для кода ОП и названия ОП
        main_df['Код_ОП'] = main_df['Статус_ОП'].apply(lambda x:extract_part_status_op(x,'Код_ОП'))
        main_df['Наименование_ОП'] = main_df['Статус_ОП'].apply(lambda x:extract_part_status_op(x,'Наименование_ОП'))
        # Добавляем разбиение по датам
        main_df = proccessing_date(raw_date, 'Дата_рождения', main_df,path_end_folder)

        # Добавляем склонение по падежам и создание инициалов
        main_df = declension_fio_by_case(main_df)

        # Обрабатываем колонки типа Подсчет если они есть
        lst_counting_name_columns = [name_column for name_column in main_df.columns if 'Подсчет_' in name_column]
        if len(lst_counting_name_columns) != 0:
            # Создаем файл в котором будут сводные данные по колонкам с Подсчетом
            dct_counting_save_name = {'Файл': 'по группам', 'Текущий_возраст': 'по возрастам', 'Год_рождения': 'по годам рождения', 'Статус_ОП': 'по ОП',
                                      'Пол': 'по полам'}  # словарь для названий колонок по которым будут создаваться файлы
            # Создаем папку для хранения сводов по колонкам подсчета
            path_counting_file = f'{path_end_folder}/Своды по колонкам Подсчетов'  #
            if not os.path.exists(path_counting_file):
                os.makedirs(path_counting_file)
            # приводим к числовому формату
            for name_counting_column in lst_counting_name_columns:
                main_df[name_counting_column] = main_df[name_counting_column].apply(convert_number)
            # создаем своды
            for name_column, name_file in dct_counting_save_name.items():
                create_svod_counting(main_df.copy(), name_column, name_file, lst_counting_name_columns,
                                     path_counting_file, current_time)

        # Создаем списки на основе которых мы создаем настраиваемый отчет
        lst_custom_wb = create_for_custom_report(main_df, params_df)
        lst_custom_wb.save(f'{path_end_folder}/Списки для свода по выбранным колонкам от {current_time}.xlsx')

        # Генерируем файлы егиссо
        # генерируем полный вариант
        if path_egisso_params != '' and path_egisso_params != 'Не выбрано':
            df_params_egisso, temp_params_egisso_error_df = extract_parameters_egisso(path_egisso_params,
                                                                                      list(main_df.columns))
            path_egisso_file = f'{path_end_folder}/ЕГИССО' # создаем папку для хранения файлов егиссо
            if not os.path.exists(path_egisso_file):
                os.makedirs(path_egisso_file)
            if len(df_params_egisso) != 0:
                egisso_full_wb, egisso_not_find_wb, egisso_error_wb = create_full_egisso_data(main_df, df_params_egisso,
                                                                                              path_egisso_file)  # создаем полный набор данных
                egisso_full_wb.save(f'{path_egisso_file}/ЕГИССО полные данные от {current_time}.xlsx')
                egisso_not_find_wb.save(f'{path_egisso_file}/ЕГИССО Не найденные льготы {current_time}.xlsx')
                egisso_error_wb.save(f'{path_egisso_file}/ЕГИССО перс данные ОШИБКИ от {current_time}.xlsx')


            else:
                # генерируем вариант только с персональными данными
                egisso_clean_wb, egisso_error_wb = create_part_egisso_data(main_df)
                egisso_clean_wb.save(f'{path_egisso_file}/ЕГИССО перс данные от {current_time}.xlsx')
                egisso_error_wb.save(f'{path_egisso_file}/ЕГИССО перс данные ОШИБКИ от {current_time}.xlsx')
        else:
            temp_params_egisso_error_df = pd.DataFrame(
            columns=['Название файла', 'Название листа', 'Значение ошибки', 'Описание ошибки']) # создаем пустой файл

        # Сохраняем лист с ошибками

        error_df = pd.concat([error_df, temp_params_egisso_error_df], axis=0, ignore_index=True)
        error_wb = write_df_to_excel({'Ошибки':error_df},write_index=False)
        error_wb.save(f'{path_end_folder}/Ошибки в файле от {current_time}.xlsx')

        # суммируем данные по листам

        all_custom_report_df = custom_report_df.sum(axis=0)
        all_custom_report_df = all_custom_report_df.drop(['Файл', 'Лист']).to_frame()  # удаляем текстовую строку

        all_custom_report_df = all_custom_report_df.reset_index()
        all_custom_report_df.columns = ['Наименование параметра', 'Количество']

        # Добавляем в разрез по файлам результирующую строку
        custom_report_df.loc['Итого'] = custom_report_df.sum(axis=0)  # создаем результирующую строку
        custom_report_df.iloc[-1,0] = 'Итого'
        custom_report_df.iloc[-1,1] = ''

        # отчет в разрезе возрастов
        custom_report_age_df = create_slice_report(main_df.copy(),'Текущий_возраст',dct_params)
        # отчет в разрезе годов рождения
        custom_report_year_df = create_slice_report(main_df.copy(),'Год_рождения',dct_params)
        # отчет в разрезе полов
        custom_report_sex_df = create_slice_report(main_df.copy(),'Пол',dct_params)

        # отчет в разрезе статусов ОП
        custom_report_op_df = create_slice_report(main_df.copy(),'Статус_ОП',dct_params)

        # сохраняем файл с данными по выбранным колонкам

        custom_report_wb = write_df_to_excel({'Общий свод': all_custom_report_df, 'Свод по файлам': custom_report_df,
                                              'Свод по возрастам':custom_report_age_df,'Свод по годам рождения':custom_report_year_df,'Свод по полам':custom_report_sex_df,'Свод по ОП':custom_report_op_df,
                                              },
                                             write_index=False)
        custom_report_wb = del_sheet(custom_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        custom_report_wb.save(f'{path_end_folder}/Свод по выбранным колонкам Статусов от {current_time}.xlsx')

        main_df.replace('Нет статуса', '', inplace=True)


        main_df.columns = list(map(str, list(main_df.columns)))


        # Обрабатываем колонки типа Список
        lst_list_name_columns = [name_column for name_column in main_df.columns if 'Список_' in name_column]
        if len(lst_list_name_columns) != 0:
            dct_list_columns = {}  # словарь в котором будут храниться датафреймы созданные для каждой колонки со списокм
            dct_values_in_list_columns = {}  # словарь в котором будут храниться названия колонок и все значения которые там встречались
            dct_df_list_in_columns = {}  # словарь где будут храниться значения в колонках и датафреймы где в указанных колонках есть соответствующее значение

            dct_list_save_name = {'Файл': 'по группам', 'Текущий_возраст': 'по возрастам','Год_рождения': 'по годам рождения', 'Статус_ОП': 'по ОП',
                                  'Пол': 'по полам'}  # словарь для названий колонок по которым будут создаваться срезы
            # Создаем папку для хранения сводов по колонкам списков
            path_list_file = f'{path_end_folder}/Своды по колонкам Списков'  #
            if not os.path.exists(path_list_file):
                os.makedirs(path_list_file)

            main_df[lst_list_name_columns] = main_df[lst_list_name_columns].astype(str)
            for name_lst_column in lst_list_name_columns:
                temp_col_value_lst = main_df[name_lst_column].tolist()  # создаем список
                temp_col_value_lst = [value for value in temp_col_value_lst if value]  # отбрасываем пустые значения
                unwrap_lst = []
                temp_col_value_lst = list(map(str, temp_col_value_lst))  # делаем строковым каждый элемент
                for value in temp_col_value_lst:
                    unwrap_lst.extend(value.split(','))
                unwrap_lst = list(map(str.strip, unwrap_lst))  # получаем список
                # убираем повторения и сортируем
                dct_values_in_list_columns[name_lst_column] = sorted(list(set(unwrap_lst)))

                dct_value_list = dict(Counter(unwrap_lst))  # Превращаем в словарь
                sorted_dct_value_lst = dict(sorted(dct_value_list.items()))  # сортируем словарь
                # создаем датафрейм
                temp_df = pd.DataFrame(list(sorted_dct_value_lst.items()), columns=['Показатель', 'Значение'])
                dct_list_columns[name_lst_column] = temp_df

            # Создаем датафреймы для подтверждения цифр
            for key, lst in dct_values_in_list_columns.items():
                for value in lst:
                    temp_list_df = main_df[main_df[key].str.contains(value)]
                    name_sheet = key.replace('Список_', '')

                    dct_df_list_in_columns[f'{name_sheet}_{value}'] = temp_list_df

                # Сохраняем
            if len(dct_df_list_in_columns) >= 252:
                lst_dct = list(dct_df_list_in_columns.items()) # превращаем в список кортежей
                start_threshold = 0
                end_threshold = 252
                count = 1
                while end_threshold <= len(lst_dct) + 1:
                    big_dct = dict(lst_dct[start_threshold:end_threshold])
                    list_columns_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                    list_columns_report_wb = del_sheet(list_columns_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                    list_columns_report_wb.save(f'{path_list_file}/Данные по сводам Списков {count}.xlsx')
                    count += 1
                    start_threshold += 252
                    end_threshold += 252
                    # контролируем диапазон
                    if end_threshold > len(lst_dct):
                        big_dct = dict(lst_dct[start_threshold:len(lst_dct) + 1])
                        list_columns_report_wb = write_df_big_dct_to_excel(big_dct, write_index=False)
                        list_columns_report_wb = del_sheet(list_columns_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                        list_columns_report_wb.save(f'{path_list_file}/Данные по сводам Списков {count}.xlsx')

            else:
                list_columns_report_wb = write_df_big_dct_to_excel(dct_df_list_in_columns, write_index=False)
                list_columns_report_wb = del_sheet(list_columns_report_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
                list_columns_report_wb.save(f'{path_list_file}/Данные по сводам Списков {current_time}.xlsx')

            # создаем срезы
            for name_column, prefix_file in dct_list_save_name.items():
                create_svod_list(main_df.copy(), name_column, prefix_file, lst_list_name_columns,
                                 path_list_file, current_time, dct_values_in_list_columns)




        main_wb = write_df_to_excel({'Общий список': main_df}, write_index=False)
        main_wb = del_sheet(main_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        main_wb.save(f'{path_end_folder}/Общий файл от {current_time}.xlsx')

        # Ищем ошибки в персональных данных
        main_df.replace('', 'Нет статуса', inplace=True)
        check_error_in_pers_data(main_df.copy(), path_end_folder, current_time)

        # Заменяем название колонки Пол на Статус_Пол чтобы обработка проходила нормально
        main_df.rename(columns={'Пол':'Статус_Пол'},inplace=True)

        # Создаем папку для хранения сводов по статусам
        path_svod_file = f'{path_end_folder}/Своды по колонкам Статусов'  #
        if not os.path.exists(path_svod_file):
            os.makedirs(path_svod_file)

            # Создаем раскладку по колонкам статусов
            lst_status_columns = [column for column in main_df.columns if 'Статус_' in column]
            dct_status_save_name = {'Файл': 'по группам', 'Текущий_возраст': 'по возрастам','Год_рождения': 'по годам рождения', 'Статус_ОП': 'по ОП',
                                    'Статус_Пол': 'по полам'}  # словарь для названий колонок по которым будут создаваться срезы

            for name_column, prefix_file in dct_status_save_name.items():
                create_svod_status(main_df.copy(), name_column, prefix_file, lst_status_columns,
                                   path_svod_file, current_time)

        # Создаем раскладку по колонкам статусов
        lst_status_columns = [column for column in main_df.columns if 'Статус_' in column]
        dct_status = {}  # словарь для хранения сводных датафреймов
        for name_column in lst_status_columns:
            svod_df = pd.pivot_table(main_df, index='Файл', columns=name_column, values='ФИО',
                                     aggfunc='count', fill_value=0, margins=True,
              margins_name='Итого').reset_index()
            name_sheet = name_column.replace('Статус_', '')
            dct_status[name_sheet] = svod_df  # сохраняем в словарь сводную таблицу

        # Создаем Свод по статусам
        # Собираем колонки содержащие слово Статус_ и Подсчет_ и Список
        lst_status = [name_column for name_column in main_df.columns if
                      'Статус_' in name_column or 'Подсчет_' in name_column or 'Список_' in name_column]

        # Создаем датафрейм с данными по статусам
        soc_df = pd.DataFrame(columns=['Показатель', 'Значение'])  # датафрейм для сбора данных отчета
        soc_df.loc[len(soc_df)] = ['Количество учебных групп', quantity_sheets]  # добавляем количество учебных групп
        # считаем количество студентов
        quantity_study_student = main_df[main_df['Статус_Учёба'] == 'Обучается'].shape[0]  # со статусом Обучается
        quantity_academ_student = main_df[main_df['Статус_Учёба'].str.contains('Академический отпуск')].shape[
            0]
        quantity_not_status_student = main_df[main_df['Статус_Учёба'].str.contains('Нет статуса')].shape[
            0]
        quantity_except_deducted = main_df[~main_df['Статус_Учёба'].str.contains('Отчислен')].shape[
            0]  # все студенты кроме отчисленных
        if checkbox_expelled == 0:
            soc_df.loc[len(soc_df)] = ['Количество студентов (контингент)',
                                       f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student}, Не указан статус - {quantity_not_status_student}, Всего - {quantity_except_deducted} ']  # добавляем количество студентов
        elif checkbox_expelled == 1:
            # Без отчисленных и академа
            soc_df.loc[len(soc_df)] = ['Количество студентов (контингент)',
                                       f'Обучается - {quantity_study_student}, Не указан статус - {quantity_not_status_student}, Всего {quantity_except_deducted} (без студентов в академическом отпуске и отчисленных)']  # добавляем количество студентов
        else:
            # Включая отчисленных и в академе
            quantity_except = main_df[main_df['Статус_Учёба'].str.contains('Отчислен')].shape[
                0]  # количество отчисленных
            soc_df.loc[len(soc_df)] = ['Количество студентов (контингент)',
                                       f'Обучается - {quantity_study_student}, Академ - {quantity_academ_student}, Не указан статус - {quantity_not_status_student},Отчислено - {quantity_except}, Всего -  {len(main_df)}']  # добавляем количество студентов

        # считаем количество совершенолетних студентов
        quantity_maturity_students = len(main_df[main_df['Совершеннолетие'] == 'совершеннолетний'])
        quantity_not_maturity_students = len(main_df[main_df['Совершеннолетие'] == 'несовершеннолетний'])
        quantity_error_maturity_students = len(
            main_df[main_df['Совершеннолетие'].isin(['отрицательный возраст', 'Ошибочное значение!!!'])])
        soc_df.loc[len(soc_df)] = ['Возраст',
                                   f'Совершеннолетних - {quantity_maturity_students}, Несовершеннолетних - {quantity_not_maturity_students}, Неправильная дата рождения - {quantity_error_maturity_students}, Всего {len(main_df)}']
        # Распределение по СПО-1
        header_spo = pd.DataFrame(columns=['Показатель', 'Значение'],
                                   data=[['Статус_СПО-1', None]])  # создаем строку с заголовком
        df_svod_by_SPO1 = main_df.groupby(['СПО_Один']).agg({'ФИО': 'count'})
        df_svod_by_SPO1 = df_svod_by_SPO1.reset_index()
        df_svod_by_SPO1.columns = ['Показатель', 'Значение']
        soc_df = pd.concat([soc_df, header_spo], axis=0)
        soc_df = pd.concat([soc_df, df_svod_by_SPO1], axis=0)
        for name_column in lst_status:
            if name_column == 'Статус_ОП':
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                # создаем строки с описанием
                new_value_df = create_value_str(main_df, name_column, 'Статус_Учёба',
                                                {'Обучается': 'Обучается', 'Академ': 'Академический отпуск',
                                                 'Не указан статус': 'Нет статуса'},checkbox_expelled)
            elif 'Статус_' in name_column:
                temp_counts = main_df[name_column].value_counts()  # делаем подсчет
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                new_value_df = temp_counts.to_frame().reset_index()  # создаем датафрейм с данными
                new_value_df.columns = ['Показатель', 'Значение']  # делаем одинаковыми названия колонок
                new_value_df['Показатель'] = new_value_df['Показатель'].astype(str)
                new_value_df.sort_values(by='Показатель', inplace=True)
            elif 'Подсчет' in name_column:
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                main_df[name_column] = main_df[name_column].apply(convert_number)
                temp_desccribe = main_df[name_column].describe()
                sum_column = main_df[name_column].sum()
                _dct_describe = temp_desccribe.to_dict()
                dct_describe = {'Среднее': round(_dct_describe['mean'], 2), 'Сумма': round(sum_column, 2),
                                'Медиана': _dct_describe['50%'],
                                'Минимум': _dct_describe['min'], 'Максимум': _dct_describe['max'],
                                'Количество': _dct_describe['count'], }
                new_value_df = pd.DataFrame(list(dct_describe.items()), columns=['Показатель', 'Значение'])
            elif 'Список_' in name_column:
                new_part_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[[name_column, None]])  # создаем строку с заголовком
                new_value_df = dct_list_columns[name_column]
            new_part_df = pd.concat([new_part_df, new_value_df], axis=0)  # соединяем
            soc_df = pd.concat([soc_df, new_part_df], axis=0)
        # Создаем раскладку по группам
        new_group_header_df = pd.DataFrame(columns=['Показатель', 'Значение'],
                                           data=[['Статус_студентов по группам', None]])  # создаем строку с заголовком
        new_group_df = create_value_str(main_df, 'Группа', 'Статус_Учёба',
                                        {'Обучается': 'Обучается', 'Академ': 'Академический отпуск',
                                         'Не указан статус': 'Нет статуса'},checkbox_expelled)
        new_group_header_df = pd.concat([new_group_header_df, new_group_df], axis=0)

        soc_df = pd.concat([soc_df, new_group_header_df], axis=0)

        soc_wb = write_df_to_excel({'Свод по статусам': soc_df}, write_index=False)
        soc_wb = del_sheet(soc_wb, ['Sheet', 'Sheet1', 'Для подсчета'])

        column_number = 0  # номер колонки в которой ищем слово Статус_
        # Создаем  стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in soc_wb['Свод по статусам'].iter_rows(min_row=1, max_row=soc_wb['Свод по статусам'].max_row,
                                                        min_col=column_number,
                                                        max_col=column_number):  # Перебираем строки
            if 'Статус_' in str(row[column_number].value) or 'Подсчет_' in str(
                    row[column_number].value) or 'Список_' in str(
                    row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill

        soc_wb.save(f'{path_end_folder}/Сводка от {current_time}.xlsx')

        # Создаем файл excel в котороым будет находится отчет
        wb = openpyxl.Workbook()

        # Проверяем наличие возможных дубликатов ,котороые могут получиться если обрезать по 30 символов
        lst_length_column = [column[:30] for column in main_df.columns]
        check_dupl_length = [k for k, v in Counter(lst_length_column).items() if v > 1]

        # проверяем наличие объединенных ячеек
        check_merge = [column for column in main_df.columns if 'Unnamed' in column]
        # если есть хоть один Unnamed то просто заменяем названия колонок на Колонка №цифра
        if check_merge or check_dupl_length:
            main_df.columns = [f'Колонка №{i}' for i in range(1, main_df.shape[1] + 1)]
        # очищаем названия колонок от символов */\ []''
        # Создаем регулярное выражение
        pattern_symbols = re.compile(r"[/*'\[\]/\\]")
        clean_main_df_columns = [re.sub(pattern_symbols, '', column) for column in main_df.columns]
        main_df.columns = clean_main_df_columns

        # Добавляем столбец для облегчения подсчета по категориям
        main_df['Для подсчета'] = 1

        # Создаем листы
        for idx, name_column in enumerate(main_df.columns):
            # Делаем короткое название не более 30 символов
            wb.create_sheet(title=name_column[:30], index=idx)

        for idx, name_column in enumerate(main_df.columns):
            group_main_df = main_df.astype({name_column: str}).groupby([name_column]).agg({'Для подсчета': 'sum'})
            group_main_df.columns = ['Количество']

            # Сортируем по убыванию
            group_main_df.sort_values(by=['Количество'], inplace=True, ascending=False)

            for r in dataframe_to_rows(group_main_df, index=True, header=True):
                if len(r) != 1:
                    wb[name_column[:30]].append(r)
            wb[name_column[:30]].column_dimensions['A'].width = 50

        # Удаляем листы
        wb = del_sheet(wb, ['Sheet', 'Sheet1', 'Для подсчета'])
        # Сохраняем итоговый файл
        wb.save(f'{path_end_folder}/Свод по каждой колонке таблицы от {current_time}.xlsx')

        if error_df.shape[0] != 0:
            messagebox.showwarning('Деметра Отчеты социальный паспорт студента',
                                   f'Обнаружены ошибки в файлах с данными.\n'
                                   f'Проверьте файл Ошибки в файле')

    except FileNotFoundError:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    except NotColumn:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Проверьте названия колонок в первом листе эталонного файла, для работы программы\n'
                             f' требуются колонки: {";".join(always_cols)}'
                             )
    except NotCustomColumn:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'В эталонном файле отсутствуют колонки указанные в файле с параметрами отчета,\n'
                             f' не найдены колонки: {";".join(custom_always_cols)}'
                             )
    except NotGoodSheet:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Заголовки ни одного листа не соответствуют эталонному файлу,\n'
                             f'Откройте файл с ошибками и устраните проблему'
                             )
    except BadEtalonFile:
        messagebox.showerror('Деметра Отчеты социальный паспорт студента',
                             f'Эталонный файл поврежден, скачайте заново или пересохраните его с помощью Excel'
                             )
    else:
        messagebox.showinfo('Деметра Отчеты социальный паспорт студента', 'Данные успешно обработаны')


if __name__ == '__main__':
    main_etalon_file = 'data/Приложение 2. Эталон для заполнения.xlsx'
    main_data_folder = 'data/Данные'
    main_result_folder = 'data/Результат'
    main_params_file = 'data/Пример Параметры отчета.xlsx'
    main_egisso_params = 'data/Параметры ЕГИССО.xlsx'
    main_checkbox_expelled = 0
    main_raw_data = '05.09.2024'
    create_local_report(main_etalon_file, main_data_folder, main_result_folder, main_params_file,main_egisso_params,
                        main_checkbox_expelled, main_raw_data)
    print('Lindy Booth')
